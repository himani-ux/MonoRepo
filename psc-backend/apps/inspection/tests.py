"""
Tests for inspection and deficiency feature backfill.

PRD Reference: Docs/PRD.md Sections 2.1, 2.2
Validation Reference: Docs/VALIDATION_RULES.md Sections 2.1, 2.2, 3.1, 3.2
RBAC Reference: Docs/BACKEND_STRUCTURE.md Section 11
"""

import shutil
import tempfile
import uuid
from datetime import date, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from django.test import TestCase, override_settings
from django.db.models import Value
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIRequestFactory, force_authenticate

from apps.accounts.models import RoleCodes
from apps.car.models import ActivityHistory, AuditLog, Evidence
from apps.inspection.deficiency_models import CAR, Deficiency, DeficiencyActionHistory
from apps.inspection.models import Inspection, InspectionReport, InspectionStatus
from apps.inspection.views import (
    InspectionListView,
    InspectionCreateView,
    InspectionDetailView,
    InspectionUpdateView,
    InspectionDeleteView,
    InspectionUploadReportView,
    InspectionSubmitView,
    InspectionPICReviewView,
    InspectionDPACloseView,
)
from apps.inspection.deficiency_views import DeficiencyCreateView
from apps.inspection.deficiency_views import DeficiencyActionCodeUpdateView
from apps.inspection.followup_views import FollowUpView
from apps.inspection.report_views import BulkCARExportView, DeficiencyExcelExportView
from apps.inspection.reports import generate_deficiency_excel
from apps.masters.models import PSCActionCode, PSCDefCode
from apps.notifications.models import Notification
from django.core.files.uploadedfile import SimpleUploadedFile


def make_user(
    *,
    role: str,
    user_type: str,
    vessel_id: str | None = None,
    user_id: str = "test-user",
    display_name: str = "Test User",
):
    return SimpleNamespace(
        id=user_id,
        role=role,
        user_type=user_type,
        vessel_id=vessel_id,
        display_name=display_name,
        username=display_name.lower().replace(" ", "_"),
        is_authenticated=True,
    )


class BaseInspectionAPITestCase(TestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.vessel_id = uuid.uuid4()
        self.other_vessel_id = uuid.uuid4()

        self.vessel_master = make_user(
            role=RoleCodes.VESSEL_MASTER,
            user_type="VESSEL",
            vessel_id=str(self.vessel_id),
            user_id="vm-1",
            display_name="Vessel Master",
        )
        self.vessel_crew = make_user(
            role=RoleCodes.VESSEL_CREW,
            user_type="VESSEL",
            vessel_id=str(self.vessel_id),
            user_id="vc-1",
            display_name="Vessel Crew",
        )
        self.office_pic = make_user(
            role=RoleCodes.OFFICE_PIC,
            user_type="OFFICE",
            vessel_id=None,
            user_id="office-1",
            display_name="Office PIC",
        )
        self.dpa = make_user(
            role=RoleCodes.DPA,
            user_type="OFFICE",
            vessel_id=None,
            user_id="dpa-1",
            display_name="DPA User",
        )

    def create_inspection(self, **overrides):
        payload = {
            "vessel_id": self.vessel_id,
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": date.today(),
            "port_place": "Singapore",
            "country": "Singapore",
            "mou_id": "TOKYO",
            "authority": "PSC Authority",
            "inspector_name": "Inspector",
            "report_reference": "REF-001",
            "is_detention": False,
        }
        payload.update(overrides)
        return Inspection.objects.create(**payload)


class TestFEAT_INS_001_CreateInspection(BaseInspectionAPITestCase):
    """
    FEAT-INS-001: Create Inspection.
    """

    def test_happy_path_vessel_master_can_create_psc_inspection(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today()),
            "port_place": "Rotterdam",
            "country": "Netherlands",
            "mou_id": "PARIS",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.vessel_master)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        inspection = Inspection.objects.get(id=response.data["data"]["id"])
        self.assertEqual(inspection.status, InspectionStatus.DRAFT)
        self.assertEqual(inspection.inspection_type, "PSC")
        self.assertEqual(inspection.psc_subtype, "INITIAL")

    def test_happy_path_non_psc_clears_psc_subtype(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "RS",
            "psc_subtype": None,
            "inspection_date": str(date.today()),
            "port_place": "Busan",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.vessel_master)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        inspection = Inspection.objects.get(id=response.data["data"]["id"])
        self.assertEqual(inspection.psc_subtype, None)
        self.assertEqual(inspection.status, InspectionStatus.DRAFT)


    def test_contract_create_supported_on_root_inspections_endpoint(self):
        view = InspectionListView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today()),
            "port_place": "Singapore",
            "country": "Singapore",
            "mou_id": "TOKYO",
        }

        request = self.factory.post("/api/psc/inspections/", payload, format="json")
        force_authenticate(request, user=self.vessel_master)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        inspection = Inspection.objects.get(id=response.data["data"]["id"])
        self.assertEqual(inspection.status, InspectionStatus.DRAFT)

    def test_validation_psc_subtype_required_for_psc(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "PSC",
            "inspection_date": str(date.today()),
            "port_place": "Mumbai",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.vessel_master)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("psc_subtype", response.data)

    def test_validation_non_psc_rejects_psc_subtype(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "RS",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today()),
            "port_place": "Chennai",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.vessel_master)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("psc_subtype", response.data)

    def test_rbac_vessel_crew_cannot_create_inspection(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today()),
            "port_place": "Shanghai",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.vessel_crew)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_office_pic_cannot_create_psc_inspection(self):
        """PSC creation is restricted to Vessel Master only."""
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today()),
            "port_place": "Hamburg",
            "mou_id": "PARIS",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.office_pic)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_office_pic_cannot_create_rightship_inspection(self):
        """RightShip creation is restricted to Vessel Master only."""
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "RS",
            "inspection_date": str(date.today()),
            "port_place": "Antwerp",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.office_pic)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_office_pic_can_create_audit_inspection(self):
        """Office users remain allowed to create non-PSC/RS inspections."""
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "AUDIT",
            "inspection_date": str(date.today()),
            "port_place": "Dubai",
            "def_reported": "NO",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.office_pic)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_rbac_vessel_user_cannot_create_for_other_vessel(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.other_vessel_id),
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today()),
            "port_place": "Singapore",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.vessel_master)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(response.data["error"], "FORBIDDEN")

    def test_rbac_unauthenticated_cannot_create_inspection(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today()),
            "port_place": "Doha",
        }
        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        response = view(request)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_gap_validation_future_inspection_date_should_be_rejected(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today() + timedelta(days=1)),
            "port_place": "Osaka",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.vessel_master)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("inspection_date", response.data)

    def test_gap_validation_mou_should_be_required_for_psc(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today()),
            "port_place": "Lisbon",
            "mou_id": "",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.vessel_master)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("mou_id", response.data)

    def test_gap_validation_port_place_min_length(self):
        view = InspectionCreateView.as_view()
        payload = {
            "vessel_id": str(self.vessel_id),
            "inspection_type": "PSC",
            "psc_subtype": "INITIAL",
            "inspection_date": str(date.today()),
            "port_place": "A",
            "mou_id": "PARIS",
        }

        request = self.factory.post("/api/psc/inspections/create/", payload, format="json")
        force_authenticate(request, user=self.vessel_master)
        response = view(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("port_place", response.data)


class TestFEAT_INS_002_UploadInspectionReport(BaseInspectionAPITestCase):
    """
    FEAT-INS-002: Upload Inspection Report.
    """

    def setUp(self):
        super().setUp()
        self.inspection = self.create_inspection(status=InspectionStatus.DRAFT)
        self.upload_dir = tempfile.mkdtemp(prefix="psc-upload-tests-")

    def tearDown(self):
        shutil.rmtree(self.upload_dir, ignore_errors=True)
        super().tearDown()

    @override_settings(PSC_UPLOAD_PATH=tempfile.gettempdir())
    def test_happy_path_upload_pdf(self):
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "inspection-report.pdf",
            b"%PDF-1.4 content",
            content_type="application/pdf",
        )

        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/upload-report/",
            {"file": upload, "description": "Initial PSC report"},
            format="multipart",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(InspectionReport.objects.filter(inspection=self.inspection).count(), 1)

    @override_settings(PSC_UPLOAD_PATH=tempfile.gettempdir())
    def test_happy_path_upload_jpeg(self):
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "inspection.jpg",
            b"jpeg-content",
            content_type="image/jpeg",
        )
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/upload-report/",
            {"file": upload, "description": "Photo report"},
            format="multipart",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    @override_settings(PSC_UPLOAD_PATH=tempfile.gettempdir())
    def test_happy_path_upload_jpg_alias(self):
        """PRD FEAT-INS-002: accepted formats include JPG alias."""
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "inspection.jpg",
            b"jpg-content",
            content_type="image/jpg",
        )
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/upload-report/",
            {"file": upload, "description": "JPG alias report"},
            format="multipart",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_validation_rejects_unsupported_file_type(self):
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "inspection.png",
            b"png-content",
            content_type="image/png",
        )
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/upload-report/",
            {"file": upload, "description": "PNG should fail"},
            format="multipart",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("file", response.data)

    def test_validation_rejects_file_larger_than_3mb(self):
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "large-report.pdf",
            b"x" * (3 * 1024 * 1024 + 1),
            content_type="application/pdf",
        )
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/upload-report/",
            {"file": upload, "description": "Large file"},
            format="multipart",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("file", response.data)

    def test_validation_description_max_length_500(self):
        """VALIDATION_RULES 8.1: description max 500 characters."""
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "inspection-report.pdf",
            b"%PDF-1.4 content",
            content_type="application/pdf",
        )
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/upload-report/",
            {"file": upload, "description": "x" * 501},
            format="multipart",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("description", response.data)

    def test_rbac_forbidden_for_other_vessel_user(self):
        other_vessel_master = make_user(
            role=RoleCodes.VESSEL_MASTER,
            user_type="VESSEL",
            vessel_id=str(self.other_vessel_id),
            user_id="vm-2",
            display_name="Other Vessel Master",
        )
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "inspection-report.pdf",
            b"%PDF-1.4 content",
            content_type="application/pdf",
        )
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/upload-report/",
            {"file": upload, "description": "Should fail access"},
            format="multipart",
        )
        force_authenticate(request, user=other_vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_unauthenticated_cannot_upload_report(self):
        """RBAC: unauthenticated report upload must be rejected."""
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "inspection-report.pdf",
            b"%PDF-1.4 content",
            content_type="application/pdf",
        )
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/upload-report/",
            {"file": upload, "description": "No auth"},
            format="multipart",
        )
        response = view(request, id=self.inspection.id)

        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_error_returns_not_found_for_missing_inspection(self):
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "inspection-report.pdf",
            b"%PDF-1.4 content",
            content_type="application/pdf",
        )
        request = self.factory.post(
            "/api/psc/inspections/00000000-0000-0000-0000-000000000000/upload-report/",
            {"file": upload, "description": "Missing inspection"},
            format="multipart",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=uuid.UUID("00000000-0000-0000-0000-000000000000"))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    @override_settings(PSC_UPLOAD_PATH=tempfile.gettempdir())
    def test_gap_validation_description_should_be_mandatory(self):
        view = InspectionUploadReportView.as_view()
        upload = SimpleUploadedFile(
            "inspection-report.pdf",
            b"%PDF-1.4 content",
            content_type="application/pdf",
        )
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/upload-report/",
            {"file": upload},
            format="multipart",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("description", response.data)


class TestFEAT_INS_003_AddDeficiency(BaseInspectionAPITestCase):
    """
    FEAT-INS-003: Add Deficiency to Inspection.
    """

    def setUp(self):
        super().setUp()
        self.inspection = self.create_inspection(status=InspectionStatus.DRAFT)

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_happy_path_add_deficiency_auto_creates_car(self, mock_def_get, mock_action_get):
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")
        mock_action_get.return_value = SimpleNamespace(action_code=30, definition="Rectify before departure")

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "Initial deficiency description for testing",
            "action_code_id": 30,
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        deficiency = Deficiency.objects.get(id=response.data["data"]["id"])
        self.assertEqual(deficiency.def_code, "10101")
        self.assertIsNotNone(deficiency.car)
        self.assertTrue(CAR.objects.filter(id=deficiency.car_id, status="DRAFT").exists())

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_happy_path_sequence_number_increments(self, mock_def_get):
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")

        view = DeficiencyCreateView.as_view()
        first_payload = {"def_code_id": "10101", "description": "First deficiency description"}
        second_payload = {"def_code_id": "10101", "description": "Second deficiency description"}

        first_request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            first_payload,
            format="json",
        )
        force_authenticate(first_request, user=self.vessel_master)
        first_response = view(first_request, inspection_id=self.inspection.id)
        self.assertEqual(first_response.status_code, status.HTTP_201_CREATED)

        second_request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            second_payload,
            format="json",
        )
        force_authenticate(second_request, user=self.vessel_master)
        second_response = view(second_request, inspection_id=self.inspection.id)
        self.assertEqual(second_response.status_code, status.HTTP_201_CREATED)

        first = Deficiency.objects.get(id=first_response.data["data"]["id"])
        second = Deficiency.objects.get(id=second_response.data["data"]["id"])
        self.assertEqual(first.sequence_no, 1)
        self.assertEqual(second.sequence_no, 2)

    @patch("apps.inspection.deficiency_serializers.CrewOnboardingHistory.objects.filter")
    @patch("apps.inspection.deficiency_serializers.HRM501.objects.filter")
    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_happy_path_assigned_crew_uuid_normalizes_to_crewid(
        self,
        mock_def_get,
        mock_hrm_filter,
        mock_onboarding_filter,
    ):
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")

        class _ExistsQS:
            def __init__(self, exists=False):
                self._exists = exists

            def exists(self):
                return self._exists

        class _ValuesListQS:
            def __init__(self, first_value=None):
                self._first_value = first_value

            def first(self):
                return self._first_value

        class _HRMQS:
            def __init__(self, exists=False, first_value=None):
                self._exists = exists
                self._first_value = first_value

            def exists(self):
                return self._exists

            def values_list(self, *args, **kwargs):
                return _ValuesListQS(self._first_value)

        def _hrm_side_effect(*args, **kwargs):
            if kwargs.get("CrewID") == "4bc48d31-9dc8-4de5-aa8b-975e5eb26f06":
                return _HRMQS(exists=False)
            if kwargs.get("id") == "4bc48d31-9dc8-4de5-aa8b-975e5eb26f06":
                return _HRMQS(first_value="KSM0171")
            return _HRMQS(exists=False)

        mock_hrm_filter.side_effect = _hrm_side_effect
        mock_onboarding_filter.return_value = _ExistsQS(exists=True)

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "Deficiency assignment should normalize UUID to CrewID",
            "assigned_crew_id": "4bc48d31-9dc8-4de5-aa8b-975e5eb26f06",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        deficiency = Deficiency.objects.get(id=response.data["data"]["id"])
        self.assertEqual(deficiency.assigned_crew_id, "KSM0171")

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_validation_invalid_def_code(self, mock_def_get):
        mock_def_get.side_effect = PSCDefCode.DoesNotExist()

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "99999",
            "description": "Deficiency with invalid code",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data["error"], "Validation failed")
        self.assertIn("def_code_id", response.data["details"])

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_validation_invalid_action_code(self, mock_action_get, mock_def_get):
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")
        mock_action_get.side_effect = PSCActionCode.DoesNotExist()

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "Deficiency with invalid action code",
            "action_code_id": 999,
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data["error"], "Validation failed")
        self.assertIn("action_code_id", response.data["details"])

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_precondition_status_must_be_draft_or_submitted(self, mock_def_get):
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")
        self.inspection.status = InspectionStatus.PIC_REVIEWED
        self.inspection.save(update_fields=["status"])

        view = DeficiencyCreateView.as_view()
        payload = {"def_code_id": "10101", "description": "Should fail by status"}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Status must be DRAFT or SUBMITTED", response.data["error"])

    def test_rbac_vessel_crew_cannot_add_deficiency(self):
        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "Crew should not add deficiency",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_crew)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_rbac_office_pic_can_add_deficiency(self, mock_def_get):
        """RBAC matrix: OFFICE_PIC can add deficiencies."""
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "Office PIC added deficiency on behalf of vessel",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn("def_code", response.data["data"])

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_rbac_dpa_can_add_deficiency(self, mock_def_get):
        """RBAC matrix: DPA can add deficiencies."""
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "DPA added deficiency for office correction",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn("def_code", response.data["data"])

    def test_rbac_unauthenticated_cannot_add_deficiency(self):
        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "No auth request",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        response = view(request, inspection_id=self.inspection.id)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_gap_validation_description_min_length(self, mock_def_get):
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "short",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("description", response.data["details"])

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_gap_validation_target_date_cannot_be_past(self, mock_def_get):
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "Deficiency with past target date",
            "target_date": str(date.today() - timedelta(days=1)),
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("target_date", response.data["details"])

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_gap_validation_description_max_length(self, mock_def_get):
        """VALIDATION_RULES 3.1: description max length should be 4000."""
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "x" * 4001,
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("description", response.data["details"])

    @patch("apps.inspection.deficiency_serializers.connection.cursor")
    @patch("apps.inspection.deficiency_serializers.HRM501.objects.filter")
    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_validation_assigned_crew_must_be_onboard_for_inspection_context(
        self,
        mock_def_get,
        mock_hrm_filter,
        mock_connection_cursor,
    ):
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")
        mock_hrm_qs = MagicMock()
        mock_hrm_qs.exists.return_value = True
        mock_hrm_filter.return_value = mock_hrm_qs

        cursor_cm = MagicMock()
        db_cursor = MagicMock()
        cursor_cm.__enter__.return_value = db_cursor
        db_cursor.fetchone.return_value = None
        mock_connection_cursor.return_value = cursor_cm

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "Crew assignment should fail if crew is not onboard vessel",
            "assigned_crew_id": "KSM0171",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("assigned_crew_id", response.data["details"])

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_gap_precondition_submitted_allows_only_office_user(self, mock_def_get):
        """
        VALIDATION_RULES 3.1 precondition:
        submitted inspections allow deficiency add only for OFFICE users.
        """
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")
        self.inspection.status = InspectionStatus.SUBMITTED
        self.inspection.save(update_fields=["status"])

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "Submitted status vessel-user addition should be blocked",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("apps.inspection.deficiency_serializers.PSCDefCode.objects.get")
    def test_gap_audit_activity_event_should_be_created(self, mock_def_get):
        mock_def_get.return_value = SimpleNamespace(def_code="10101", def_name="Certificate issue")
        before_count = ActivityHistory.objects.filter(
            entity_type__startswith="DEFICIENCY",
            entity_id__isnull=False,
        ).count()

        view = DeficiencyCreateView.as_view()
        payload = {
            "def_code_id": "10101",
            "description": "Deficiency should create activity event",
        }
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/deficiencies/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, inspection_id=self.inspection.id)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        after_count = ActivityHistory.objects.filter(
            entity_type__startswith="DEFICIENCY",
            entity_id__isnull=False,
        ).count()
        self.assertEqual(after_count, before_count + 1)


class TestFEAT_INS_004_SubmitInspection(BaseInspectionAPITestCase):
    """
    FEAT-INS-004: Submit Inspection.

    PRD Reference: Docs/PRD.md FEAT-INS-004
    Validation Reference: Docs/VALIDATION_RULES.md Section 2.2
    RBAC Reference: Docs/BACKEND_STRUCTURE.md Section 11
    """

    def setUp(self):
        super().setUp()
        self.inspection = self.create_inspection(status=InspectionStatus.DRAFT)

    def attach_report(self):
        InspectionReport.objects.create(
            inspection=self.inspection,
            file_name="inspection-report.pdf",
            file_path="/tmp/inspection-report.pdf",
            file_size=1024,
            mime_type="application/pdf",
            description="Initial inspection report",
            uploaded_by=str(self.vessel_master.id),
        )

    def test_happy_path_vessel_master_can_submit_draft_with_report(self):
        self.attach_report()

        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.inspection.refresh_from_db()
        self.assertEqual(self.inspection.status, InspectionStatus.SUBMITTED)
        self.assertEqual(response.data["data"]["status"], InspectionStatus.SUBMITTED)

    def test_happy_path_office_pic_can_submit_on_behalf(self):
        self.attach_report()

        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.inspection.refresh_from_db()
        self.assertEqual(self.inspection.status, InspectionStatus.SUBMITTED)

    def test_happy_path_dpa_can_submit_on_behalf(self):
        self.attach_report()

        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.inspection.refresh_from_db()
        self.assertEqual(self.inspection.status, InspectionStatus.SUBMITTED)

    def test_validation_requires_at_least_one_report(self):
        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("reports", response.data)

    def test_precondition_status_must_be_draft(self):
        self.inspection.status = InspectionStatus.SUBMITTED
        self.inspection.save(update_fields=["status"])
        self.attach_report()

        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn("Only draft inspections can be submitted", str(response.data))

    def test_gap_rbac_vessel_crew_should_not_submit(self):
        self.attach_report()

        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.vessel_crew)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_other_vessel_user_cannot_submit(self):
        self.attach_report()
        other_vessel_master = make_user(
            role=RoleCodes.VESSEL_MASTER,
            user_type="VESSEL",
            vessel_id=str(self.other_vessel_id),
            user_id="vm-2",
            display_name="Other Vessel Master",
        )

        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        force_authenticate(request, user=other_vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_unauthenticated_cannot_submit(self):
        self.attach_report()

        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        response = view(request, id=self.inspection.id)

        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_gap_validation_all_deficiencies_must_have_car_before_submit(self):
        self.attach_report()
        deficiency = Deficiency.objects.create(
            inspection=self.inspection,
            def_code_id="10101",
            def_code="10101",
            description="Deficiency without linked CAR should block submit",
            car=None,
        )
        # Force an orphan deficiency state to verify submit precondition.
        Deficiency.objects.filter(id=deficiency.id).update(car=None)

        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("deficiencies", str(response.data).lower())

    def test_gap_audit_activity_event_should_be_created_on_submit(self):
        self.attach_report()
        before_count = ActivityHistory.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
            event_type="INSPECTION_SUBMITTED",
        ).count()

        view = InspectionSubmitView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/submit/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        after_count = ActivityHistory.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
            event_type="INSPECTION_SUBMITTED",
        ).count()
        self.assertEqual(after_count, before_count + 1)


class TestFEAT_INS_005_PICReviewInspection(BaseInspectionAPITestCase):
    """
    FEAT-INS-005: PIC Review Inspection.

    PRD Reference: Docs/PRD.md FEAT-INS-005
    Validation Reference: Docs/VALIDATION_RULES.md Section 2.3
    RBAC Reference: Docs/BACKEND_STRUCTURE.md Section 11
    """

    def setUp(self):
        super().setUp()
        self.inspection = self.create_inspection(status=InspectionStatus.SUBMITTED)

    def test_happy_path_office_pic_can_review_submitted_inspection(self):
        view = InspectionPICReviewView.as_view()
        payload = {"comment": "Reviewed and acknowledged by PIC."}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.inspection.refresh_from_db()
        self.assertEqual(self.inspection.status, InspectionStatus.PIC_REVIEWED)
        self.assertEqual(self.inspection.pic_comment, payload["comment"])
        self.assertEqual(self.inspection.pic_reviewed_by, str(self.office_pic.id))

    def test_happy_path_office_ssqe_can_review(self):
        office_ssqe = make_user(
            role=RoleCodes.OFFICE_SSQE,
            user_type="OFFICE",
            vessel_id=None,
            user_id="office-ssqe-1",
            display_name="Office SSQE",
        )

        view = InspectionPICReviewView.as_view()
        payload = {"comment": "Reviewed by SSQE and accepted."}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=office_ssqe)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.inspection.refresh_from_db()
        self.assertEqual(self.inspection.status, InspectionStatus.PIC_REVIEWED)

    def test_happy_path_office_supt_can_review(self):
        office_supt = make_user(
            role=RoleCodes.OFFICE_SUPT,
            user_type="OFFICE",
            vessel_id=None,
            user_id="office-supt-1",
            display_name="Office Superintendent",
        )

        view = InspectionPICReviewView.as_view()
        payload = {"comment": "Reviewed by superintendent for closure readiness."}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=office_supt)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.inspection.refresh_from_db()
        self.assertEqual(self.inspection.status, InspectionStatus.PIC_REVIEWED)

    def test_validation_comment_is_required(self):
        view = InspectionPICReviewView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("comment", response.data)

    def test_validation_comment_minimum_length(self):
        view = InspectionPICReviewView.as_view()
        payload = {"comment": "too short"}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("comment", response.data)

    def test_validation_comment_minimum_boundary_accepted(self):
        """VALIDATION_RULES 2.3: 10-char PIC comment should pass."""
        view = InspectionPICReviewView.as_view()
        payload = {"comment": "1234567890"}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_precondition_status_must_be_submitted(self):
        self.inspection.status = InspectionStatus.DRAFT
        self.inspection.save(update_fields=["status"])

        view = InspectionPICReviewView.as_view()
        payload = {"comment": "This should fail because status is not submitted."}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn("Only submitted inspections can be reviewed", str(response.data))

    def test_rbac_vessel_master_cannot_review(self):
        view = InspectionPICReviewView.as_view()
        payload = {"comment": "Vessel user should not be able to review."}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_dpa_cannot_review(self):
        view = InspectionPICReviewView.as_view()
        payload = {"comment": "DPA cannot perform PIC review by RBAC."}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_unauthenticated_cannot_review(self):
        view = InspectionPICReviewView.as_view()
        payload = {"comment": "No auth should be rejected for PIC review."}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        response = view(request, id=self.inspection.id)

        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_error_returns_not_found_for_missing_inspection(self):
        view = InspectionPICReviewView.as_view()
        payload = {"comment": "Review missing inspection should return not found."}
        missing_id = uuid.UUID("00000000-0000-0000-0000-000000000000")
        request = self.factory.post(
            f"/api/psc/inspections/{missing_id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, id=missing_id)

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_gap_audit_activity_event_should_be_created_on_pic_review(self):
        before_count = ActivityHistory.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
            event_type="INSPECTION_PIC_REVIEWED",
        ).count()

        view = InspectionPICReviewView.as_view()
        payload = {"comment": "PIC review should create activity history event."}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/pic-review/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, id=self.inspection.id)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        after_count = ActivityHistory.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
            event_type="INSPECTION_PIC_REVIEWED",
        ).count()
        self.assertEqual(after_count, before_count + 1)


class TestFEAT_INS_006_DPACloseInspection(BaseInspectionAPITestCase):
    """
    FEAT-INS-006: DPA Close Inspection.

    PRD Reference: Docs/PRD.md FEAT-INS-006
    Validation Reference: Docs/VALIDATION_RULES.md Section 2.4
    RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.3, 11
    """

    def setUp(self):
        super().setUp()
        self.inspection = self.create_inspection(
            status=InspectionStatus.PIC_REVIEWED,
            pic_comment="PIC reviewed and ready for DPA close",
            pic_reviewed_by=str(self.office_pic.id),
        )

    def test_happy_path_dpa_can_close_pic_reviewed_inspection(self):
        """PRD FEAT-INS-006: Only DPA can close and status changes to DPA_CLOSED."""
        view = InspectionDPACloseView.as_view()
        payload = {"comment": "Closed by DPA after verification."}
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            payload,
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.inspection.refresh_from_db()
        self.assertEqual(self.inspection.status, InspectionStatus.DPA_CLOSED)
        self.assertEqual(self.inspection.dpa_comment, payload["comment"])
        self.assertEqual(self.inspection.dpa_closed_by, str(self.dpa.id))

    def test_validation_comment_is_required(self):
        """VALIDATION_RULES 2.4: DPA comment is mandatory."""
        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {},
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("comment", response.data)

    def test_validation_comment_minimum_length(self):
        """VALIDATION_RULES 2.4: DPA comment minimum length is 10."""
        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "too short"},
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("comment", response.data)

    def test_validation_comment_minimum_boundary_accepted(self):
        """VALIDATION_RULES 2.4: 10-char DPA comment should pass."""
        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "1234567890"},
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_precondition_status_must_be_pic_reviewed(self):
        """VALIDATION_RULES 2.4: Only PIC_REVIEWED inspections can be DPA closed."""
        self.inspection.status = InspectionStatus.SUBMITTED
        self.inspection.save(update_fields=["status"])

        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "Attempt close from wrong status."},
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn("Only PIC-reviewed inspections can be closed by DPA", str(response.data))

    def test_rbac_office_pic_cannot_close(self):
        """RBAC matrix: OFFICE_* cannot execute DPA close action."""
        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "PIC should not close inspection."},
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_office_ssqe_cannot_close(self):
        """RBAC matrix: OFFICE_SSQE cannot execute DPA close action."""
        office_ssqe = make_user(
            role=RoleCodes.OFFICE_SSQE,
            user_type="OFFICE",
            vessel_id=None,
            user_id="office-ssqe-2",
            display_name="Office SSQE",
        )
        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "SSQE should not close inspection."},
            format="json",
        )
        force_authenticate(request, user=office_ssqe)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_office_supt_cannot_close(self):
        """RBAC matrix: OFFICE_SUPT cannot execute DPA close action."""
        office_supt = make_user(
            role=RoleCodes.OFFICE_SUPT,
            user_type="OFFICE",
            vessel_id=None,
            user_id="office-supt-2",
            display_name="Office Superintendent",
        )
        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "Superintendent should not close inspection."},
            format="json",
        )
        force_authenticate(request, user=office_supt)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_vessel_master_cannot_close(self):
        """RBAC matrix: VESSEL_MASTER cannot execute DPA close action."""
        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "Vessel master should not close inspection."},
            format="json",
        )
        force_authenticate(request, user=self.vessel_master)
        response = view(request, id=self.inspection.id)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_unauthenticated_cannot_close(self):
        """RBAC: unauthenticated DPA-close request must be rejected."""
        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "No auth should fail."},
            format="json",
        )
        response = view(request, id=self.inspection.id)

        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_error_returns_not_found_for_missing_inspection(self):
        """BACKEND_STRUCTURE 10.3: missing inspection returns NOT_FOUND."""
        view = InspectionDPACloseView.as_view()
        missing_id = uuid.UUID("00000000-0000-0000-0000-000000000000")
        request = self.factory.post(
            f"/api/psc/inspections/{missing_id}/dpa-close/",
            {"comment": "Close missing inspection should fail."},
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=missing_id)

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_gap_audit_activity_event_should_be_created_on_dpa_close(self):
        """PRD FEAT-INS-006: DPA close must create activity event."""
        before_count = ActivityHistory.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
            event_type="INSPECTION_DPA_CLOSED",
        ).count()

        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "Close with activity logging expected."},
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=self.inspection.id)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        after_count = ActivityHistory.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
            event_type="INSPECTION_DPA_CLOSED",
        ).count()
        self.assertEqual(after_count, before_count + 1)

    def test_gap_notification_to_vessel_master_should_be_sent_on_dpa_close(self):
        """PRD FEAT-INS-006: DPA close must notify vessel master."""
        before_count = Notification.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
        ).count()

        view = InspectionDPACloseView.as_view()
        request = self.factory.post(
            f"/api/psc/inspections/{self.inspection.id}/dpa-close/",
            {"comment": "Close should trigger vessel notification."},
            format="json",
        )
        force_authenticate(request, user=self.dpa)
        response = view(request, id=self.inspection.id)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        after_count = Notification.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
        ).count()
        self.assertEqual(after_count, before_count + 1)


class TestFEAT_INS_007_EditInspectionDraft(BaseInspectionAPITestCase):
    """
    FEAT-INS-007: Edit Inspection (Draft).

    PRD Reference: Docs/PRD.md FEAT-INS-007
    API/RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.3, 11
    """

    def setUp(self):
        super().setUp()
        self.view = InspectionUpdateView.as_view()
        self.draft_inspection = self.create_inspection(
            status=InspectionStatus.DRAFT,
            inspection_type="PSC",
            psc_subtype="INITIAL",
            port_place="Singapore",
        )

    def _payload(self, **overrides):
        payload = {
            "inspection_type": "PSC",
            "psc_subtype": "CIC",
            "inspection_date": str(date.today()),
            "port_place": "Rotterdam",
            "country": "Netherlands",
            "mou_id": "PARIS",
            "authority": "Updated Authority",
            "inspector_name": "Updated Inspector",
            "report_reference": "REF-UPDATED-001",
            "is_detention": False,
        }
        payload.update(overrides)
        return payload

    def _update(self, *, user=None, inspection=None, payload=None):
        target = inspection or self.draft_inspection
        request = self.factory.put(
            f"/api/psc/inspections/{target.id}/update/",
            payload or self._payload(),
            format="json",
        )
        if user is not None:
            force_authenticate(request, user=user)
        return self.view(request, id=target.id)

    def test_feat_ins_007_happy_path_vessel_master_can_edit_draft(self):
        """PRD FEAT-INS-007: vessel master has full edit access for DRAFT inspection."""
        response = self._update(user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.draft_inspection.refresh_from_db()
        self.assertEqual(self.draft_inspection.port_place, "Rotterdam")
        self.assertEqual(self.draft_inspection.psc_subtype, "CIC")
        self.assertEqual(self.draft_inspection.updated_by, str(self.vessel_master.id))

    def test_feat_ins_007_happy_path_office_can_edit_assist_draft(self):
        """PRD FEAT-INS-007: office users can perform edit-assist on DRAFT inspection."""
        response = self._update(user=self.office_pic, payload=self._payload(port_place="Hamburg"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.draft_inspection.refresh_from_db()
        self.assertEqual(self.draft_inspection.port_place, "Hamburg")
        self.assertEqual(self.draft_inspection.updated_by, str(self.office_pic.id))

    def test_feat_ins_007_validation_psc_requires_psc_subtype_on_update(self):
        """VALIDATION_RULES 2.1: PSC subtype remains mandatory when inspection_type is PSC."""
        response = self._update(
            user=self.vessel_master,
            payload=self._payload(psc_subtype=None),
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("psc_subtype", response.data)

    def test_feat_ins_007_precondition_vessel_master_cannot_edit_submitted(self):
        """PRD FEAT-INS-007: DRAFT-only edit for vessel master; submitted should be blocked."""
        self.draft_inspection.status = InspectionStatus.SUBMITTED
        self.draft_inspection.save(update_fields=["status"])

        response = self._update(user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_feat_ins_007_rbac_vessel_crew_cannot_edit_draft(self):
        """RBAC matrix: VESSEL_CREW cannot edit inspections."""
        response = self._update(user=self.vessel_crew)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_feat_ins_007_rbac_other_vessel_master_cannot_edit(self):
        """RBAC/data visibility: vessel users cannot edit inspections outside own vessel."""
        other_vessel_master = make_user(
            role=RoleCodes.VESSEL_MASTER,
            user_type="VESSEL",
            vessel_id=str(self.other_vessel_id),
            user_id="vm-other",
            display_name="Other Vessel Master",
        )
        response = self._update(user=other_vessel_master)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_feat_ins_007_rbac_unauthenticated_edit_rejected(self):
        """RBAC: unauthenticated update request is rejected."""
        response = self._update(user=None)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_feat_ins_007_gap_office_edit_assist_should_create_audit_log(self):
        """PRD FEAT-INS-007: office edit-assist should be logged separately in audit trail."""
        before_count = AuditLog.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.draft_inspection.id,
            action="UPDATE",
        ).count()

        response = self._update(user=self.office_pic, payload=self._payload(port_place="Lisbon"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        after_count = AuditLog.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.draft_inspection.id,
            action="UPDATE",
            is_office_edit_assist=True,
        ).count()
        self.assertEqual(after_count, before_count + 1)


class TestFEAT_INS_008_EditInspectionPostSubmit(BaseInspectionAPITestCase):
    """
    FEAT-INS-008: Edit Inspection (Post-Submit).

    PRD Reference: Docs/PRD.md FEAT-INS-008
    API/RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.3, 11
    """

    def setUp(self):
        super().setUp()
        self.view = InspectionUpdateView.as_view()
        self.submitted_inspection = self.create_inspection(
            status=InspectionStatus.SUBMITTED,
            inspection_type="PSC",
            psc_subtype="INITIAL",
            port_place="Singapore",
            revision_no=1,
        )
        self.pic_reviewed_inspection = self.create_inspection(
            status=InspectionStatus.PIC_REVIEWED,
            inspection_type="PSC",
            psc_subtype="EXPANDED",
            port_place="Busan",
            revision_no=3,
        )

    def _payload(self, **overrides):
        payload = {
            "inspection_type": "PSC",
            "psc_subtype": "CIC",
            "inspection_date": str(date.today()),
            "port_place": "Antwerp",
            "country": "Belgium",
            "mou_id": "PARIS",
            "authority": "Office Updated Authority",
            "inspector_name": "Office Updated Inspector",
            "report_reference": "REF-POST-001",
            "is_detention": False,
        }
        payload.update(overrides)
        return payload

    def _update(self, *, user=None, inspection=None, payload=None):
        target = inspection or self.submitted_inspection
        request = self.factory.put(
            f"/api/psc/inspections/{target.id}/update/",
            payload or self._payload(),
            format="json",
        )
        if user is not None:
            force_authenticate(request, user=user)
        return self.view(request, id=target.id)

    def test_feat_ins_008_happy_path_office_can_edit_submitted(self):
        """PRD FEAT-INS-008: office roles can edit inspection in SUBMITTED status."""
        response = self._update(
            user=self.office_pic,
            inspection=self.submitted_inspection,
            payload=self._payload(port_place="Dubai"),
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.submitted_inspection.refresh_from_db()
        self.assertEqual(self.submitted_inspection.port_place, "Dubai")

    def test_feat_ins_008_happy_path_dpa_can_edit_pic_reviewed(self):
        """PRD FEAT-INS-008: DPA can edit inspection in PIC_REVIEWED status."""
        response = self._update(
            user=self.dpa,
            inspection=self.pic_reviewed_inspection,
            payload=self._payload(port_place="Qatar"),
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.pic_reviewed_inspection.refresh_from_db()
        self.assertEqual(self.pic_reviewed_inspection.port_place, "Qatar")

    def test_feat_ins_008_rbac_vessel_master_cannot_edit_submitted(self):
        """RBAC matrix: VESSEL_MASTER cannot edit post-submit inspections."""
        response = self._update(user=self.vessel_master, inspection=self.submitted_inspection)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_feat_ins_008_rbac_vessel_crew_cannot_edit_pic_reviewed(self):
        """RBAC matrix: VESSEL_CREW cannot edit post-submit inspections."""
        response = self._update(user=self.vessel_crew, inspection=self.pic_reviewed_inspection)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_feat_ins_008_validation_non_psc_rejects_psc_subtype_on_update(self):
        """VALIDATION_RULES 2.1: non-PSC inspection cannot keep PSC subtype."""
        response = self._update(
            user=self.office_pic,
            inspection=self.submitted_inspection,
            payload=self._payload(inspection_type="RS", psc_subtype="INITIAL"),
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("psc_subtype", response.data)

    def test_feat_ins_008_error_missing_inspection_returns_404(self):
        """BACKEND_STRUCTURE 10.3: update on missing inspection should return NOT_FOUND."""
        missing_id = uuid.UUID("00000000-0000-0000-0000-000000000000")
        request = self.factory.put(
            f"/api/psc/inspections/{missing_id}/update/",
            self._payload(),
            format="json",
        )
        force_authenticate(request, user=self.office_pic)
        response = self.view(request, id=missing_id)

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_feat_ins_008_rbac_unauthenticated_edit_rejected(self):
        """RBAC: unauthenticated post-submit edit request is rejected."""
        response = self._update(user=None, inspection=self.submitted_inspection)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_feat_ins_008_gap_revision_no_should_increment_on_post_submit_edit(self):
        """PRD FEAT-INS-008: post-submit edit must increment revision_no."""
        before_revision = self.submitted_inspection.revision_no
        response = self._update(
            user=self.office_pic,
            inspection=self.submitted_inspection,
            payload=self._payload(port_place="Manila"),
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.submitted_inspection.refresh_from_db()
        self.assertEqual(self.submitted_inspection.revision_no, before_revision + 1)

    def test_feat_ins_008_gap_post_submit_edit_should_create_audit_log(self):
        """PRD FEAT-INS-008: post-submit edits require full field-level audit logging."""
        before_count = AuditLog.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.submitted_inspection.id,
            action="UPDATE",
        ).count()

        response = self._update(
            user=self.office_pic,
            inspection=self.submitted_inspection,
            payload=self._payload(port_place="Fujairah"),
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        after_count = AuditLog.objects.filter(
            entity_type="INSPECTION",
            entity_id=self.submitted_inspection.id,
            action="UPDATE",
        ).count()
        self.assertEqual(after_count, before_count + 1)


class TestFEAT_INS_009_DeleteDraftInspection(BaseInspectionAPITestCase):
    """
    FEAT-INS-009: Delete Draft Inspection.

    PRD Reference: Docs/PRD.md FEAT-INS-009
    Validation Reference: Docs/VALIDATION_RULES.md Section 2.5
    API/RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.3, 11
    """

    def setUp(self):
        super().setUp()
        self.view = InspectionDeleteView.as_view()
        self.draft_inspection = self.create_inspection(
            status=InspectionStatus.DRAFT,
            inspection_type="PSC",
            psc_subtype="INITIAL",
        )
        self.submitted_inspection = self.create_inspection(
            status=InspectionStatus.SUBMITTED,
            inspection_type="PSC",
            psc_subtype="INITIAL",
        )

    def _delete(self, *, user=None, inspection=None):
        target = inspection or self.draft_inspection
        request = self.factory.delete(f"/api/psc/inspections/{target.id}/delete/")
        if user is not None:
            force_authenticate(request, user=user)
        return self.view(request, id=target.id)

    def test_feat_ins_009_happy_path_vessel_master_soft_deletes_draft(self):
        """PRD FEAT-INS-009: Vessel Master can soft-delete DRAFT inspection."""
        response = self._delete(user=self.vessel_master, inspection=self.draft_inspection)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.draft_inspection.refresh_from_db()
        self.assertTrue(self.draft_inspection.is_deleted)

    def test_feat_ins_009_precondition_non_draft_cannot_be_deleted(self):
        """VALIDATION_RULES 2.5: only DRAFT status can be deleted."""
        response = self._delete(user=self.vessel_master, inspection=self.submitted_inspection)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.submitted_inspection.refresh_from_db()
        self.assertFalse(self.submitted_inspection.is_deleted)

    def test_feat_ins_009_rbac_office_cannot_delete(self):
        """RBAC matrix: office users cannot delete inspections."""
        response = self._delete(user=self.office_pic, inspection=self.draft_inspection)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_feat_ins_009_rbac_vessel_crew_cannot_delete(self):
        """RBAC matrix: VESSEL_CREW cannot delete inspections."""
        response = self._delete(user=self.vessel_crew, inspection=self.draft_inspection)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_feat_ins_009_rbac_other_vessel_master_cannot_delete(self):
        """RBAC/data visibility: vessel users cannot delete inspections outside own vessel."""
        other_vessel_master = make_user(
            role=RoleCodes.VESSEL_MASTER,
            user_type="VESSEL",
            vessel_id=str(self.other_vessel_id),
            user_id="vm-other-delete",
            display_name="Other Vessel Master",
        )
        response = self._delete(user=other_vessel_master, inspection=self.draft_inspection)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_feat_ins_009_rbac_unauthenticated_delete_rejected(self):
        """RBAC: unauthenticated delete request is rejected."""
        response = self._delete(user=None, inspection=self.draft_inspection)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_feat_ins_009_error_missing_inspection_returns_404(self):
        """BACKEND_STRUCTURE 10.3: deleting a missing inspection should return NOT_FOUND."""
        missing_id = uuid.UUID("00000000-0000-0000-0000-000000000000")
        request = self.factory.delete(f"/api/psc/inspections/{missing_id}/delete/")
        force_authenticate(request, user=self.vessel_master)
        response = self.view(request, id=missing_id)

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_feat_ins_009_gap_delete_should_soft_delete_related_deficiencies_and_car(self):
        """PRD FEAT-INS-009: associated deficiencies and CARs should also be soft-deleted."""
        car = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-{uuid.uuid4().hex[:6].upper()}",
            status="DRAFT",
        )
        deficiency = Deficiency.objects.create(
            inspection=self.draft_inspection,
            def_code_id="10101",
            def_code="10101",
            description="Deficiency linked to draft inspection delete flow",
            car=car,
            created_by=str(self.vessel_master.id),
        )

        response = self._delete(user=self.vessel_master, inspection=self.draft_inspection)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        deficiency.refresh_from_db()
        car.refresh_from_db()
        self.assertTrue(deficiency.is_deleted)
        self.assertTrue(car.is_deleted)


class TestFEAT_INS_010_ViewInspectionList(BaseInspectionAPITestCase):
    """
    FEAT-INS-010: View Inspection List.

    PRD Reference: Docs/PRD.md FEAT-INS-010
    API/RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.3, 11
    """

    def setUp(self):
        super().setUp()
        self.view = InspectionListView.as_view()

        self.own_draft = self.create_inspection(
            inspection_type="PSC",
            psc_subtype="INITIAL",
            status=InspectionStatus.DRAFT,
            inspection_date=date(2026, 1, 15),
            port_place="Singapore",
            is_detention=False,
        )
        self.own_submitted = self.create_inspection(
            inspection_type="RS",
            psc_subtype=None,
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 20),
            port_place="Rotterdam",
            is_detention=False,
        )
        self.own_detention = self.create_inspection(
            inspection_type="AUDIT",
            psc_subtype=None,
            status=InspectionStatus.PIC_REVIEWED,
            inspection_date=date(2026, 1, 25),
            port_place="Dubai",
            is_detention=True,
        )
        self.other_vessel_inspection = self.create_inspection(
            vessel_id=self.other_vessel_id,
            inspection_type="PSC",
            psc_subtype="INITIAL",
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 22),
            port_place="Busan",
            is_detention=False,
        )

        InspectionReport.objects.create(
            inspection=self.own_submitted,
            file_name="report.pdf",
            file_path="/tmp/report.pdf",
            file_size=1024,
            mime_type="application/pdf",
            uploaded_by=str(self.vessel_master.id),
        )

        linked_car = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-{uuid.uuid4().hex[:6].upper()}",
            status="DRAFT",
        )
        Deficiency.objects.create(
            inspection=self.own_submitted,
            def_code_id="10101",
            def_code="10101",
            description="Open deficiency for list count check",
            is_cleared=False,
            car=linked_car,
        )
        Deficiency.objects.create(
            inspection=self.own_submitted,
            def_code_id="10201",
            def_code="10201",
            description="Cleared deficiency for list count check",
            is_cleared=True,
        )

    def _list(self, user=None, params=None):
        request = self.factory.get("/api/psc/inspections/", params or {})
        if user:
            force_authenticate(request, user=user)

        with patch("apps.inspection.views.vessel_name_annotation", return_value=Value("MV Example")), patch(
            "apps.inspection.views.vessel_code_annotation",
            return_value=Value("EXM"),
        ):
            return self.view(request)

    def test_happy_path_vessel_user_sees_only_own_vessel_inspections(self):
        """PRD FEAT-INS-010: vessel users must only see their vessel inspections."""
        response = self._list(user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        ids = {row["id"] for row in response.data["data"]}
        self.assertIn(str(self.own_draft.id), ids)
        self.assertIn(str(self.own_submitted.id), ids)
        self.assertIn(str(self.own_detention.id), ids)
        self.assertNotIn(str(self.other_vessel_inspection.id), ids)

    def test_happy_path_office_user_can_filter_by_vessel(self):
        """PRD FEAT-INS-010: office users can filter inspection list by vessel."""
        response = self._list(
            user=self.office_pic,
            params={"vessel_id": str(self.other_vessel_id)},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["data"]), 1)
        self.assertEqual(response.data["data"][0]["id"], str(self.other_vessel_inspection.id))

    def test_happy_path_filter_by_status(self):
        """BACKEND_STRUCTURE 10.3: status query filter is supported."""
        response = self._list(
            user=self.office_pic,
            params={"status": InspectionStatus.SUBMITTED},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(len(response.data["data"]), 1)
        self.assertTrue(all(row["status"] == InspectionStatus.SUBMITTED for row in response.data["data"]))

    def test_happy_path_filter_by_inspection_type(self):
        """BACKEND_STRUCTURE 10.3: inspection_type query filter is supported."""
        response = self._list(
            user=self.office_pic,
            params={"inspection_type": "AUDIT"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["data"]), 1)
        self.assertEqual(response.data["data"][0]["id"], str(self.own_detention.id))

    def test_happy_path_filter_by_date_range(self):
        """BACKEND_STRUCTURE 10.3: date_from/date_to filters are supported."""
        response = self._list(
            user=self.vessel_master,
            params={"date_from": "2026-01-19", "date_to": "2026-01-23"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["data"]), 1)
        self.assertEqual(response.data["data"][0]["id"], str(self.own_submitted.id))

    def test_happy_path_pagination_defaults_to_20(self):
        """PRD FEAT-INS-010: default list pagination size is 20."""
        for i in range(22):
            self.create_inspection(
                inspection_type="RS",
                psc_subtype=None,
                status=InspectionStatus.DRAFT,
                inspection_date=date(2026, 2, 1),
                port_place=f"Bulk Port {i}",
            )

        response = self._list(user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["pagination"]["page_size"], 20)
        self.assertEqual(len(response.data["data"]), 20)

    def test_happy_path_page_size_is_capped_at_100(self):
        """BACKEND_STRUCTURE 10.3: max page_size is 100."""
        response = self._list(
            user=self.office_pic,
            params={"page_size": 500},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["pagination"]["page_size"], 100)

    def test_happy_path_detention_flag_is_returned(self):
        """PRD FEAT-INS-010: detention inspections should be distinguishable in list data."""
        response = self._list(
            user=self.vessel_master,
            params={"inspection_type": "AUDIT"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["data"][0]["is_detention"])

    def test_rbac_unauthenticated_list_request_is_rejected(self):
        """RBAC: inspection list requires authentication."""
        response = self._list(user=None)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_gap_list_should_include_deficiency_and_open_counts(self):
        """PRD FEAT-INS-010: list must show deficiency count and open count."""
        response = self._list(
            user=self.vessel_master,
            params={"status": InspectionStatus.SUBMITTED},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        submitted_row = next(
            row for row in response.data["data"] if row["id"] == str(self.own_submitted.id)
        )
        self.assertEqual(submitted_row["deficiency_count"], 2)
        self.assertEqual(submitted_row["open_deficiency_count"], 1)


class TestFEAT_INS_011_ViewInspectionDetail(BaseInspectionAPITestCase):
    """
    FEAT-INS-011: View Inspection Detail.

    PRD Reference: Docs/PRD.md FEAT-INS-011
    API/RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.3, 11
    """

    def setUp(self):
        super().setUp()
        self.view = InspectionDetailView.as_view()

        self.inspection = self.create_inspection(
            status=InspectionStatus.SUBMITTED,
            inspection_type="PSC",
            psc_subtype="INITIAL",
            inspection_date=date(2026, 1, 15),
            port_place="Singapore",
        )

        self.car = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-{uuid.uuid4().hex[:6].upper()}",
            status="DRAFT",
        )
        self.deficiency = Deficiency.objects.create(
            inspection=self.inspection,
            def_code_id="10101",
            def_code="10101",
            description="Certificate deficiency for detail rendering",
            action_code_id=30,
            action_code="30",
            is_cleared=False,
            car=self.car,
        )

        InspectionReport.objects.create(
            inspection=self.inspection,
            file_name="active-report.pdf",
            file_path="/tmp/active-report.pdf",
            file_size=1024,
            mime_type="application/pdf",
            description="Active report",
            is_deleted=False,
            uploaded_by=str(self.vessel_master.id),
        )
        InspectionReport.objects.create(
            inspection=self.inspection,
            file_name="deleted-report.pdf",
            file_path="/tmp/deleted-report.pdf",
            file_size=1024,
            mime_type="application/pdf",
            description="Deleted report",
            is_deleted=True,
            uploaded_by=str(self.vessel_master.id),
        )

        ActivityHistory.objects.create(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
            vessel_id=self.vessel_id,
            event_type="INSPECTION_SUBMITTED",
            event_description="Inspection submitted for office review",
            performed_by=str(self.vessel_master.id),
            performed_by_name=self.vessel_master.display_name,
        )

        AuditLog.objects.create(
            entity_type="INSPECTION",
            entity_id=self.inspection.id,
            action="UPDATE",
            field_name="status",
            old_value="DRAFT",
            new_value="SUBMITTED",
            performed_by=str(self.office_pic.id),
            performed_by_role=RoleCodes.OFFICE_PIC,
            is_office_edit_assist=True,
        )

    def _detail(self, user=None, inspection_id=None):
        target_id = inspection_id or self.inspection.id
        request = self.factory.get(f"/api/psc/inspections/{target_id}/")
        if user:
            force_authenticate(request, user=user)

        with patch("apps.inspection.views.vessel_name_annotation", return_value=Value("MV Example")), patch(
            "apps.inspection.views.vessel_code_annotation",
            return_value=Value("EXM"),
        ), patch("apps.inspection.views.imo_number_annotation", return_value=Value("1234567")), patch(
            "apps.inspection.deficiency_serializers.DeficiencyListSerializer.get_def_code_description",
            return_value="Certificates - International Tonnage",
        ), patch(
            "apps.inspection.deficiency_serializers.DeficiencyListSerializer.get_action_code_description",
            return_value="Deficiency to be rectified",
        ):
            return self.view(request, id=target_id)

    def test_happy_path_returns_full_detail_with_deficiencies_car_and_reports(self):
        """PRD FEAT-INS-011: detail includes inspection fields, deficiencies, and CAR status."""
        response = self._detail(user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.data["data"]
        self.assertEqual(data["id"], str(self.inspection.id))
        self.assertEqual(data["deficiencies"][0]["def_code"], "10101")
        self.assertEqual(data["deficiencies"][0]["car"]["status"], "DRAFT")
        self.assertEqual(len(data["reports"]), 1)
        self.assertEqual(data["reports"][0]["file_name"], "active-report.pdf")

    def test_happy_path_includes_activity_history(self):
        """PRD FEAT-INS-011: detail shows activity history timeline."""
        response = self._detail(user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        activity = response.data["data"]["activity_history"]
        self.assertGreaterEqual(len(activity), 1)
        self.assertEqual(activity[0]["event_type"], "INSPECTION_SUBMITTED")

    def test_happy_path_office_user_can_view_detail(self):
        """RBAC matrix: office users can view inspection details."""
        response = self._detail(user=self.office_pic)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_happy_path_dpa_can_view_detail(self):
        """RBAC matrix: DPA can view inspection details."""
        response = self._detail(user=self.dpa)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_rbac_other_vessel_master_cannot_view_detail(self):
        """BACKEND_STRUCTURE data visibility: vessel users cannot view other vessel inspection."""
        other_vessel_master = make_user(
            role=RoleCodes.VESSEL_MASTER,
            user_type="VESSEL",
            vessel_id=str(self.other_vessel_id),
            user_id="vm-2",
            display_name="Other Vessel Master",
        )
        response = self._detail(user=other_vessel_master)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_unauthenticated_detail_request_is_rejected(self):
        """RBAC: inspection detail requires authentication."""
        response = self._detail(user=None)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_error_returns_not_found_for_missing_inspection(self):
        """BACKEND_STRUCTURE 10.3: detail endpoint returns 404 for missing inspection."""
        missing_id = uuid.UUID("00000000-0000-0000-0000-000000000000")
        response = self._detail(user=self.vessel_master, inspection_id=missing_id)
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_error_soft_deleted_inspection_is_not_returned(self):
        """BACKEND_STRUCTURE 10.3: soft-deleted inspections are excluded from detail endpoint."""
        self.inspection.is_deleted = True
        self.inspection.save(update_fields=["is_deleted"])

        response = self._detail(user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_gap_office_and_dpa_should_see_audit_log_in_detail(self):
        """PRD FEAT-INS-011: office and DPA should have full audit log visibility."""
        response = self._detail(user=self.office_pic)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("audit_log", response.data["data"])
        self.assertGreaterEqual(len(response.data["data"]["audit_log"]), 1)


class TestFEAT_DEF_001_UpdateActionCode(BaseInspectionAPITestCase):
    """
    FEAT-DEF-001: Update Action Code.

    PRD Reference: Docs/PRD.md FEAT-DEF-001
    Validation Reference: Docs/VALIDATION_RULES.md Section 3.2
    API/RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.4, 11
    """

    def setUp(self):
        super().setUp()
        self.view = DeficiencyActionCodeUpdateView.as_view()
        self.inspection = self.create_inspection(
            inspection_type="PSC",
            psc_subtype="INITIAL",
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 15),
            port_place="Singapore",
            mou_id="TOKYO",
        )
        self.car = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-{uuid.uuid4().hex[:6].upper()}",
            status="DRAFT",
        )
        self.deficiency = Deficiency.objects.create(
            inspection=self.inspection,
            def_code_id="10101",
            def_code="10101",
            description="Deficiency for action code update",
            action_code_id=30,
            action_code="30",
            is_cleared=False,
            car=self.car,
            created_by=str(self.vessel_master.id),
        )
        self.follow_up = self.create_inspection(
            inspection_type="PSC",
            psc_subtype="FOLLOW_UP",
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 20),
            port_place="Busan",
            mou_id="TOKYO",
            parent_inspection=self.inspection,
        )

    def _update(self, payload, user=None, deficiency_id=None):
        target_id = deficiency_id or self.deficiency.id
        request = self.factory.put(
            f"/api/psc/deficiencies/{target_id}/action-code/",
            payload,
            format="json",
        )
        if user:
            force_authenticate(request, user=user)
        with patch(
            "apps.inspection.deficiency_serializers.DeficiencyDetailSerializer.get_def_code_description",
            return_value="Mock Def Code",
        ), patch(
            "apps.inspection.deficiency_serializers.DeficiencyDetailSerializer.get_action_code_description",
            return_value="Mock Action Code",
        ):
            return self.view(request, id=target_id)

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_happy_path_vessel_master_updates_action_code_and_creates_history(self, mock_action_get):
        """PRD FEAT-DEF-001: Vessel Master can update action code and history is recorded."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")
        before = self.deficiency.action_history.count()

        response = self._update(
            {
                "action_code_id": 10,
                "change_reason": "Rectified during follow-up inspection.",
            },
            user=self.vessel_master,
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.deficiency.refresh_from_db()
        self.assertEqual(self.deficiency.action_code_id, 10)
        self.assertEqual(self.deficiency.action_code, "10")
        self.assertTrue(self.deficiency.is_cleared)
        self.assertIsNotNone(self.deficiency.cleared_date)
        self.assertEqual(self.deficiency.action_history.count(), before + 1)
        history = self.deficiency.action_history.first()
        self.assertEqual(history.previous_action_code_id, 30)
        self.assertEqual(history.new_action_code_id, 10)

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_happy_path_action_code_30_can_transition_to_non_clearing_code(self, mock_action_get):
        """PRD FEAT-DEF-001: action code 30 can transition to any code."""
        mock_action_get.return_value = SimpleNamespace(action_code=17, definition="To be rectified")

        response = self._update(
            {
                "action_code_id": 17,
                "change_reason": "Updated after authority discussion.",
            },
            user=self.vessel_master,
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.deficiency.refresh_from_db()
        self.assertEqual(self.deficiency.action_code_id, 17)
        self.assertEqual(self.deficiency.action_code, "17")
        self.assertFalse(self.deficiency.is_cleared)

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_happy_path_follow_up_reference_links_deficiency_and_history(self, mock_action_get):
        """PRD FEAT-DEF-001: update can be linked to follow-up inspection."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")

        response = self._update(
            {
                "action_code_id": 10,
                "follow_up_inspection_id": str(self.follow_up.id),
                "change_reason": "Cleared during PSC follow-up.",
            },
            user=self.vessel_master,
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.deficiency.refresh_from_db()
        self.assertEqual(self.deficiency.cleared_by_follow_up_id, self.follow_up.id)
        history = self.deficiency.action_history.first()
        self.assertEqual(history.follow_up_inspection_id, self.follow_up.id)

    def test_validation_action_code_is_required(self):
        """VALIDATION_RULES 3.2: action_code_id is mandatory."""
        response = self._update({"change_reason": "Missing action code"}, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("action_code_id", response.data["details"])

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_validation_invalid_action_code_returns_400(self, mock_action_get):
        """VALIDATION_RULES 3.2: action_code_id must exist in master table."""
        mock_action_get.side_effect = PSCActionCode.DoesNotExist
        response = self._update({"action_code_id": 999}, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("action_code_id", response.data["details"])

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_validation_invalid_follow_up_inspection_returns_400(self, mock_action_get):
        """VALIDATION_RULES 3.2: follow_up_inspection_id must exist when provided."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")
        response = self._update(
            {
                "action_code_id": 10,
                "follow_up_inspection_id": str(uuid.uuid4()),
            },
            user=self.vessel_master,
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("follow_up_inspection_id", response.data["details"])

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_validation_change_reason_over_500_chars_returns_400(self, mock_action_get):
        """VALIDATION_RULES 3.2: change_reason max length is 500."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")
        response = self._update(
            {
                "action_code_id": 10,
                "change_reason": "x" * 501,
            },
            user=self.vessel_master,
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("change_reason", response.data["details"])

    def test_rbac_vessel_crew_cannot_update_action_code(self):
        """RBAC matrix: VESSEL_CREW cannot update action code."""
        response = self._update({"action_code_id": 10}, user=self.vessel_crew)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_office_user_cannot_update_action_code(self):
        """RBAC matrix: OFFICE users cannot update action code."""
        response = self._update({"action_code_id": 10}, user=self.office_pic)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_vessel_master_cannot_update_other_vessel_deficiency(self):
        """RBAC/data visibility: vessel users can only update deficiencies on own vessel."""
        other_inspection = self.create_inspection(
            vessel_id=self.other_vessel_id,
            inspection_type="PSC",
            psc_subtype="INITIAL",
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 16),
            port_place="Doha",
            mou_id="TOKYO",
        )
        other_deficiency = Deficiency.objects.create(
            inspection=other_inspection,
            def_code_id="10102",
            def_code="10102",
            description="Other vessel deficiency",
            action_code_id=30,
            action_code="30",
        )
        response = self._update(
            {"action_code_id": 10},
            user=self.vessel_master,
            deficiency_id=other_deficiency.id,
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_unauthenticated_update_is_rejected(self):
        """RBAC: endpoint requires authentication."""
        response = self._update({"action_code_id": 10}, user=None)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_error_missing_deficiency_returns_404(self):
        """API contract: missing deficiency should return 404."""
        response = self._update(
            {"action_code_id": 10},
            user=self.vessel_master,
            deficiency_id=uuid.UUID("00000000-0000-0000-0000-000000000000"),
        )
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_gap_follow_up_inspection_should_be_follow_up_type(self, mock_action_get):
        """VALIDATION GAP: follow_up_inspection_id should require FOLLOW_UP subtype."""
        non_follow_up = self.create_inspection(
            inspection_type="PSC",
            psc_subtype="INITIAL",
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 21),
            port_place="Kobe",
            mou_id="TOKYO",
        )
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")

        response = self._update(
            {
                "action_code_id": 10,
                "follow_up_inspection_id": str(non_follow_up.id),
            },
            user=self.vessel_master,
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("follow_up_inspection_id", response.data["details"])


class TestFEAT_DEF_002_RegisterPSCFollowUp(BaseInspectionAPITestCase):
    """
    FEAT-DEF-002: Register PSC Follow-up.

    PRD Reference: Docs/PRD.md FEAT-DEF-002
    Validation Reference: Docs/VALIDATION_RULES.md Sections 2.1, 3.2
    API/RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.4, 11
    """

    def setUp(self):
        super().setUp()
        self.view = FollowUpView.as_view()
        self.parent_psc = self.create_inspection(
            inspection_type="PSC",
            psc_subtype="INITIAL",
            status=InspectionStatus.DPA_CLOSED,
            inspection_date=date(2026, 1, 10),
            port_place="Singapore",
            mou_id="TOKYO",
        )
        self.parent_rs = self.create_inspection(
            inspection_type="RS",
            psc_subtype=None,
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 10),
            port_place="Busan",
            mou_id=None,
        )

        self.car_1 = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-{uuid.uuid4().hex[:6].upper()}",
            status="DRAFT",
        )
        self.car_2 = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-{uuid.uuid4().hex[:6].upper()}",
            status="DRAFT",
        )
        self.def_1 = Deficiency.objects.create(
            inspection=self.parent_psc,
            def_code_id="10101",
            def_code="10101",
            description="First deficiency pending follow-up",
            action_code_id=30,
            action_code="30",
            is_cleared=False,
            car=self.car_1,
        )
        self.def_2 = Deficiency.objects.create(
            inspection=self.parent_psc,
            def_code_id="10102",
            def_code="10102",
            description="Second deficiency pending follow-up",
            action_code_id=30,
            action_code="30",
            is_cleared=False,
            car=self.car_2,
        )

    def _register(self, payload, user=None):
        request = self.factory.post("/api/psc/psc-follow-up/register/", payload, format="json")
        if user:
            force_authenticate(request, user=user)
        return self.view(request)

    def _action_code_getter(self, action_code):
        return SimpleNamespace(action_code=action_code, definition=f"Action {action_code}")

    @patch("apps.inspection.followup_views.notify_psc_follow_up_recorded")
    def test_happy_path_creates_follow_up_inspection_with_required_fields(self, mock_notify):
        """PRD FEAT-DEF-002: creates linked PSC FOLLOW_UP inspection with captured metadata."""
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
            "country": "Japan",
            "authority": "MLIT",
            "inspector_name": "Inspector A",
            "report_reference": "FU-001",
            "deficiency_updates": [],
        }
        response = self._register(payload, user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        data = response.data["data"]["follow_up_inspection"]
        self.assertEqual(data["inspection_type"], "PSC")
        self.assertEqual(data["psc_subtype"], "FOLLOW_UP")
        self.assertEqual(data["status"], "SUBMITTED")
        self.assertEqual(data["parent_inspection_id"], str(self.parent_psc.id))
        follow_up = Inspection.objects.get(id=data["id"])
        self.assertEqual(follow_up.port_place, "Yokohama")
        self.assertEqual(follow_up.authority, "MLIT")
        self.assertEqual(follow_up.mou_id, self.parent_psc.mou_id)
        mock_notify.assert_called_once_with(follow_up, str(self.parent_psc.vessel_id))

    @patch("apps.inspection.followup_serializers.PSCActionCode.objects.get")
    @patch("apps.inspection.followup_views.notify_psc_follow_up_recorded")
    def test_happy_path_batch_updates_deficiencies_and_creates_history(self, mock_notify, mock_action_get):
        """PRD FEAT-DEF-002: supports batch action updates and links records to follow-up."""
        mock_action_get.side_effect = self._action_code_getter
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
            "authority": "MLIT",
            "deficiency_updates": [
                {"deficiency_id": str(self.def_1.id), "action_code_id": 10},
                {"deficiency_id": str(self.def_2.id), "action_code_id": 10},
            ],
        }
        before_1 = self.def_1.action_history.count()
        before_2 = self.def_2.action_history.count()
        response = self._register(payload, user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["data"]["updated_deficiencies_count"], 2)
        follow_up_id = response.data["data"]["follow_up_inspection"]["id"]
        self.def_1.refresh_from_db()
        self.def_2.refresh_from_db()
        self.assertEqual(self.def_1.action_code_id, 10)
        self.assertEqual(self.def_2.action_code_id, 10)
        self.assertTrue(self.def_1.is_cleared)
        self.assertTrue(self.def_2.is_cleared)
        self.assertEqual(str(self.def_1.cleared_by_follow_up_id), follow_up_id)
        self.assertEqual(str(self.def_2.cleared_by_follow_up_id), follow_up_id)
        self.assertEqual(self.def_1.action_history.count(), before_1 + 1)
        self.assertEqual(self.def_2.action_history.count(), before_2 + 1)
        mock_notify.assert_called_once()

    @patch("apps.inspection.followup_serializers.PSCActionCode.objects.get")
    def test_happy_path_non_clearing_action_does_not_mark_deficiency_cleared(self, mock_action_get):
        """Business rule: only clearing action code (10) auto-clears deficiency."""
        mock_action_get.side_effect = self._action_code_getter
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
            "deficiency_updates": [
                {"deficiency_id": str(self.def_1.id), "action_code_id": 17},
            ],
        }
        response = self._register(payload, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.def_1.refresh_from_db()
        self.assertEqual(self.def_1.action_code_id, 17)
        self.assertFalse(self.def_1.is_cleared)

    def test_rbac_vessel_crew_cannot_register_follow_up(self):
        """RBAC matrix: only Vessel Master can register PSC follow-up."""
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
        }
        response = self._register(payload, user=self.vessel_crew)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_office_user_cannot_register_follow_up(self):
        """RBAC matrix: Office roles cannot register PSC follow-up."""
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
        }
        response = self._register(payload, user=self.office_pic)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_rbac_unauthenticated_register_is_rejected(self):
        """RBAC: follow-up endpoint requires authentication."""
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
        }
        response = self._register(payload, user=None)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_validation_parent_inspection_must_be_psc(self):
        """PRD FEAT-DEF-002: follow-up can only be registered for PSC inspections."""
        payload = {
            "parent_inspection_id": str(self.parent_rs.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
        }
        response = self._register(payload, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("parent_inspection_id", response.data["details"])

    def test_validation_follow_up_date_cannot_be_before_parent_date(self):
        """Validation: follow-up date cannot be before original inspection date."""
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-09",
            "port_place": "Yokohama",
        }
        response = self._register(payload, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("inspection_date", response.data["details"])

    @patch("apps.inspection.followup_serializers.PSCActionCode.objects.get")
    def test_validation_deficiency_updates_must_belong_to_parent(self, mock_action_get):
        """Validation: batch deficiency updates must reference parent inspection deficiencies."""
        mock_action_get.side_effect = self._action_code_getter
        other_inspection = self.create_inspection(
            inspection_type="PSC",
            psc_subtype="INITIAL",
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 12),
            port_place="Osaka",
            mou_id="TOKYO",
        )
        other_deficiency = Deficiency.objects.create(
            inspection=other_inspection,
            def_code_id="10202",
            def_code="10202",
            description="Deficiency from another inspection",
            action_code_id=30,
            action_code="30",
        )
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
            "deficiency_updates": [
                {"deficiency_id": str(other_deficiency.id), "action_code_id": 10},
            ],
        }
        response = self._register(payload, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("deficiency_updates", response.data["details"])

    @patch("apps.inspection.followup_serializers.PSCActionCode.objects.get")
    def test_validation_invalid_action_code_returns_400(self, mock_action_get):
        """Validation: action_code_id in updates must exist in master action codes."""
        mock_action_get.side_effect = PSCActionCode.DoesNotExist
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
            "deficiency_updates": [
                {"deficiency_id": str(self.def_1.id), "action_code_id": 999},
            ],
        }
        response = self._register(payload, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("deficiency_updates", response.data["details"])

    def test_validation_required_fields_return_standard_error_payload(self):
        """Error format: validation failures return error/message/details."""
        response = self._register({}, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data["error"], "VALIDATION_ERROR")
        self.assertEqual(response.data["message"], "Invalid follow-up registration data")
        self.assertIn("details", response.data)

    @patch("apps.inspection.followup_serializers.PSCActionCode.objects.get")
    def test_gap_follow_up_registration_should_create_activity_event(self, mock_action_get):
        """GAP: FEAT-DEF-002 requires follow-up event history record."""
        mock_action_get.side_effect = self._action_code_getter
        before = ActivityHistory.objects.filter(
            entity_type="INSPECTION",
            event_type="PSC_FOLLOW_UP_RECORDED",
        ).count()
        payload = {
            "parent_inspection_id": str(self.parent_psc.id),
            "inspection_date": "2026-01-20",
            "port_place": "Yokohama",
            "deficiency_updates": [
                {"deficiency_id": str(self.def_1.id), "action_code_id": 10},
            ],
        }
        response = self._register(payload, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        after = ActivityHistory.objects.filter(
            entity_type="INSPECTION",
            event_type="PSC_FOLLOW_UP_RECORDED",
        ).count()
        self.assertEqual(after, before + 1)


class TestFEAT_DEF_003_MarkDeficiencyCleared(BaseInspectionAPITestCase):
    """
    FEAT-DEF-003: Mark Deficiency Cleared.

    PRD Reference: Docs/PRD.md FEAT-DEF-003
    Validation Reference: Docs/VALIDATION_RULES.md Sections 3.2 and 13.1
    API/RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.3, 10.4, 11
    """

    def setUp(self):
        super().setUp()
        self.update_view = DeficiencyActionCodeUpdateView.as_view()
        self.detail_view = InspectionDetailView.as_view()
        self.export_view = DeficiencyExcelExportView.as_view()

        self.inspection = self.create_inspection(
            inspection_type="PSC",
            psc_subtype="INITIAL",
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 15),
            port_place="Singapore",
            mou_id="TOKYO",
        )
        self.follow_up = self.create_inspection(
            inspection_type="PSC",
            psc_subtype="FOLLOW_UP",
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 20),
            port_place="Yokohama",
            mou_id="TOKYO",
            parent_inspection=self.inspection,
        )
        self.car = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-{uuid.uuid4().hex[:6].upper()}",
            status="DRAFT",
        )
        self.deficiency = Deficiency.objects.create(
            inspection=self.inspection,
            def_code_id="10101",
            def_code="10101",
            description="Deficiency pending rectification.",
            action_code_id=30,
            action_code="30",
            is_cleared=False,
            car=self.car,
            created_by=str(self.vessel_master.id),
        )

    def _update_action_code(self, payload, user=None):
        request = self.factory.put(
            f"/api/psc/deficiencies/{self.deficiency.id}/action-code/",
            payload,
            format="json",
        )
        if user:
            force_authenticate(request, user=user)
        with patch(
            "apps.inspection.deficiency_serializers.DeficiencyDetailSerializer.get_def_code_description",
            return_value="Mock Def Code",
        ), patch(
            "apps.inspection.deficiency_serializers.DeficiencyDetailSerializer.get_action_code_description",
            return_value="Mock Action Code",
        ):
            return self.update_view(request, id=self.deficiency.id)

    def _detail(self, user=None):
        request = self.factory.get(f"/api/psc/inspections/{self.inspection.id}/")
        if user:
            force_authenticate(request, user=user)
        with patch("apps.inspection.views.vessel_name_annotation", return_value=Value("MV Example")), patch(
            "apps.inspection.views.vessel_code_annotation",
            return_value=Value("EXM"),
        ), patch("apps.inspection.views.imo_number_annotation", return_value=Value("1234567")), patch(
            "apps.inspection.deficiency_serializers.DeficiencyListSerializer.get_def_code_description",
            return_value="Certificates - International Tonnage",
        ), patch(
            "apps.inspection.deficiency_serializers.DeficiencyListSerializer.get_action_code_description",
            return_value="Deficiency rectified",
        ):
            return self.detail_view(request, id=self.inspection.id)

    def _export_excel(self, user=None):
        request = self.factory.get("/api/psc/inspections/export-excel/")
        if user:
            force_authenticate(request, user=user)
        return self.export_view(request)

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_feat_def_003_happy_path_action_code_10_marks_deficiency_cleared(self, mock_action_get):
        """PRD FEAT-DEF-003: action code 10 automatically sets is_cleared and cleared_date."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")

        response = self._update_action_code({"action_code_id": 10}, user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.deficiency.refresh_from_db()
        self.assertTrue(self.deficiency.is_cleared)
        self.assertIsNotNone(self.deficiency.cleared_date)

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_feat_def_003_happy_path_follow_up_reference_is_recorded_on_clear(self, mock_action_get):
        """PRD FEAT-DEF-003: clearing can store cleared_by_follow_up_id reference."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")

        response = self._update_action_code(
            {
                "action_code_id": 10,
                "follow_up_inspection_id": str(self.follow_up.id),
            },
            user=self.vessel_master,
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.deficiency.refresh_from_db()
        self.assertEqual(self.deficiency.cleared_by_follow_up_id, self.follow_up.id)
        self.assertTrue(self.deficiency.is_cleared)
        self.assertIsNotNone(self.deficiency.cleared_date)

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_feat_def_003_happy_path_detail_list_reflects_cleared_status(self, mock_action_get):
        """PRD FEAT-DEF-003: deficiency clear status is reflected in inspection detail lists."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")
        update_response = self._update_action_code({"action_code_id": 10}, user=self.vessel_master)
        self.assertEqual(update_response.status_code, status.HTTP_200_OK)

        detail_response = self._detail(user=self.vessel_master)

        self.assertEqual(detail_response.status_code, status.HTTP_200_OK)
        deficiency_item = detail_response.data["data"]["deficiencies"][0]
        self.assertTrue(deficiency_item["is_cleared"])
        self.assertIsNotNone(deficiency_item["cleared_date"])

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_feat_def_003_happy_path_report_export_payload_reflects_cleared_status(self, mock_action_get):
        """PRD FEAT-DEF-003: cleared status should flow into deficiency report export payload."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")
        update_response = self._update_action_code({"action_code_id": 10}, user=self.vessel_master)
        self.assertEqual(update_response.status_code, status.HTTP_200_OK)

        with patch(
            "apps.inspection.report_views._lookup_vessel_names",
            return_value={},
        ), patch(
            "apps.inspection.report_views.generate_deficiency_excel",
            return_value=b"mock-excel-bytes",
        ) as mock_generate:
            export_response = self._export_excel(user=self.vessel_master)

        self.assertEqual(export_response.status_code, status.HTTP_200_OK)
        deficiency_payload = mock_generate.call_args.args[0]
        self.assertEqual(len(deficiency_payload), 1)
        self.assertTrue(deficiency_payload[0]["is_cleared"])
        self.assertIsNotNone(deficiency_payload[0]["cleared_date"])

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_feat_def_003_edge_non_rectified_action_does_not_mark_cleared(self, mock_action_get):
        """PRD FEAT-DEF-003: only action code 10 triggers auto-clear behavior."""
        mock_action_get.return_value = SimpleNamespace(action_code=17, definition="To be rectified")

        response = self._update_action_code({"action_code_id": 17}, user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.deficiency.refresh_from_db()
        self.assertFalse(self.deficiency.is_cleared)
        self.assertIsNone(self.deficiency.cleared_date)

    # ---- Gap-closing tests (Session 50) ----

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_feat_def_003_gap_clearing_creates_action_history_entry(self, mock_action_get):
        """PRD FEAT-DEF-003: action code change to rectified creates DeficiencyActionHistory."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")

        response = self._update_action_code({"action_code_id": 10}, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        history_count = DeficiencyActionHistory.objects.filter(
            deficiency=self.deficiency,
        ).count()
        self.assertGreaterEqual(history_count, 1, "Action code change should create history entry.")

    @patch("apps.inspection.deficiency_serializers.PSCActionCode.objects.get")
    def test_feat_def_003_gap_multi_deficiency_only_targeted_is_cleared(self, mock_action_get):
        """PRD FEAT-DEF-003: clearing one deficiency does not affect others on same inspection."""
        mock_action_get.return_value = SimpleNamespace(action_code=10, definition="Rectified")

        second_def = Deficiency.objects.create(
            inspection=self.inspection,
            def_code_id="20202",
            def_code="20202",
            description="Second deficiency, not cleared.",
            action_code_id=30,
            action_code="30",
            is_cleared=False,
            car=CAR.objects.create(
                car_number=f"PSC-{date.today().year}-{uuid.uuid4().hex[:6].upper()}",
                status="DRAFT",
            ),
            created_by=str(self.vessel_master.id),
        )

        response = self._update_action_code({"action_code_id": 10}, user=self.vessel_master)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.deficiency.refresh_from_db()
        second_def.refresh_from_db()
        self.assertTrue(self.deficiency.is_cleared)
        self.assertFalse(second_def.is_cleared, "Second deficiency should NOT be cleared.")


class TestFEAT_RPT_002_BulkCARPDFExport(BaseInspectionAPITestCase):
    """Inspection-level CAR PDF export should attach evidence preview URLs for every CAR."""

    def setUp(self):
        super().setUp()
        self.view = BulkCARExportView.as_view()
        self.inspection = self.create_inspection(status=InspectionStatus.SUBMITTED)

        self.car_one = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-BULK01",
            status="SUBMITTED_TO_PIC",
            created_by=str(self.vessel_master.id),
        )
        self.car_two = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-BULK02",
            status="SUBMITTED_TO_PIC",
            created_by=str(self.vessel_master.id),
        )

        Deficiency.objects.create(
            inspection=self.inspection,
            def_code_id="10101",
            def_code="10101",
            description="Bulk export deficiency one.",
            action_code_id=30,
            action_code="30",
            car=self.car_one,
            created_by=str(self.vessel_master.id),
        )
        Deficiency.objects.create(
            inspection=self.inspection,
            def_code_id="20202",
            def_code="20202",
            description="Bulk export deficiency two.",
            action_code_id=17,
            action_code="17",
            car=self.car_two,
            created_by=str(self.vessel_master.id),
        )

        Evidence.objects.create(
            car=self.car_one,
            evidence_type="BEFORE",
            file_name="bulk-one-before.jpg",
            file_path="/tmp/bulk-one-before.jpg",
            file_size=1024,
            mime_type="image/jpeg",
            description="First bulk-export evidence attachment.",
            uploaded_by=str(self.vessel_master.id),
        )
        Evidence.objects.create(
            car=self.car_two,
            evidence_type="AFTER",
            file_name="bulk-two-after.jpg",
            file_path="/tmp/bulk-two-after.jpg",
            file_size=2048,
            mime_type="image/jpeg",
            description="Second bulk-export evidence attachment.",
            uploaded_by=str(self.vessel_master.id),
        )

    def _export(self, user=None, audience=None):
        url = f"/api/psc/inspections/{self.inspection.id}/cars/export-pdf/"
        if audience:
            url = f"{url}?audience={audience}"
        request = self.factory.get(url)
        force_authenticate(request, user=user or self.vessel_master)
        return self.view(request, inspection_id=self.inspection.id)

    @patch("apps.inspection.report_views.generate_car_pdf", return_value=b"%PDF-1.4 bulk")
    def test_bulk_export_attaches_report_preview_urls_to_each_car_payload(self, mock_generate):
        response = self._export()

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertEqual(mock_generate.call_count, 2)

        for call in mock_generate.call_args_list:
            payload = call.args[0]
            self.assertTrue(payload["evidence"])
            preview_url = payload["evidence"][0].get("report_preview_url", "")
            self.assertIn("/api/psc/evidence/", preview_url)
            self.assertIn("report_token=", preview_url)

    @patch("apps.inspection.report_views.generate_car_pdf", return_value=b"%PDF-1.4 bulk")
    def test_bulk_export_passes_vessel_name_to_pdf_payload(self, mock_generate):
        vessel_lookup = {
            str(self.inspection.vessel_id).replace("-", "").lower(): "MV Example",
        }
        with patch("apps.inspection.report_views._lookup_vessel_names", return_value=vessel_lookup):
            response = self._export()

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(mock_generate.call_count, 2)

        for call in mock_generate.call_args_list:
            payload = call.args[0]
            self.assertEqual(payload["inspection"]["vessel_name"], "MV Example")
            self.assertEqual(payload["inspection"]["vessel"]["name"], "MV Example")

    @patch("apps.inspection.report_views.generate_car_pdf", return_value=b"%PDF-1.4 bulk")
    def test_bulk_export_supports_external_audience(self, mock_generate):
        response = self._export(audience="external")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertEqual(mock_generate.call_count, 2)
        for call in mock_generate.call_args_list:
            self.assertEqual(call.kwargs.get("audience"), "external")


class TestFEAT_RPT_002_DeficiencyExcelExport(BaseInspectionAPITestCase):
    """
    FEAT-RPT-002: Deficiency Excel Export.

    PRD Reference: Docs/PRD.md FEAT-RPT-002
    API/RBAC Reference: Docs/BACKEND_STRUCTURE.md Sections 10.3, 11.2
    """

    def setUp(self):
        super().setUp()
        self.view = DeficiencyExcelExportView.as_view()
        self.office_supt = make_user(
            role=RoleCodes.OFFICE_SUPT,
            user_type="OFFICE",
            user_id="office-2",
            display_name="Office Superintendent",
        )

        self.inspection_own = self.create_inspection(
            vessel_id=self.vessel_id,
            inspection_type="PSC",
            psc_subtype="INITIAL",
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 1, 15),
            port_place="Singapore",
            is_detention=True,
        )
        self.car_own = CAR.objects.create(
            car_number=f"PSC-{date.today().year}-RPT001",
            status="SUBMITTED",
            root_cause_summary="Own vessel CAR root cause narrative for export coverage.",
            created_by=str(self.vessel_master.id),
        )
        self.def_own = Deficiency.objects.create(
            inspection=self.inspection_own,
            def_code_id="10101",
            def_code="10101",
            description="Own vessel deficiency for export tests.",
            action_code_id=30,
            action_code="30",
            car=self.car_own,
            is_deleted=False,
            created_by=str(self.vessel_master.id),
        )

        self.inspection_other = self.create_inspection(
            vessel_id=self.other_vessel_id,
            inspection_type="RS",
            psc_subtype=None,
            status=InspectionStatus.SUBMITTED,
            inspection_date=date(2026, 2, 20),
            port_place="Busan",
            is_detention=False,
        )
        self.car_other = CAR.objects.create(
            car_number=f"RS-{date.today().year}-RPT002",
            status="DRAFT",
            root_cause_summary="Other vessel CAR root cause narrative for export coverage.",
            created_by=str(self.vessel_master.id),
        )
        self.def_other = Deficiency.objects.create(
            inspection=self.inspection_other,
            def_code_id="20202",
            def_code="20202",
            description="Other vessel deficiency for export tests.",
            action_code_id=17,
            action_code="17",
            car=self.car_other,
            is_deleted=False,
            created_by=str(self.vessel_master.id),
        )

    def _export(self, user=None, params=None):
        request = self.factory.get("/api/psc/inspections/export-excel/", params or {})
        if user:
            force_authenticate(request, user=user)
        return self.view(request)

    def test_feat_rpt_002_happy_path_export_returns_xlsx_attachment(self):
        """PRD FEAT-RPT-002: export endpoint returns downloadable XLSX workbook."""
        with patch(
            "apps.inspection.report_views._lookup_vessel_names",
            return_value={},
        ):
            response = self._export(user=self.office_pic)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment; filename=", response["Content-Disposition"])
        self.assertIn("Deficiency_Export_", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_feat_rpt_002_rbac_unauthenticated_export_rejected(self):
        """RBAC BACKEND_STRUCTURE 11.1: export endpoint requires authentication."""
        response = self._export(user=None)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_feat_rpt_002_happy_path_vessel_export_is_scoped_to_own_vessel(self):
        """BACKEND_STRUCTURE 11.2: vessel users should only export own vessel deficiencies."""
        with patch(
            "apps.inspection.report_views._lookup_vessel_names",
            return_value={},
        ), patch(
            "apps.inspection.report_views.generate_deficiency_excel",
            return_value=b"mock-excel",
        ) as mock_generate:
            response = self._export(user=self.vessel_master)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = mock_generate.call_args.args[0]
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]["def_code"], self.def_own.def_code)

    def test_feat_rpt_002_happy_path_filters_are_applied_before_export(self):
        """PRD FEAT-RPT-002: export should honor inspection filters from list context."""
        with patch(
            "apps.inspection.report_views._lookup_vessel_names",
            return_value={},
        ), patch(
            "apps.inspection.report_views.generate_deficiency_excel",
            return_value=b"mock-excel",
        ) as mock_generate:
            response = self._export(
                user=self.office_supt,
                params={
                    "inspection_type": "PSC",
                    "date_from": "2026-01-01",
                    "date_to": "2026-01-31",
                },
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = mock_generate.call_args.args[0]
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]["def_code"], self.def_own.def_code)
        self.assertEqual(payload[0]["inspection"]["inspection_type"], "PSC")

    def test_feat_rpt_002_gap_vessel_name_lookup_should_use_cursor_mapping(self):
        """GAP PRD FEAT-RPT-002: exported rows should show resolved vessel name, not raw vessel UUID."""
        vessel_key = str(self.vessel_id).replace('-', '').lower()
        with patch(
            "apps.inspection.report_views._lookup_vessel_names",
            return_value={vessel_key: "MV Example"},
        ), patch(
            "apps.inspection.report_views.generate_deficiency_excel",
            return_value=b"mock-excel",
        ) as mock_generate:
            response = self._export(
                user=self.office_pic,
                params={
                    "inspection_type": "PSC",
                    "date_from": "2026-01-01",
                    "date_to": "2026-01-31",
                },
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = mock_generate.call_args.args[0]
        self.assertEqual(payload[0]["inspection"]["vessel_name"], "MV Example")

    def test_feat_rpt_002_happy_path_generator_builds_multi_sheet_workbook(self):
        """PRD FEAT-RPT-002: workbook must contain Deficiency Summary, CAR Status, and Applied Filters sheets."""
        deficiencies = [
            {
                "def_code": "10101",
                "description": "Summary row one",
                "action_code": "30",
                "target_date": "2026-01-22",
                "is_cleared": False,
                "cleared_date": None,
                "inspection": {
                    "vessel_name": "MV Example",
                    "inspection_type": "PSC",
                    "inspection_date": "2026-01-15",
                    "port_place": "Singapore",
                    "is_detention": True,
                },
                "car": {
                    "car_number": "PSC-2026-001",
                    "status_display": "Submitted",
                    "root_cause_summary": "R" * 60,
                    "target_date": "2026-01-25",
                    "evidence_count": 2,
                    "actions_count": 3,
                },
            },
            {
                "def_code": "20202",
                "description": "Summary row two",
                "action_code": "17",
                "target_date": "2026-02-10",
                "is_cleared": True,
                "cleared_date": "2026-02-12",
                "inspection": {
                    "vessel_name": "MV Other",
                    "inspection_type": "RS",
                    "inspection_date": "2026-02-01",
                    "port_place": "Busan",
                    "is_detention": False,
                },
                "car": {
                    "car_number": "RS-2026-002",
                    "status_display": "Draft",
                    "root_cause_summary": "R" * 55,
                    "target_date": "2026-02-18",
                    "evidence_count": 0,
                    "actions_count": 1,
                },
            },
        ]
        excel_bytes = generate_deficiency_excel(deficiencies, {"total_count": 2})
        workbook = load_workbook(BytesIO(excel_bytes))

        self.assertEqual(workbook.sheetnames, ["Deficiency Summary", "CAR Status", "Applied Filters"])

    def test_feat_rpt_002_happy_path_generator_applies_detention_highlight_and_auto_filter(self):
        """PRD FEAT-RPT-002: detention rows should be highlighted and auto-filter enabled."""
        deficiencies = [
            {
                "def_code": "10101",
                "description": "Detention row",
                "action_code": "30",
                "target_date": "2026-01-22",
                "is_cleared": False,
                "cleared_date": None,
                "inspection": {
                    "vessel_name": "MV Example",
                    "inspection_type": "PSC",
                    "inspection_date": "2026-01-15",
                    "port_place": "Singapore",
                    "is_detention": True,
                },
                "car": {},
            },
            {
                "def_code": "20202",
                "description": "Non detention row",
                "action_code": "17",
                "target_date": "2026-02-10",
                "is_cleared": False,
                "cleared_date": None,
                "inspection": {
                    "vessel_name": "MV Other",
                    "inspection_type": "RS",
                    "inspection_date": "2026-02-01",
                    "port_place": "Busan",
                    "is_detention": False,
                },
                "car": {},
            },
        ]
        excel_bytes = generate_deficiency_excel(deficiencies, {"total_count": 2})
        workbook = load_workbook(BytesIO(excel_bytes))
        summary_sheet = workbook["Deficiency Summary"]

        self.assertEqual(summary_sheet.auto_filter.ref, "A1:L3")
        self.assertTrue(summary_sheet["A2"].fill.start_color.rgb.endswith("FADBD8"))
        self.assertTrue(summary_sheet["A3"].fill.start_color.rgb.endswith("F8F9FA"))

    def test_feat_rpt_002_happy_path_generator_sets_car_status_sheet_filter_and_filters_sheet_totals(self):
        """PRD FEAT-RPT-002: CAR status sheet should be filterable and filters sheet should include total count."""
        deficiencies = [
            {
                "def_code": "10101",
                "description": "Row one",
                "action_code": "30",
                "target_date": "2026-01-22",
                "is_cleared": False,
                "cleared_date": None,
                "inspection": {
                    "vessel_name": "MV Example",
                    "inspection_type": "PSC",
                    "inspection_date": "2026-01-15",
                    "port_place": "Singapore",
                    "is_detention": False,
                },
                "car": {
                    "car_number": "PSC-2026-001",
                    "status_display": "Submitted",
                    "root_cause_summary": "R" * 60,
                    "target_date": "2026-01-25",
                    "evidence_count": 1,
                    "actions_count": 2,
                },
            },
        ]
        excel_bytes = generate_deficiency_excel(
            deficiencies,
            {"inspection_type": "PSC", "total_count": 1},
        )
        workbook = load_workbook(BytesIO(excel_bytes))
        car_status_sheet = workbook["CAR Status"]
        filter_sheet = workbook["Applied Filters"]

        self.assertEqual(car_status_sheet.auto_filter.ref, "A1:I2")

        labels = [filter_sheet[f"A{row}"].value for row in range(1, 20)]
        self.assertIn("Total Records", labels)
        total_row = labels.index("Total Records") + 1
        self.assertEqual(filter_sheet[f"B{total_row}"].value, 1)
