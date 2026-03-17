"""
Phase 3 DefIntel checklist preview/export views.
"""

import io
import logging
import uuid
from collections import defaultdict
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path

from django.http import HttpResponse
from django.db import connection
from rest_framework import serializers, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.models import RoleCodes, VesselData
from core.vessel_access import apply_office_vessel_filter

from .defintel_access import can_access_defintel_reports
from .deficiency_models import Deficiency
from .defintel_models import OpenSourceDeficiencyRecord, OpenSourceImportRun
from .defintel_views import (
    _normalize_action_code,
    _normalize_def_code,
    _normalize_text,
)

logger = logging.getLogger(__name__)

SCOPE_VESSEL = 'VESSEL'
SCOPE_FLEET = 'FLEET'
SCOPE_INSPECTOR = 'INSPECTOR'
SCOPE_FILTER_COMBINED = 'FILTER_COMBINED'
SCOPE_CHOICES = [
    SCOPE_VESSEL,
    SCOPE_FLEET,
    SCOPE_INSPECTOR,
    SCOPE_FILTER_COMBINED,
]

CHECKLIST_TOTAL_COLUMNS = 12
CHECKLIST_TABLE_HEADER_ROW = 8
CHECKLIST_DATA_START_ROW = 9
CHECKLIST_TABLE_HEADERS = [
    'Sr No.',
    'Significance',
    'DEF Code',
    'Deficiency / Description',
    'Action Code',
    'Action Label',
    'MoU',
    'Port',
    'Country',
    'Ship Status (OK / Not OK / N.A.)',
    'Ship Remarks / Evidence',
    'Office Comments',
]
CHECKLIST_HEADER_MERGE_END_COLUMN = 9
CHECKLIST_LEGEND_TITLE = (
    'Deficiency Code Significance Designation listing '
    '(in order of Importance with corresponding Color):'
)
CHECKLIST_LEGEND_ORDER = ['DETXX', 'CIC', 'CSP', 'CMP', 'BLANK', 'S10', 'C10']
DETENTION_ACTION_CODE = '30'
DETENTION_SIGNIFICANCE_CODE = 'DET1'
DETENTION_STYLE_FALLBACK_KEY = 'DETXX'


class ChecklistFiltersSerializer(serializers.Serializer):
    def_code = serializers.ListField(
        child=serializers.CharField(),
        required=False,
        default=list,
    )
    action_code = serializers.ListField(
        child=serializers.CharField(),
        required=False,
        default=list,
    )
    mou = serializers.ListField(
        child=serializers.CharField(),
        required=False,
        default=list,
    )
    port = serializers.ListField(
        child=serializers.CharField(),
        required=False,
        default=list,
    )
    country = serializers.ListField(
        child=serializers.CharField(),
        required=False,
        default=list,
    )


class VesselPrepRequestSerializer(serializers.Serializer):
    scope_mode = serializers.ChoiceField(choices=SCOPE_CHOICES)
    vessel_id = serializers.UUIDField(required=False)
    vessel_name = serializers.CharField(required=False, allow_blank=True, max_length=255)
    inspector_name = serializers.CharField(required=False)
    filters = ChecklistFiltersSerializer(required=False, default=dict)
    date_from = serializers.DateField(required=False)
    date_to = serializers.DateField(required=False)
    dedup = serializers.BooleanField(required=False, default=True)

    def validate(self, attrs):
        scope_mode = attrs['scope_mode']
        date_from = attrs.get('date_from')
        date_to = attrs.get('date_to')

        if scope_mode == SCOPE_VESSEL and not attrs.get('vessel_id'):
            raise serializers.ValidationError({'vessel_id': 'Required when scope is VESSEL'})

        if scope_mode == SCOPE_INSPECTOR:
            try:
                _normalize_inspector_name(attrs.get('inspector_name'))
            except ValueError:
                raise serializers.ValidationError(
                    {'inspector_name': 'inspector_name is required for INSPECTOR scope.'}
                )

        if date_from and date_to and date_from > date_to:
            raise serializers.ValidationError({'date_to': 'date_to must be greater than or equal to date_from.'})

        return attrs


class ChecklistBuildError(Exception):
    """Raised when checklist data cannot be built."""

    def __init__(self, error_code, message, http_status=status.HTTP_400_BAD_REQUEST):
        super().__init__(message)
        self.error_code = error_code
        self.message = message
        self.http_status = http_status


def _normalize_filter_values(raw_filters):
    raw_filters = raw_filters or {}

    def normalized_upper_list(values):
        result = []
        for value in values or []:
            text = _normalize_text(value, field_name='filter_value')
            if text and text not in result:
                result.append(text)
        return result

    def normalized_def_codes(values):
        result = []
        for value in values or []:
            code = _normalize_def_code(value)
            if code not in result:
                result.append(code)
        return result

    def normalized_action_codes(values):
        result = []
        for value in values or []:
            code = _normalize_action_code(value)
            if code not in result:
                result.append(code)
        return result

    return {
        'def_code': normalized_def_codes(raw_filters.get('def_code')),
        'action_code': normalized_action_codes(raw_filters.get('action_code')),
        'mou': normalized_upper_list(raw_filters.get('mou')),
        'port': normalized_upper_list(raw_filters.get('port')),
        'country': normalized_upper_list(raw_filters.get('country')),
    }


def _record_matches_text_filters(record, filters):
    if filters['mou'] and record['mou'] not in filters['mou']:
        return False
    if filters['port'] and record['port'] not in filters['port']:
        return False
    if filters['country'] and record['country'] not in filters['country']:
        return False
    return True


def _normalize_inspector_name(value):
    if value is None:
        raise ValueError('inspector_name is required')

    normalized = ' '.join(str(value).strip().split()).lower()
    if not normalized:
        raise ValueError('inspector_name is required')
    return normalized


def _apply_user_access_scope(queryset, user):
    if user.user_type == 'VESSEL':
        queryset = queryset.filter(inspection__vessel_id=user.vessel_id)
        if user.role == RoleCodes.VESSEL_CREW:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(assigned_crew_id=user.crew_id) |
                Q(assigned_crew_id=user.id)
            )
        return queryset
    return apply_office_vessel_filter(queryset, user, 'inspection__vessel_id')


def _build_internal_records(*, user, scope_mode, vessel_id, inspector_name, filters, date_from, date_to):
    queryset = (
        Deficiency.objects.filter(
            is_deleted=False,
            inspection__is_deleted=False,
        )
        .select_related('inspection')
        .order_by('inspection__inspection_date', 'inspection_id', 'sequence_no')
    )
    queryset = _apply_user_access_scope(queryset, user)

    if scope_mode == SCOPE_VESSEL:
        queryset = queryset.filter(inspection__vessel_id=vessel_id)
    elif scope_mode == SCOPE_INSPECTOR:
        queryset = queryset.filter(inspection__inspector_name__isnull=False)

    normalized_inspector = None
    if scope_mode == SCOPE_INSPECTOR:
        normalized_inspector = _normalize_inspector_name(inspector_name)

    if date_from:
        queryset = queryset.filter(inspection__inspection_date__gte=date_from)
    if date_to:
        queryset = queryset.filter(inspection__inspection_date__lte=date_to)

    if filters['def_code']:
        queryset = queryset.filter(def_code_id__in=filters['def_code'])
    if filters['action_code']:
        queryset = queryset.filter(action_code_id__in=filters['action_code'])

    records = []
    invalid_rows = 0

    for deficiency in queryset.iterator():
        inspection = deficiency.inspection
        try:
            if scope_mode == SCOPE_INSPECTOR:
                current = _normalize_inspector_name(inspection.inspector_name)
                if current != normalized_inspector:
                    continue

            def_code = _normalize_def_code(deficiency.def_code_id or deficiency.def_code)

            if deficiency.action_code_id is not None:
                action_code = _normalize_action_code(deficiency.action_code_id)
            elif deficiency.action_code:
                action_code = _normalize_action_code(deficiency.action_code)
            else:
                raise ValueError('action_code is required')

            mou = _normalize_text(inspection.mou_id, field_name='mou')
            port = _normalize_text(inspection.port_place, field_name='port')
            country = (
                _normalize_text(inspection.country, field_name='country', optional=True)
                or ''
            )
            inspection_date = inspection.inspection_date

            record = {
                'source': 'internal',
                'year': inspection_date.year if inspection_date else None,
                'def_code': def_code,
                'action_code': action_code,
                'mou': mou,
                'port': port,
                'country': country,
                'date': inspection_date,
                'description': (deficiency.description or '').strip(),
            }

            if _record_matches_text_filters(record, filters):
                records.append(record)
        except ValueError:
            invalid_rows += 1

    return records, invalid_rows


def _build_opensource_records(*, filters, date_from, date_to):
    queryset = OpenSourceDeficiencyRecord.objects.all().order_by('year', 'id')

    if filters['def_code']:
        queryset = queryset.filter(def_code_norm__in=filters['def_code'])
    if filters['action_code']:
        queryset = queryset.filter(action_code_norm__in=filters['action_code'])
    if date_from:
        queryset = queryset.filter(year__gte=date_from.year)
    if date_to:
        queryset = queryset.filter(year__lte=date_to.year)

    records = []
    for row in queryset.iterator():
        record = {
            'source': 'opensource',
            'year': row.year,
            'def_code': row.def_code_norm,
            'action_code': row.action_code_norm,
            'mou': row.mou_norm,
            'port': row.port_norm,
            'country': row.country_norm or '',
            'date': None,
            'description': (row.description_raw or '').strip(),
        }

        if _record_matches_text_filters(record, filters):
            records.append(record)

    return records


def _apply_post_merge_dedup(records, dedup_enabled):
    stats = {
        'dedup_enabled': bool(dedup_enabled),
        'input_rows': len(records),
        'removed_rows': 0,
        'output_rows': len(records),
    }
    if not dedup_enabled:
        return records, stats

    seen = set()
    deduped = []

    for record in records:
        # Post-merge dedup uses the strict identity plus source split for count integrity.
        key = (
            record['source'],
            record['year'],
            record['def_code'],
            record['action_code'],
            record['port'],
            record['mou'],
        )
        if key in seen:
            stats['removed_rows'] += 1
            continue
        seen.add(key)
        deduped.append(record)

    stats['output_rows'] = len(deduped)
    return deduped, stats


def _aggregate_checklist_rows(records):
    grouped = defaultdict(
        lambda: {
            'def_code': '',
            'action_code': '',
            'mou': '',
            'port': '',
            'country': '',
            'occurrence_count_total': 0,
            'occurrence_count_internal': 0,
            'occurrence_count_opensource': 0,
            'last_seen_date': None,
            'example_description_internal': '',
            'example_description_opensource': '',
        }
    )

    for record in records:
        key = (
            record['def_code'],
            str(record['action_code']),
            record['mou'],
            record['port'],
            record['country'],
        )
        bucket = grouped[key]

        bucket['def_code'] = record['def_code']
        bucket['action_code'] = str(record['action_code'])
        bucket['mou'] = record['mou']
        bucket['port'] = record['port']
        bucket['country'] = record['country']

        bucket['occurrence_count_total'] += 1
        if record['source'] == 'internal':
            bucket['occurrence_count_internal'] += 1
            if record['description'] and not bucket['example_description_internal']:
                bucket['example_description_internal'] = record['description']
        else:
            bucket['occurrence_count_opensource'] += 1
            if record['description'] and not bucket['example_description_opensource']:
                bucket['example_description_opensource'] = record['description']

        if record['date'] and (
            bucket['last_seen_date'] is None or record['date'] > bucket['last_seen_date']
        ):
            bucket['last_seen_date'] = record['date']

    rows = []
    for bucket in grouped.values():
        rows.append(
            {
                'def_code': bucket['def_code'],
                'action_code': bucket['action_code'],
                'mou': bucket['mou'],
                'port': bucket['port'],
                'country': bucket['country'],
                'occurrence_count_total': bucket['occurrence_count_total'],
                'occurrence_count_internal': bucket['occurrence_count_internal'],
                'occurrence_count_opensource': bucket['occurrence_count_opensource'],
                'last_seen_date': (
                    bucket['last_seen_date'].isoformat() if bucket['last_seen_date'] else None
                ),
                'example_description': (
                    bucket['example_description_internal']
                    or bucket['example_description_opensource']
                ),
            }
        )

    rows.sort(
        key=lambda row: (
            -row['occurrence_count_total'],
            -row['occurrence_count_internal'],
            row['def_code'],
            row['action_code'],
            row['mou'],
            row['port'],
            row['country'],
        )
    )
    return rows


def _build_checklist_payload(validated_data, user):
    scope_mode = validated_data['scope_mode']
    vessel_id = validated_data.get('vessel_id')
    inspector_name = validated_data.get('inspector_name')
    date_from = validated_data.get('date_from')
    date_to = validated_data.get('date_to')
    dedup = validated_data.get('dedup', True)

    try:
        filters = _normalize_filter_values(validated_data.get('filters', {}))
    except ValueError as exc:
        raise ChecklistBuildError('VALIDATION_ERROR', str(exc))

    if scope_mode != SCOPE_FILTER_COMBINED:
        filters = {
            'def_code': [],
            'action_code': [],
            'mou': [],
            'port': [],
            'country': [],
        }

    if scope_mode == SCOPE_FILTER_COMBINED and not OpenSourceImportRun.objects.exists():
        raise ChecklistBuildError(
            'IMPORT_REQUIRED',
            'OpenSource import required before using FILTER_COMBINED scope.',
        )

    internal_records, internal_invalid_rows = _build_internal_records(
        user=user,
        scope_mode=scope_mode,
        vessel_id=vessel_id,
        inspector_name=inspector_name,
        filters=filters,
        date_from=date_from,
        date_to=date_to,
    )

    opensource_records = []
    if scope_mode == SCOPE_FILTER_COMBINED:
        opensource_records = _build_opensource_records(
            filters=filters,
            date_from=date_from,
            date_to=date_to,
        )

    merged_records = internal_records + opensource_records
    deduped_records, dedup_stats = _apply_post_merge_dedup(merged_records, dedup)
    checklist_rows = _aggregate_checklist_rows(deduped_records)

    occurrence_internal = sum(row['occurrence_count_internal'] for row in checklist_rows)
    occurrence_opensource = sum(row['occurrence_count_opensource'] for row in checklist_rows)
    occurrence_total = sum(row['occurrence_count_total'] for row in checklist_rows)
    unique_def_codes = len({row['def_code'] for row in checklist_rows if row.get('def_code')})
    unique_action_codes = len(
        {str(row['action_code']) for row in checklist_rows if row.get('action_code') not in (None, '')}
    )

    payload = {
        'scope_mode': scope_mode,
        'date_from': date_from.isoformat() if date_from else None,
        'date_to': date_to.isoformat() if date_to else None,
        'filters': filters,
        'dedup': bool(dedup),
        'rows': checklist_rows,
        'summary': {
            'row_count': len(checklist_rows),
            'occurrence_count_total': occurrence_total,
            'occurrence_count_internal': occurrence_internal,
            'occurrence_count_opensource': occurrence_opensource,
            'unique_def_code_count': unique_def_codes,
            'unique_action_code_count': unique_action_codes,
            'internal_invalid_rows': internal_invalid_rows,
            'input_internal_rows': len(internal_records),
            'input_opensource_rows': len(opensource_records),
            'dedup_stats': dedup_stats,
            'last_seen_rule': (
                'Uses max internal inspection_date. OpenSource rows have no date and do not increase last_seen_date.'
            ),
        },
    }
    return payload


def _build_export_metadata(validated_data, payload, user):
    generated_on = datetime.utcnow().date()
    validity_until = generated_on + timedelta(days=30)

    scope_mode = payload.get('scope_mode') or validated_data.get('scope_mode')
    request_vessel_id = validated_data.get('vessel_id') if scope_mode == SCOPE_VESSEL else None
    request_vessel_name = (
        str(validated_data.get('vessel_name') or '').strip()
        if scope_mode == SCOPE_VESSEL
        else ''
    )
    user_vessel_id = getattr(user, 'vessel_id', None)
    candidate_vessel_id = request_vessel_id or user_vessel_id
    if scope_mode == SCOPE_VESSEL and not candidate_vessel_id:
        logger.debug('Vessel id not provided for vessel-prep export metadata (scope_mode=VESSEL).')
    vessel_label = request_vessel_name or _lookup_vessel_name(candidate_vessel_id)
    if not vessel_label and str(getattr(user, 'user_type', '') or '').upper() == 'VESSEL':
        vessel_label = str(getattr(user, 'vessel_name', '') or '').strip()
    if not vessel_label:
        vessel_label = '(Not specified)'

    filters = payload.get('filters') or validated_data.get('filters') or {}
    requested_ports = filters.get('port') or []
    generated_for_port = ', '.join(requested_ports) if requested_ports else 'All Ports'
    date_from_value = validated_data.get('date_from')
    date_to_value = validated_data.get('date_to')

    return {
        'scope_mode': scope_mode,
        'vessel_label': vessel_label,
        'generated_for_port': generated_for_port,
        'date_from': date_from_value.isoformat() if date_from_value else '',
        'date_to': date_to_value.isoformat() if date_to_value else '',
        'generated_on': generated_on.isoformat(),
        'valid_until': validity_until.isoformat(),
    }


def _lookup_vessel_name(vessel_id):
    if not vessel_id:
        return ''

    normalized_vessel_id = _normalize_uuid_text(vessel_id)
    if not normalized_vessel_id:
        return ''

    try:
        vessel_name = (
            VesselData.objects.filter(
                id=normalized_vessel_id,
                is_active=True,
                is_deleted=False,
            )
            .values_list('vesselName', flat=True)
            .first()
        )
        if vessel_name:
            return str(vessel_name).strip()
    except Exception as exc:
        logger.debug('Vessel lookup via ORM failed for %s: %s', normalized_vessel_id, exc)

    # SQL Server fallback: force explicit uniqueidentifier cast.
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT TOP 1 vesselName
                FROM VesselData
                WHERE id = CAST(%s AS uniqueidentifier)
                  AND is_active = 1
                  AND is_deleted = 0
                """,
                [normalized_vessel_id],
            )
            row = cursor.fetchone()
            if row and row[0]:
                return str(row[0]).strip()
    except Exception as exc:
        logger.debug('Vessel lookup via raw SQL failed for %s: %s', normalized_vessel_id, exc)

    return ''


def _normalize_uuid_text(value):
    """Normalize UUID-like values to canonical hyphenated lower-case text."""
    if value is None:
        return ''

    text = str(value).strip()
    if not text:
        return ''

    try:
        return str(uuid.UUID(text))
    except (ValueError, TypeError, AttributeError):
        compact = text.replace('-', '')
        if len(compact) == 32:
            try:
                return str(uuid.UUID(compact))
            except (ValueError, TypeError, AttributeError):
                return ''
        return ''



def _resolve_template_path():
    project_root = Path(__file__).resolve().parents[3]
    candidates = [
        project_root / 'docs' / 'templates' / 'Preparation Checklist.xlsx',
        project_root / 'Docs' / 'templates' / 'Preparation Checklist.xlsx',
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise ChecklistBuildError(
        'MISSING_TEMPLATE',
        f"Missing required template: {candidates[0]}",
        http_status=status.HTTP_500_INTERNAL_SERVER_ERROR,
    )


def _find_checklist_footer_row(checklist_sheet):
    from openpyxl.cell.cell import MergedCell

    footer_markers = (
        'Deficiency Code Significance Designation listing',
        'Deficiency Code Significance',
    )
    for row_idx in range(1, checklist_sheet.max_row + 1):
        value = checklist_sheet.cell(row=row_idx, column=1).value
        if isinstance(value, str) and any(value.startswith(marker) for marker in footer_markers):
            return row_idx

    merged_rows = sorted(
        rng.min_row
        for rng in checklist_sheet.merged_cells.ranges
        if rng.min_row >= 5 and rng.min_col <= 2 <= rng.max_col
    )
    if merged_rows:
        return merged_rows[0]

    for row_idx in range(5, checklist_sheet.max_row + 1):
        if isinstance(checklist_sheet.cell(row=row_idx, column=2), MergedCell):
            return row_idx

    return checklist_sheet.max_row + 1


def _copy_template_row_style(checklist_sheet, *, source_row, start_row, row_count, column_count=9):
    if row_count <= 0:
        return
    source_max_column = max(1, checklist_sheet.max_column)
    for offset in range(row_count):
        target_row = start_row + offset
        for col_idx in range(1, column_count + 1):
            source_col_idx = min(col_idx, source_max_column)
            checklist_sheet.cell(row=target_row, column=col_idx)._style = copy(
                checklist_sheet.cell(row=source_row, column=source_col_idx)._style
            )
        source_height = checklist_sheet.row_dimensions[source_row].height
        if source_height is not None:
            checklist_sheet.row_dimensions[target_row].height = source_height


def _normalize_designation_key(value):
    normalized = str(value or '').strip().upper()
    if not normalized:
        return ''
    if normalized == 'BLANK':
        return 'BLANK'
    return normalized


def _extract_template_significance_by_def_code(checklist_sheet, footer_row):
    mapping = {}
    for row_idx in range(5, footer_row):
        raw_def_code = checklist_sheet.cell(row=row_idx, column=3).value
        if raw_def_code in (None, ''):
            continue
        def_code = str(raw_def_code).strip()
        if not def_code:
            continue

        designation = str(checklist_sheet.cell(row=row_idx, column=2).value or '').strip()
        if designation:
            mapping[def_code] = designation
    return mapping


def _read_bottom_legend_entries(checklist_sheet, legend_title_row):
    entries = []
    for row_idx in range(legend_title_row + 1, legend_title_row + 30):
        designation = checklist_sheet.cell(row=row_idx, column=1).value
        description = checklist_sheet.cell(row=row_idx, column=2).value

        if designation in (None, '') and description in (None, ''):
            if entries:
                break
            continue
        if designation in (None, ''):
            continue

        designation_text = str(designation).strip()
        description_text = str(description or '').strip()
        entries.append(
            {
                'designation': designation_text,
                'description': description_text,
                'fill': copy(checklist_sheet.cell(row=row_idx, column=1).fill),
                'font': copy(checklist_sheet.cell(row=row_idx, column=1).font),
                'designation_style': copy(checklist_sheet.cell(row=row_idx, column=1)._style),
                'description_style': copy(checklist_sheet.cell(row=row_idx, column=2)._style),
            }
        )
    return entries


def _read_bottom_notes_block(checklist_sheet, legend_title_row):
    notes_title_row = None
    for row_idx in range(legend_title_row + 1, legend_title_row + 60):
        value = checklist_sheet.cell(row=row_idx, column=1).value
        if isinstance(value, str) and value.strip().upper().startswith('IMPORTANT NOTES'):
            notes_title_row = row_idx
            break

    if notes_title_row is None:
        return {
            'title': 'IMPORTANT NOTES:',
            'title_style': None,
            'lines': [],
        }

    lines = []
    for row_idx in range(notes_title_row + 1, notes_title_row + 30):
        value = checklist_sheet.cell(row=row_idx, column=1).value
        if value in (None, ''):
            if lines:
                break
            continue
        lines.append(
            {
                'text': str(value).strip(),
                'style': copy(checklist_sheet.cell(row=row_idx, column=1)._style),
            }
        )

    return {
        'title': str(checklist_sheet.cell(row=notes_title_row, column=1).value or 'IMPORTANT NOTES:').strip(),
        'title_style': copy(checklist_sheet.cell(row=notes_title_row, column=1)._style),
        'lines': lines,
    }


def _order_legend_entries(entries):
    entry_by_key = {
        _normalize_designation_key(entry.get('designation')): entry
        for entry in entries
    }
    ordered_entries = []
    for legend_key in CHECKLIST_LEGEND_ORDER:
        entry = entry_by_key.get(legend_key)
        if entry:
            ordered_entries.append(entry)

    for entry in entries:
        entry_key = _normalize_designation_key(entry.get('designation'))
        if entry_key not in CHECKLIST_LEGEND_ORDER:
            ordered_entries.append(entry)

    return ordered_entries


def _resolve_row_significance(row, template_significance_by_def_code):
    if str(row.get('action_code') or '').strip() == DETENTION_ACTION_CODE:
        return DETENTION_SIGNIFICANCE_CODE

    def_code = str(row.get('def_code') or '').strip()
    if def_code and def_code in template_significance_by_def_code:
        return template_significance_by_def_code[def_code]
    return None


def _resolve_action_label(action_code):
    return 'DETENTION' if str(action_code).strip() == DETENTION_ACTION_CODE else ''


def _set_checklist_cell_value_or_raise(
    checklist_sheet,
    *,
    row_number,
    column_number,
    value,
    footer_row,
    dataset_length,
):
    from openpyxl.cell.cell import MergedCell

    cell = checklist_sheet.cell(row=row_number, column=column_number)
    if isinstance(cell, MergedCell):
        raise ChecklistBuildError(
            'EXPORT_TEMPLATE_ERROR',
            (
                'Checklist export write attempted on merged/protected template area '
                f'(row_number={row_number}, column={column_number}, '
                f'footer_row={footer_row}, dataset_length={dataset_length}).'
            ),
            http_status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
    cell.value = value


def _build_checklist_export_workbook(payload):
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = load_workbook(_resolve_template_path())
    checklist_sheet = workbook.active
    checklist_sheet.title = 'Preparation Checklist'

    # Capture base styles from template before clearing body values.
    template_header_style = copy(checklist_sheet.cell(row=4, column=1)._style)
    template_data_styles = {
        col_idx: copy(checklist_sheet.cell(row=5, column=min(col_idx, 9))._style)
        for col_idx in range(1, CHECKLIST_TOTAL_COLUMNS + 1)
    }

    template_footer_row = _find_checklist_footer_row(checklist_sheet)
    template_significance_by_def_code = _extract_template_significance_by_def_code(checklist_sheet, template_footer_row)
    table_header_row = CHECKLIST_TABLE_HEADER_ROW
    table_data_start_row = CHECKLIST_DATA_START_ROW
    required_last_data_row = (
        table_data_start_row + len(payload['rows']) - 1
        if payload['rows']
        else table_header_row
    )
    footer_start_row = required_last_data_row + 5

    inserted_row_count = 0
    if footer_start_row > template_footer_row:
        inserted_row_count = footer_start_row - template_footer_row
        checklist_sheet.insert_rows(
            template_footer_row,
            amount=inserted_row_count,
        )
        _copy_template_row_style(
            checklist_sheet,
            source_row=max(5, template_footer_row - 1),
            start_row=template_footer_row,
            row_count=inserted_row_count,
            column_count=CHECKLIST_TOTAL_COLUMNS,
        )
    source_footer_row = template_footer_row + inserted_row_count
    effective_footer_row = footer_start_row
    bottom_legend_entries = _order_legend_entries(
        _read_bottom_legend_entries(checklist_sheet, source_footer_row)
    )
    notes_block = _read_bottom_notes_block(checklist_sheet, source_footer_row)
    designation_style_map = {
        _normalize_designation_key(entry['designation']): entry
        for entry in bottom_legend_entries
    }
    det_style_entry = (
        designation_style_map.get(DETENTION_SIGNIFICANCE_CODE)
        or designation_style_map.get(DETENTION_STYLE_FALLBACK_KEY)
    )
    if det_style_entry and DETENTION_SIGNIFICANCE_CODE not in designation_style_map:
        designation_style_map[DETENTION_SIGNIFICANCE_CODE] = det_style_entry
    no_fill = PatternFill(fill_type=None)

    def clear_range(
        start_row,
        end_row,
        start_col=1,
        end_col=CHECKLIST_TOTAL_COLUMNS,
        excluded_rows=None,
        clear_fill=False,
        clear_border=False,
    ):
        if end_row < start_row:
            return
        excluded_rows = excluded_rows or set()
        for row_idx in range(start_row, end_row + 1):
            if row_idx in excluded_rows:
                continue
            for col_idx in range(start_col, end_col + 1):
                cell = checklist_sheet.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                if clear_fill:
                    cell.fill = PatternFill(fill_type=None)
                if clear_border:
                    cell.border = Border()

    # Remove template artifacts from visible main area while preserving template merge/style cells.
    clear_range(
        1,
        289,
        1,
        CHECKLIST_TOTAL_COLUMNS,
        excluded_rows={1, 2, 3, 4, 5, 6, table_header_row},
        clear_fill=True,
        clear_border=True,
    )

    export_meta = payload.get('export_meta') or {}
    date_from = export_meta.get('date_from') or 'N/A'
    date_to = export_meta.get('date_to') or 'N/A'
    generated_on = export_meta.get('generated_on') or datetime.utcnow().date().isoformat()
    valid_until = export_meta.get('valid_until') or (
        datetime.utcnow().date() + timedelta(days=30)
    ).isoformat()
    vessel_label = export_meta.get('vessel_label') or 'N/A'
    generated_for_port = export_meta.get('generated_for_port') or 'All Ports'
    scope_mode = export_meta.get('scope_mode') or payload.get('scope_mode') or ''

    header_rows = [
        'PSC Preparation Checklist',
        f'Vessel: {vessel_label}',
        f'Generated for: Scope = {scope_mode}, Port = {generated_for_port}',
        f'Date range: {date_from} to {date_to}',
        f'Generated on: {generated_on}',
        f'Validity: {valid_until}',
    ]

    for row_idx in range(1, 7):
        for merged_range in list(checklist_sheet.merged_cells.ranges):
            if (
                merged_range.min_row <= row_idx <= merged_range.max_row
                and merged_range.min_col <= 1 <= merged_range.max_col
                and merged_range.min_col <= CHECKLIST_HEADER_MERGE_END_COLUMN <= merged_range.max_col
            ):
                checklist_sheet.unmerge_cells(str(merged_range))

    for row_idx, header_value in enumerate(header_rows, start=1):
        checklist_sheet.merge_cells(
            start_row=row_idx,
            start_column=1,
            end_row=row_idx,
            end_column=CHECKLIST_HEADER_MERGE_END_COLUMN,
        )
        header_cell = checklist_sheet.cell(row=row_idx, column=1, value=header_value)
        if row_idx == 1:
            header_cell.font = Font(bold=True, size=16)

    # Explicitly keep row 7 blank before the table header.
    clear_range(7, 7, 1, CHECKLIST_TOTAL_COLUMNS)

    for col_idx, header in enumerate(CHECKLIST_TABLE_HEADERS, start=1):
        cell = checklist_sheet.cell(row=table_header_row, column=col_idx, value=header)
        cell._style = copy(template_header_style)
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = 'center'
        cell.alignment = alignment

    wrap_text_columns = {4, 11, 12}
    for idx, row in enumerate(payload['rows'], start=1):
        row_number = table_data_start_row + idx - 1
        significance_value = _resolve_row_significance(row, template_significance_by_def_code)
        action_label = _resolve_action_label(row['action_code'])
        style_entry = None
        if significance_value not in (None, ''):
            significance_key = _normalize_designation_key(significance_value)
            style_entry = designation_style_map.get(significance_key)
            if not style_entry and significance_key.startswith('DET'):
                style_entry = det_style_entry

        row_values = [
            idx,
            significance_value or '',
            row['def_code'],
            row['example_description'] or '',
            row['action_code'],
            action_label,
            row['mou'],
            row['port'],
            row['country'] or '',
            '',
            '',
            '',
        ]

        for column_number, cell_value in enumerate(row_values, start=1):
            _set_checklist_cell_value_or_raise(
                checklist_sheet,
                row_number=row_number,
                column_number=column_number,
                value=cell_value,
                footer_row=effective_footer_row,
                dataset_length=len(payload['rows']),
            )
            cell = checklist_sheet.cell(row=row_number, column=column_number)
            cell._style = copy(template_data_styles[column_number])

            if column_number == 2:
                if style_entry:
                    cell.fill = copy(style_entry['fill'])
                    if style_entry.get('font') is not None:
                        cell.font = copy(style_entry['font'])
                else:
                    cell.fill = no_fill
            else:
                cell.fill = no_fill

            alignment = copy(cell.alignment)
            alignment.vertical = 'top'
            alignment.wrap_text = column_number in wrap_text_columns
            if column_number == 10:
                alignment.horizontal = 'center'
            cell.alignment = alignment

    if payload['rows']:
        last_data_row = table_data_start_row + len(payload['rows']) - 1
        status_validation = DataValidation(
            type='list',
            formula1='"OK,Not OK,N.A."',
            allow_blank=True,
        )
        checklist_sheet.add_data_validation(status_validation)
        status_validation.add(f'J{table_data_start_row}:J{last_data_row}')
    else:
        last_data_row = table_header_row

    clear_range(
        last_data_row + 1,
        effective_footer_row - 1,
        1,
        CHECKLIST_TOTAL_COLUMNS,
        clear_fill=True,
        clear_border=True,
    )

    checklist_sheet.freeze_panes = f'A{table_data_start_row}'
    checklist_sheet.auto_filter.ref = f'A{table_header_row}:L{table_header_row}'

    for col_idx, header in enumerate(CHECKLIST_TABLE_HEADERS, start=1):
        if col_idx == 1:
            checklist_sheet.column_dimensions[get_column_letter(col_idx)].width = 8
            continue

        max_length = len(str(header))
        if payload['rows']:
            for row_idx in range(table_data_start_row, last_data_row + 1):
                value = checklist_sheet.cell(row=row_idx, column=col_idx).value
                if value not in (None, ''):
                    max_length = max(max_length, len(str(value)))

        width = min(50, max(10, max_length + 2))
        if col_idx in (4, 11, 12):
            width = min(50, max(width, 30))
        checklist_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    if 'Input Summary' in workbook.sheetnames:
        input_summary_sheet = workbook['Input Summary']
    else:
        input_summary_sheet = workbook.create_sheet('Input Summary')
    clear_sheet_max_row = max(input_summary_sheet.max_row, 80)
    clear_sheet_max_col = max(input_summary_sheet.max_column, 6)
    for row_idx in range(1, clear_sheet_max_row + 1):
        for col_idx in range(1, clear_sheet_max_col + 1):
            input_summary_sheet.cell(row=row_idx, column=col_idx, value=None)

    summary = payload['summary']
    filter_summary = payload.get('filters', {})
    system_summary_rows = [
        ('occurrence_summary', 'Occurrence Summary'),
        ('occurrence_count_total', summary['occurrence_count_total']),
        ('occurrence_count_internal', summary['occurrence_count_internal']),
        ('occurrence_count_opensource', summary['occurrence_count_opensource']),
        ('unique_def_code_count', summary.get('unique_def_code_count', 0)),
        ('unique_action_code_count', summary.get('unique_action_code_count', 0)),
        ('input_payload', 'Input Payload'),
        ('scope_mode', payload['scope_mode']),
        ('date_from', payload.get('date_from') or ''),
        ('date_to', payload.get('date_to') or ''),
        ('dedup_enabled', str(payload.get('dedup'))),
        ('filter_def_code', ', '.join(filter_summary.get('def_code', []))),
        ('filter_action_code', ', '.join(str(v) for v in filter_summary.get('action_code', []))),
        ('filter_mou', ', '.join(filter_summary.get('mou', []))),
        ('filter_port', ', '.join(filter_summary.get('port', []))),
        ('filter_country', ', '.join(filter_summary.get('country', []))),
        ('row_count', summary['row_count']),
        ('occurrence_count_total', summary['occurrence_count_total']),
        ('occurrence_count_internal', summary['occurrence_count_internal']),
        ('occurrence_count_opensource', summary['occurrence_count_opensource']),
        ('unique_def_code_count', summary.get('unique_def_code_count', 0)),
        ('unique_action_code_count', summary.get('unique_action_code_count', 0)),
        ('internal_invalid_rows', summary['internal_invalid_rows']),
        ('input_internal_rows', summary['input_internal_rows']),
        ('input_opensource_rows', summary['input_opensource_rows']),
        ('dedup_input_rows', summary['dedup_stats']['input_rows']),
        ('dedup_removed_rows', summary['dedup_stats']['removed_rows']),
        ('dedup_output_rows', summary['dedup_stats']['output_rows']),
        ('last_seen_rule', summary['last_seen_rule']),
    ]
    for row_idx, (key, value) in enumerate(system_summary_rows, start=1):
        input_summary_sheet.cell(row=row_idx, column=1, value=key)
        input_summary_sheet.cell(row=row_idx, column=2, value=str(value))
    input_summary_sheet.sheet_state = 'hidden'

    if 'System Summary' in workbook.sheetnames:
        del workbook['System Summary']
    if 'Legend' in workbook.sheetnames:
        del workbook['Legend']

    footer_title_cell = checklist_sheet.cell(row=source_footer_row, column=1)
    footer_title_style = copy(footer_title_cell._style)
    clear_range(
        source_footer_row,
        source_footer_row + 80,
        1,
        CHECKLIST_TOTAL_COLUMNS,
        clear_fill=True,
        clear_border=True,
    )
    clear_range(
        effective_footer_row,
        effective_footer_row + 80,
        1,
        CHECKLIST_TOTAL_COLUMNS,
        clear_fill=True,
        clear_border=True,
    )

    footer_title_cell = checklist_sheet.cell(row=effective_footer_row, column=1, value=CHECKLIST_LEGEND_TITLE)
    footer_title_cell._style = footer_title_style
    footer_title_cell.font = Font(bold=True)

    legend_row_start = effective_footer_row + 1
    for offset, entry in enumerate(bottom_legend_entries):
        row_idx = legend_row_start + offset
        designation_cell = checklist_sheet.cell(row=row_idx, column=1, value=entry['designation'])
        designation_cell._style = copy(entry['designation_style'])
        designation_cell.fill = copy(entry['fill'])
        designation_cell.font = copy(entry['font'])
        description_cell = checklist_sheet.cell(row=row_idx, column=2, value=entry['description'])
        description_cell._style = copy(entry['description_style'])

    notes_title_row = legend_row_start + len(bottom_legend_entries) + 1
    notes_title_cell = checklist_sheet.cell(
        row=notes_title_row,
        column=1,
        value=notes_block.get('title') or 'IMPORTANT NOTES:',
    )
    if notes_block.get('title_style') is not None:
        notes_title_cell._style = copy(notes_block['title_style'])
    else:
        notes_title_cell.font = Font(bold=True)

    for offset, line in enumerate(notes_block.get('lines') or [], start=1):
        note_cell = checklist_sheet.cell(
            row=notes_title_row + offset,
            column=1,
            value=line['text'],
        )
        if line.get('style') is not None:
            note_cell._style = copy(line['style'])

    # Remove leftover template headers from non-operational columns.
    for col_idx in range(CHECKLIST_TOTAL_COLUMNS + 1, checklist_sheet.max_column + 1):
        cell = checklist_sheet.cell(row=table_header_row, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None

    workbook.active = workbook.index(checklist_sheet)
    if checklist_sheet.sheet_view.selection:
        checklist_sheet.sheet_view.selection[0].activeCell = 'A1'
        checklist_sheet.sheet_view.selection[0].sqref = 'A1'

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


class VesselPrepPreviewView(APIView):
    """
    POST /api/psc/reports/vessel-prep/preview/
    """

    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        if not can_access_defintel_reports(request.user):
            return Response(
                {
                    'error': 'FORBIDDEN',
                    'message': 'DefIntel checklist is available to Office users and vessel ranks Master/CO/CE/2E only.',
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = VesselPrepRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Invalid request payload.',
                    'details': serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            payload = _build_checklist_payload(serializer.validated_data, request.user)
        except ChecklistBuildError as exc:
            return Response(
                {'error': exc.error_code, 'message': exc.message},
                status=exc.http_status,
            )

        return Response(
            {
                'data': payload,
                'message': 'Vessel preparation checklist preview generated successfully.',
            },
            status=status.HTTP_200_OK,
        )


class VesselPrepExportView(APIView):
    """
    POST /api/psc/reports/vessel-prep/export/
    """

    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        if not can_access_defintel_reports(request.user):
            return Response(
                {
                    'error': 'FORBIDDEN',
                    'message': 'DefIntel checklist is available to Office users and vessel ranks Master/CO/CE/2E only.',
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = VesselPrepRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Invalid request payload.',
                    'details': serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            payload = _build_checklist_payload(serializer.validated_data, request.user)
            payload['export_meta'] = _build_export_metadata(
                serializer.validated_data,
                payload,
                request.user,
            )
            workbook_bytes = _build_checklist_export_workbook(payload)
        except ChecklistBuildError as exc:
            return Response(
                {'error': exc.error_code, 'message': exc.message},
                status=exc.http_status,
            )

        filename = f"Vessel_Preparation_Checklist_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
