"""
Phase 2 tests for OpenSource monthly import endpoint.
"""

import hashlib
import re
import uuid
from datetime import date
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.db.utils import OperationalError, ProgrammingError
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APIRequestFactory, force_authenticate

from apps.accounts.models import RoleCodes, VesselData
from apps.inspection.deficiency_models import Deficiency
from apps.inspection.defintel_checklist import VesselPrepExportView, VesselPrepPreviewView
from apps.inspection.defintel_models import OpenSourceDeficiencyRecord, OpenSourceImportRun
from apps.inspection.defintel_views import OpenSourceImportView
from apps.inspection.models import Inspection


def make_vessel_master_user():
    return SimpleNamespace(
        id='vm-1',
        role=RoleCodes.VESSEL_MASTER,
        user_type='VESSEL',
        vessel_id='00000000-0000-0000-0000-000000000001',
        display_name='Vessel Master',
        username='vessel_master',
        rank='Master',
        employee_id=None,
        crew_id='CRW001',
        full_name='Vessel Master',
        is_active=True,
        pk='vm-1',
        is_anonymous=False,
        is_staff=False,
        is_superuser=False,
        groups=[],
        user_permissions=[],
        is_authenticated=True,
    )


def make_office_user():
    return SimpleNamespace(
        id='office-1',
        role=RoleCodes.OFFICE_PIC,
        user_type='OFFICE',
        vessel_id=None,
        display_name='Office PIC',
        username='office_pic',
        rank=None,
        employee_id='EMP001',
        crew_id=None,
        full_name='Office PIC',
        is_active=True,
        pk='office-1',
        is_anonymous=False,
        is_staff=False,
        is_superuser=False,
        groups=[],
        user_permissions=[],
        is_authenticated=True,
    )


def make_xlsx_file(rows, headers=None, name='opensource.xlsx'):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'DeficiencyData'

    worksheet.append(
        headers
        or [
            'No.',
            'Inspection Port',
            'Inspection Country',
            'MOU',
            'year',
            'InspectionDate',
            'ship_type',
            'Deficiency Code',
            'ActionCode',
            'Description',
        ]
    )
    for row in rows:
        worksheet.append(row)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return SimpleUploadedFile(
        name=name,
        content=stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


class TestOpenSourceImportPhase2(TestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.user = make_office_user()
        self.view = OpenSourceImportView.as_view()

    def _post_import(self, uploaded_file):
        request = self.factory.post(
            '/api/psc/reports/opensource/import/',
            {'file': uploaded_file},
            format='multipart',
        )
        force_authenticate(request, user=self.user)
        return self.view(request)

    def test_happy_path_import_inserts_normalized_rows(self):
        upload = make_xlsx_file(
            [
                [1, ' Singapore ', 'Singapore', 'tokyo', 2025, '2025-01-11', 'Bulk', '1101', '30', 'First row'],
                [2, 'Busan', 'Korea', 'TOKYO', 2025, '2025-01-12', 'Bulk', '20101', 17, 'Second row'],
            ]
        )

        response = self._post_import(upload)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(OpenSourceImportRun.objects.count(), 1)
        self.assertEqual(OpenSourceDeficiencyRecord.objects.count(), 2)

        payload = response.data['data']
        self.assertEqual(payload['total_rows'], 2)
        self.assertEqual(payload['valid_rows'], 2)
        self.assertEqual(payload['inserted_rows'], 2)
        self.assertEqual(payload['duplicate_rows'], 0)
        self.assertEqual(payload['invalid_rows'], 0)

        first = OpenSourceDeficiencyRecord.objects.order_by('port_norm').first()
        self.assertEqual(first.def_code_norm, '20101')
        self.assertEqual(first.action_code_norm, 17)

    def test_importing_same_file_twice_counts_duplicates(self):
        rows = [
            [1, 'Singapore', 'Singapore', 'TOKYO', 2025, '2025-01-11', 'Bulk', '1101', '30', 'First row'],
            [2, 'Busan', 'Korea', 'TOKYO', 2025, '2025-01-12', 'Bulk', '20101', 17, 'Second row'],
        ]

        first_response = self._post_import(make_xlsx_file(rows, name='monthly-1.xlsx'))
        second_response = self._post_import(make_xlsx_file(rows, name='monthly-1-repeat.xlsx'))

        self.assertEqual(first_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(second_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(OpenSourceDeficiencyRecord.objects.count(), 2)

        second_payload = second_response.data['data']
        self.assertEqual(second_payload['inserted_rows'], 0)
        self.assertEqual(second_payload['duplicate_rows'], 2)
        self.assertEqual(second_payload['valid_rows'], 2)

    def test_import_accepts_partial_invalid_rows_with_201(self):
        upload = make_xlsx_file(
            [
                [1, 'Singapore', 'Singapore', 'TOKYO', 2025, '2025-01-11', 'Bulk', '1101', '30', 'First row'],
                [2, ' singapore ', 'Singapore', 'tokyo', '2025', '2025-01-15', 'Bulk', '01101', 30, 'Duplicate row'],
                [3, 'Busan', 'Korea', 'TOKYO', 2025, '2025-01-12', 'Bulk', '20101', 'X', 'Invalid action code'],
            ]
        )

        response = self._post_import(upload)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        payload = response.data['data']
        self.assertEqual(payload['valid_rows'], 2)
        self.assertEqual(payload['inserted_rows'], 1)
        self.assertEqual(payload['duplicate_rows'], 1)
        self.assertEqual(payload['invalid_rows'], 1)
        self.assertEqual(len(payload['invalid_rows_sample']), 1)
        self.assertEqual(OpenSourceImportRun.objects.count(), 1)
        self.assertEqual(OpenSourceDeficiencyRecord.objects.count(), 1)

    def test_import_without_year_header_derives_year_from_inspection_date(self):
        upload = make_xlsx_file(
            [
                [1, 'Singapore', 'Singapore', 'TOKYO', '2025.01.11', 'Bulk', '1101', '30', 'First row'],
                [2, 'Busan', 'Korea', 'TOKYO', '2025.01.12', 'Bulk', '20101', 17, 'Second row'],
            ],
            headers=[
                'No.',
                'Inspection Port',
                'Inspection Country',
                'MOU',
                'InspectionDate',
                'ship_type',
                'Deficiency Code',
                'ActionCode',
                'Description',
            ],
            name='missing-year.xlsx',
        )

        response = self._post_import(upload)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        payload = response.data['data']
        self.assertEqual(payload['valid_rows'], 2)
        self.assertEqual(payload['inserted_rows'], 2)
        self.assertEqual(payload['invalid_rows'], 0)
        self.assertEqual(OpenSourceDeficiencyRecord.objects.count(), 2)
        self.assertEqual(OpenSourceDeficiencyRecord.objects.order_by('id').first().year, 2025)

    def test_import_without_year_invalid_inspection_date_skips_invalid_with_201(self):
        upload = make_xlsx_file(
            [
                [1, 'Singapore', 'Singapore', 'TOKYO', 'bad-date', 'Bulk', '1101', '30', 'Invalid date'],
                [2, 'Busan', 'Korea', 'TOKYO', '2025.01.12', 'Bulk', '20101', 17, 'Valid row'],
            ],
            headers=[
                'No.',
                'Inspection Port',
                'Inspection Country',
                'MOU',
                'InspectionDate',
                'ship_type',
                'Deficiency Code',
                'ActionCode',
                'Description',
            ],
            name='missing-year-partial-invalid.xlsx',
        )

        response = self._post_import(upload)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        payload = response.data['data']
        self.assertEqual(payload['valid_rows'], 1)
        self.assertEqual(payload['inserted_rows'], 1)
        self.assertEqual(payload['invalid_rows'], 1)
        self.assertEqual(len(payload['invalid_rows_sample']), 1)
        self.assertEqual(OpenSourceDeficiencyRecord.objects.count(), 1)

    def test_import_without_year_invalid_inspection_date_all_invalid_returns_400(self):
        upload = make_xlsx_file(
            [
                [1, 'Singapore', 'Singapore', 'TOKYO', 'bad-date', 'Bulk', '1101', '30', 'Invalid date'],
            ],
            headers=[
                'No.',
                'Inspection Port',
                'Inspection Country',
                'MOU',
                'InspectionDate',
                'ship_type',
                'Deficiency Code',
                'ActionCode',
                'Description',
            ],
            name='missing-year-all-invalid.xlsx',
        )

        response = self._post_import(upload)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data['error'], 'VALIDATION_ERROR')
        self.assertEqual(response.data['invalid_rows'], 1)
        self.assertEqual(OpenSourceImportRun.objects.count(), 0)
        self.assertEqual(OpenSourceDeficiencyRecord.objects.count(), 0)

    def test_import_rejects_missing_required_headers(self):
        upload = make_xlsx_file(
            [[1, 'Singapore', 'Singapore', 2025, '2025-01-11', 'Bulk', '1101', '30', 'First row']],
            headers=[
                'No.',
                'Inspection Port',
                'Inspection Country',
                'year',
                'InspectionDate',
                'ship_type',
                'Deficiency Code',
                'ActionCode',
                'Description',
            ],
            name='missing-mou.xlsx',
        )

        response = self._post_import(upload)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data['error'], 'VALIDATION_ERROR')
        self.assertIn('MOU', response.data['missing_headers'])

    def test_vessel_master_cannot_import_opensource(self):
        self.user = make_vessel_master_user()
        upload = make_xlsx_file(
            [
                [1, 'Singapore', 'Singapore', 'TOKYO', 2025, '2025-01-11', 'Bulk', '1101', '30', 'First row'],
            ]
        )

        response = self._post_import(upload)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(response.data['error'], 'FORBIDDEN')
        self.assertIn('Office only', response.data['message'])

    def test_import_handles_missing_storage_table_with_structured_500(self):
        upload = make_xlsx_file(
            [
                [1, 'Singapore', 'Singapore', 'TOKYO', 2025, '2025-01-11', 'Bulk', '1101', '30', 'First row'],
            ]
        )

        with patch(
            'apps.inspection.defintel_views.OpenSourceDeficiencyRecord.objects.filter',
            side_effect=ProgrammingError("Invalid object name 'psc_opensource_deficiency_record'."),
        ):
            response = self._post_import(upload)

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertEqual(response.data['error'], 'SERVER_ERROR')
        self.assertIn('not initialized', response.data['message'])


class TestVesselPrepChecklistPhase3(TestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.preview_view = VesselPrepPreviewView.as_view()
        self.export_view = VesselPrepExportView.as_view()

        self.vessel_id = uuid.uuid4()
        self.other_vessel_id = uuid.uuid4()
        self.user = SimpleNamespace(
            id=str(uuid.uuid4()),
            role=RoleCodes.VESSEL_MASTER,
            user_type='VESSEL',
            vessel_id=str(self.vessel_id),
            display_name='Vessel Master',
            username='vessel_master',
            rank='Master',
            is_authenticated=True,
        )

        self.inspection = Inspection.objects.create(
            vessel_id=self.vessel_id,
            inspection_type='PSC',
            psc_subtype='INITIAL',
            inspection_date=date(2025, 1, 10),
            port_place='Singapore',
            country='Singapore',
            mou_id='Tokyo',
            inspector_name=' John   Doe ',
            created_by=self.user.id,
        )
        Deficiency.objects.create(
            inspection=self.inspection,
            def_code_id='01101',
            def_code='01101',
            description='Life-saving appliances not ready',
            action_code_id=30,
            action_code='30',
            sequence_no=1,
        )

        same_inspector = Inspection.objects.create(
            vessel_id=self.vessel_id,
            inspection_type='PSC',
            psc_subtype='INITIAL',
            inspection_date=date(2025, 1, 12),
            port_place='Shanghai',
            country='China',
            mou_id='Tokyo',
            inspector_name='john doe',
            created_by=str(uuid.uuid4()),
        )
        Deficiency.objects.create(
            inspection=same_inspector,
            def_code_id='01234',
            def_code='01234',
            description='Fire safety deficiency',
            action_code_id=17,
            action_code='17',
            sequence_no=1,
        )

        other_inspection = Inspection.objects.create(
            vessel_id=self.other_vessel_id,
            inspection_type='PSC',
            psc_subtype='INITIAL',
            inspection_date=date(2025, 1, 11),
            port_place='Busan',
            country='Korea',
            mou_id='Tokyo',
            inspector_name='JOHN DOE',
            created_by=str(uuid.uuid4()),
        )
        Deficiency.objects.create(
            inspection=other_inspection,
            def_code_id='02020',
            def_code='02020',
            description='Other vessel deficiency',
            action_code_id=17,
            action_code='17',
            sequence_no=1,
        )

        different_inspector = Inspection.objects.create(
            vessel_id=self.vessel_id,
            inspection_type='PSC',
            psc_subtype='INITIAL',
            inspection_date=date(2025, 1, 13),
            port_place='Tokyo',
            country='Japan',
            mou_id='Tokyo',
            inspector_name='Jane Roe',
            created_by=str(uuid.uuid4()),
        )
        Deficiency.objects.create(
            inspection=different_inspector,
            def_code_id='09999',
            def_code='09999',
            description='Different inspector deficiency',
            action_code_id=30,
            action_code='30',
            sequence_no=1,
        )

    def _post_preview(self, payload):
        request = self.factory.post(
            '/api/psc/reports/vessel-prep/preview/',
            payload,
            format='json',
        )
        force_authenticate(request, user=self.user)
        return self.preview_view(request)

    def _post_export(self, payload):
        request = self.factory.post(
            '/api/psc/reports/vessel-prep/export/',
            payload,
            format='json',
        )
        force_authenticate(request, user=self.user)
        return self.export_view(request)

    @staticmethod
    def _find_export_table_header_row(sheet):
        for row_idx in range(1, sheet.max_row + 1):
            if (
                sheet.cell(row=row_idx, column=1).value == 'Sr No.'
                and sheet.cell(row=row_idx, column=2).value == 'Significance'
            ):
                return row_idx
        return None

    @staticmethod
    def _count_export_table_header_rows(sheet):
        count = 0
        for row_idx in range(1, sheet.max_row + 1):
            if (
                sheet.cell(row=row_idx, column=1).value == 'Sr No.'
                and sheet.cell(row=row_idx, column=2).value == 'Significance'
                and sheet.cell(row=row_idx, column=3).value == 'DEF Code'
            ):
                count += 1
        return count

    @staticmethod
    def _find_row_with_text(sheet, text, max_col=12):
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, max_col + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if isinstance(value, str) and text in value:
                    return row_idx
        return None

    @staticmethod
    def _cell_has_fill(cell):
        fill = getattr(cell, 'fill', None)
        return bool(fill and fill.fill_type not in (None, 'none'))

    @staticmethod
    def _cell_has_border(cell):
        border = getattr(cell, 'border', None)
        if border is None:
            return False
        sides = [border.left, border.right, border.top, border.bottom, border.diagonal]
        return any(getattr(side, 'style', None) not in (None, '') for side in sides)

    @staticmethod
    def _style_signature(cell):
        fill = getattr(cell, 'fill', None)
        font = getattr(cell, 'font', None)
        fill_color = getattr(fill, 'start_color', None)
        font_color = getattr(font, 'color', None)
        return (
            getattr(fill, 'fill_type', None),
            getattr(fill_color, 'type', None),
            getattr(fill_color, 'rgb', None),
            getattr(fill_color, 'indexed', None),
            getattr(fill_color, 'tint', None),
            getattr(font, 'bold', None),
            getattr(font_color, 'type', None),
            getattr(font_color, 'rgb', None),
            getattr(font_color, 'indexed', None),
            getattr(font_color, 'tint', None),
        )

    @staticmethod
    def _find_legend_designation_row(sheet, designation):
        legend_title_row = TestVesselPrepChecklistPhase3._find_row_with_text(
            sheet,
            'Deficiency Code Significance Designation listing',
            max_col=12,
        )
        if legend_title_row is None:
            return None

        target = str(designation).strip().upper()
        for row_idx in range(legend_title_row + 1, legend_title_row + 30):
            value = str(sheet.cell(row=row_idx, column=1).value or '').strip().upper()
            if value == target:
                return row_idx
        return None

    @staticmethod
    def _get_input_summary_value(summary_sheet, key):
        for row_idx in range(1, summary_sheet.max_row + 1):
            if summary_sheet.cell(row=row_idx, column=1).value == key:
                return summary_sheet.cell(row=row_idx, column=2).value
        return None

    @staticmethod
    def _ensure_vessel_data_table():
        table_name = VesselData._meta.db_table
        with connection.cursor() as cursor:
            try:
                cursor.execute(
                    f"""
                    CREATE TABLE IF NOT EXISTS {table_name} (
                        id TEXT PRIMARY KEY,
                        vesselName TEXT NOT NULL,
                        vesselCode TEXT NOT NULL,
                        imoNumber TEXT NULL,
                        is_active INTEGER NOT NULL DEFAULT 1,
                        is_deleted INTEGER NOT NULL DEFAULT 0
                    )
                    """
                )
            except OperationalError:
                pass

    def test_preview_vessel_scope_uses_internal_only(self):
        response = self._post_preview(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']['rows']
        self.assertEqual(len(rows), 3)
        returned_codes = {row['def_code'] for row in rows}
        self.assertEqual(returned_codes, {'01101', '01234', '09999'})
        for row in rows:
            self.assertEqual(row['occurrence_count_internal'], 1)
            self.assertEqual(row['occurrence_count_opensource'], 0)

    def test_vessel_scope_requires_vessel_id_preview(self):
        response = self._post_preview(
            {
                'scope_mode': 'VESSEL',
            }
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data['error'], 'VALIDATION_ERROR')
        self.assertEqual(
            response.data['details']['vessel_id'][0],
            'Required when scope is VESSEL',
        )

    def test_vessel_scope_requires_vessel_id_export(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
            }
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data['error'], 'VALIDATION_ERROR')
        self.assertEqual(
            response.data['details']['vessel_id'][0],
            'Required when scope is VESSEL',
        )

    def test_filter_combined_requires_import(self):
        response = self._post_preview(
            {
                'scope_mode': 'FILTER_COMBINED',
                'filters': {
                    'def_code': ['01101'],
                },
            }
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data['error'], 'IMPORT_REQUIRED')

    def test_inspector_scope_uses_inspector_name_case_insensitive_normalized(self):
        response = self._post_preview(
            {
                'scope_mode': 'INSPECTOR',
                'inspector_name': '  john    DOE  ',
            }
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']['rows']
        returned_codes = {row['def_code'] for row in rows}
        self.assertEqual(returned_codes, {'01101', '01234'})
        self.assertNotIn('09999', returned_codes)

    def test_inspector_scope_requires_inspector_name(self):
        response = self._post_preview(
            {
                'scope_mode': 'INSPECTOR',
            }
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data['error'], 'VALIDATION_ERROR')
        self.assertIn('inspector_name', response.data['details'])

    def test_filter_combined_merges_internal_and_opensource_counts(self):
        import_run = OpenSourceImportRun.objects.create(
            uploaded_by='vm-1',
            filename='monthly.xlsx',
            file_hash=hashlib.sha256(b'monthly').hexdigest(),
            total_rows=1,
            valid_rows=1,
            inserted_rows=1,
            duplicate_rows=0,
            invalid_rows=0,
        )
        dedup_key = hashlib.sha256('2025|01101|30|SINGAPORE|TOKYO'.encode('utf-8')).hexdigest()
        OpenSourceDeficiencyRecord.objects.create(
            import_run=import_run,
            year=2025,
            def_code_norm='01101',
            action_code_norm=30,
            port_norm='SINGAPORE',
            mou_norm='TOKYO',
            country_norm='SINGAPORE',
            description_raw='OpenSource sample row',
            dedup_key_hash=dedup_key,
        )

        response = self._post_preview(
            {
                'scope_mode': 'FILTER_COMBINED',
                'filters': {
                    'def_code': ['01101'],
                    'action_code': ['30'],
                    'mou': ['TOKYO'],
                    'port': ['SINGAPORE'],
                },
                'dedup': False,
            }
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']['rows']
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['occurrence_count_internal'], 1)
        self.assertEqual(rows[0]['occurrence_count_opensource'], 1)
        self.assertEqual(rows[0]['occurrence_count_total'], 2)

    def test_filter_combined_export_with_410_opensource_rows_returns_xlsx(self):
        import_run = OpenSourceImportRun.objects.create(
            uploaded_by='vm-1',
            filename='monthly-large.xlsx',
            file_hash=hashlib.sha256(b'monthly-large').hexdigest(),
            total_rows=410,
            valid_rows=410,
            inserted_rows=410,
            duplicate_rows=0,
            invalid_rows=0,
        )
        records = []
        for idx in range(410):
            def_code = str(10000 + idx)
            port = f'PORT-{idx}'
            dedup_key = hashlib.sha256(f'2025|{def_code}|17|{port}|TOKYO'.encode('utf-8')).hexdigest()
            records.append(
                OpenSourceDeficiencyRecord(
                    import_run=import_run,
                    year=2025,
                    def_code_norm=def_code,
                    action_code_norm=17,
                    port_norm=port,
                    mou_norm='TOKYO',
                    country_norm='COUNTRY',
                    description_raw=f'OpenSource row {idx}',
                    dedup_key_hash=dedup_key,
                )
            )
        OpenSourceDeficiencyRecord.objects.bulk_create(records)

        response = self._post_export(
            {
                'scope_mode': 'FILTER_COMBINED',
                'date_from': '2025-11-01',
                'date_to': '2025-12-31',
                'dedup': True,
                'filters': {
                    'action_code': ['17'],
                },
            }
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment; filename="Vessel_Preparation_Checklist_', response['Content-Disposition'])
        self.assertGreater(len(response.content), 0)

        loaded = load_workbook(filename=BytesIO(response.content))
        checklist_sheet = loaded['Preparation Checklist']
        header_row = self._find_export_table_header_row(checklist_sheet)
        self.assertIsNotNone(header_row)
        self.assertEqual(checklist_sheet.cell(row=header_row + 410, column=3).value, '10409')

        summary_sheet = loaded['Input Summary']
        row_count_value = self._get_input_summary_value(summary_sheet, 'row_count')
        self.assertEqual(row_count_value, '410')

    def test_export_returns_required_sheet_names(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        loaded = load_workbook(filename=BytesIO(response.content))
        self.assertIn('Preparation Checklist', loaded.sheetnames)
        self.assertIn('Input Summary', loaded.sheetnames)
        self.assertNotIn('System Summary', loaded.sheetnames)
        self.assertNotIn('Legend', loaded.sheetnames)
        sheet = loaded['Preparation Checklist']
        header_row_index = self._find_export_table_header_row(sheet)
        self.assertIsNotNone(header_row_index)
        self.assertEqual(header_row_index, 8)
        header_row = [sheet.cell(row=header_row_index, column=index).value for index in range(1, 13)]
        self.assertEqual(
            header_row,
            [
                'Sr No.',
                'Significance',
                'DEF Code',
                'Deficiency / Description',
                'Action Code',
                'Action Label',
                'MoU',
                'Port',
                'Country',
                'Ship Status (OK / Not OK / N.A.)',
                'Ship Remarks / Evidence',
                'Office Comments',
            ],
        )
        self.assertEqual(sheet.cell(row=1, column=1).value, 'PSC Preparation Checklist')
        generated_for_value = str(sheet.cell(row=3, column=1).value or '')
        self.assertIn('Generated for:', generated_for_value)

    def test_table_has_sr_no_and_counts_match_row_count(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        expected_headers = [
            'Sr No.',
            'Significance',
            'DEF Code',
            'Deficiency / Description',
            'Action Code',
            'Action Label',
            'MoU',
            'Port',
            'Country',
            'Ship Status (OK / Not OK / N.A.)',
            'Ship Remarks / Evidence',
            'Office Comments',
        ]
        actual_headers = [sheet.cell(row=8, column=index).value for index in range(1, 13)]
        self.assertEqual(actual_headers, expected_headers)

        row_count = int(self._get_input_summary_value(loaded['Input Summary'], 'row_count'))
        last_data_row = 8 + row_count
        self.assertGreater(row_count, 0)
        sr_no_values = [sheet.cell(row=row_idx, column=1).value for row_idx in range(9, last_data_row + 1)]
        self.assertEqual(sr_no_values, list(range(1, row_count + 1)))

    def test_last_seen_column_removed(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        headers = [sheet.cell(row=8, column=index).value for index in range(1, 13)]
        self.assertEqual(
            headers,
            [
                'Sr No.',
                'Significance',
                'DEF Code',
                'Deficiency / Description',
                'Action Code',
                'Action Label',
                'MoU',
                'Port',
                'Country',
                'Ship Status (OK / Not OK / N.A.)',
                'Ship Remarks / Evidence',
                'Office Comments',
            ],
        )
        self.assertNotIn('Last Seen', headers)
        self.assertEqual(len(headers), 12)

    def test_input_summary_sheet_is_hidden(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        self.assertEqual(loaded['Input Summary'].sheet_state, 'hidden')
        self.assertNotIn('System Summary', loaded.sheetnames)
        self.assertNotIn('Legend', loaded.sheetnames)

    def test_main_sheet_has_no_occurrence_summary(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=12):
            for cell in row:
                if isinstance(cell.value, str):
                    self.assertNotIn('Occurrence Summary', cell.value)

    def test_main_sheet_has_no_occurrence_columns(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        headers = [sheet.cell(row=8, column=index).value for index in range(1, 13)]
        header_text = ' | '.join(str(value or '') for value in headers)
        self.assertNotIn('Occurrences', header_text)
        self.assertNotIn('Internal', header_text)
        self.assertNotIn('OpenSource', header_text)

    def test_preparation_checklist_has_no_stray_template_text(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        row_count = int(self._get_input_summary_value(loaded['Input Summary'], 'row_count'))
        last_data_row = 8 + row_count
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, 13):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if not isinstance(value, str):
                    continue
                if 'Excessive corrosion' not in value:
                    continue
                is_table_description_cell = col_idx == 4 and 9 <= row_idx <= last_data_row
                self.assertTrue(is_table_description_cell)

    def test_footer_starts_5_rows_after_last_data_row(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        row_count = int(self._get_input_summary_value(loaded['Input Summary'], 'row_count'))
        last_data_row = 8 + row_count
        legend_row = self._find_row_with_text(sheet, 'Deficiency Code Significance Designation listing')
        self.assertIsNotNone(legend_row)
        self.assertIn(legend_row - last_data_row, (5, 6))

    def test_no_stray_fills_in_cleared_region(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        row_count = int(self._get_input_summary_value(loaded['Input Summary'], 'row_count'))
        last_data_row = 8 + row_count
        legend_row = self._find_row_with_text(sheet, 'Deficiency Code Significance Designation listing')
        self.assertIsNotNone(legend_row)

        sample_start_row = last_data_row + 1
        sample_end_row = max(sample_start_row, legend_row - 1)
        for row_idx in range(sample_start_row, sample_end_row + 1):
            for col_idx in range(1, 13):
                cell = sheet.cell(row=row_idx, column=col_idx)
                self.assertIsNone(cell.value)
                self.assertFalse(self._cell_has_fill(cell))
                self.assertFalse(self._cell_has_border(cell))

    def test_footer_legend_moves_below_table_when_table_is_long(self):
        import_run = OpenSourceImportRun.objects.create(
            uploaded_by='vm-1',
            filename='monthly-large-footer.xlsx',
            file_hash=hashlib.sha256(b'monthly-large-footer').hexdigest(),
            total_rows=410,
            valid_rows=410,
            inserted_rows=410,
            duplicate_rows=0,
            invalid_rows=0,
        )
        records = []
        for idx in range(410):
            def_code = str(20000 + idx)
            port = f'LONG-PORT-{idx}'
            dedup_key = hashlib.sha256(f'2025|{def_code}|17|{port}|TOKYO'.encode('utf-8')).hexdigest()
            records.append(
                OpenSourceDeficiencyRecord(
                    import_run=import_run,
                    year=2025,
                    def_code_norm=def_code,
                    action_code_norm=17,
                    port_norm=port,
                    mou_norm='TOKYO',
                    country_norm='COUNTRY',
                    description_raw=f'Footer row {idx}',
                    dedup_key_hash=dedup_key,
                )
            )
        OpenSourceDeficiencyRecord.objects.bulk_create(records)

        response = self._post_export(
            {
                'scope_mode': 'FILTER_COMBINED',
                'date_from': '2025-11-01',
                'date_to': '2025-12-31',
                'dedup': True,
                'filters': {
                    'action_code': ['17'],
                },
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        row_count = int(self._get_input_summary_value(loaded['Input Summary'], 'row_count'))
        last_data_row = 8 + row_count
        legend_row = self._find_row_with_text(sheet, 'Deficiency Code Significance Designation listing')
        self.assertIsNotNone(legend_row)
        self.assertIn(legend_row - last_data_row, (5, 6))

    def test_action_30_labelled_detention(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        row_count = int(self._get_input_summary_value(loaded['Input Summary'], 'row_count'))
        last_data_row = 8 + row_count

        detention_rows = []
        for row_idx in range(9, last_data_row + 1):
            action_code = sheet.cell(row=row_idx, column=5).value
            if str(action_code).strip() == '30':
                detention_rows.append(row_idx)
                self.assertEqual(sheet.cell(row=row_idx, column=6).value, 'DETENTION')

        self.assertGreater(len(detention_rows), 0)

    def test_action_code_30_forces_significance_det1_value(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        row_count = int(self._get_input_summary_value(loaded['Input Summary'], 'row_count'))
        last_data_row = 8 + row_count

        detention_rows = []
        for row_idx in range(9, last_data_row + 1):
            action_code = sheet.cell(row=row_idx, column=5).value
            if str(action_code).strip() == '30':
                detention_rows.append(row_idx)
                self.assertEqual(sheet.cell(row=row_idx, column=2).value, 'DET1')

        self.assertGreater(len(detention_rows), 0)

    def test_action_code_30_uses_det1_style(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        row_count = int(self._get_input_summary_value(loaded['Input Summary'], 'row_count'))
        last_data_row = 8 + row_count

        legend_style_row = self._find_legend_designation_row(sheet, 'DET1')
        if legend_style_row is None:
            legend_style_row = self._find_legend_designation_row(sheet, 'DETXX')
        self.assertIsNotNone(legend_style_row)

        expected_style = self._style_signature(sheet.cell(row=legend_style_row, column=1))
        detention_rows = []
        for row_idx in range(9, last_data_row + 1):
            action_code = sheet.cell(row=row_idx, column=5).value
            if str(action_code).strip() == '30':
                detention_rows.append(row_idx)
                self.assertEqual(
                    self._style_signature(sheet.cell(row=row_idx, column=2)),
                    expected_style,
                )

        self.assertGreater(len(detention_rows), 0)

    def test_office_comments_column_present_and_visible(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        headers = [sheet.cell(row=8, column=index).value for index in range(1, 13)]
        self.assertEqual(len(headers), 12)
        self.assertIn('Office Comments', headers)
        self.assertEqual(sheet.cell(row=8, column=12).value, 'Office Comments')

    def test_export_has_single_table_header_once(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']

        self.assertEqual(self._count_export_table_header_rows(sheet), 1)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=12):
            for cell in row:
                self.assertNotEqual(cell.value, 'Type')

    def test_export_vessel_header_not_uuid_or_username(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        vessel_line = str(sheet.cell(row=2, column=1).value or '')
        self.assertTrue(vessel_line.startswith('Vessel:'))

        vessel_value = vessel_line.split(':', 1)[1].strip()
        self.assertNotRegex(
            vessel_value,
            r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$',
        )
        self.assertNotEqual(vessel_value.lower(), str(self.user.display_name).lower())
        self.assertNotEqual(vessel_value.lower(), str(self.user.username).lower())

    def test_export_has_no_AUTO_visible(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=12):
            for cell in row:
                if isinstance(cell.value, str):
                    self.assertNotEqual(cell.value.strip().upper(), 'AUTO')

    def test_export_header_uses_request_port_filter(self):
        OpenSourceImportRun.objects.create(
            uploaded_by='vm-1',
            filename='monthly-empty.xlsx',
            file_hash=hashlib.sha256(b'monthly-empty').hexdigest(),
            total_rows=0,
            valid_rows=0,
            inserted_rows=0,
            duplicate_rows=0,
            invalid_rows=0,
        )
        response = self._post_export(
            {
                'scope_mode': 'FILTER_COMBINED',
                'filters': {
                    'port': ['SINGAPORE'],
                },
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        generated_for_value = str(sheet.cell(row=3, column=1).value or '')
        self.assertIn('Generated for:', generated_for_value)
        self.assertIn('SINGAPORE', generated_for_value)

    def test_export_colors_only_on_significance_when_available(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        header_row_index = self._find_export_table_header_row(sheet)
        self.assertIsNotNone(header_row_index)

        for row_idx in range(header_row_index + 1, header_row_index + 4):
            significance = sheet.cell(row=row_idx, column=2).value
            if significance not in (None, ''):
                self.assertTrue(self._cell_has_fill(sheet.cell(row=row_idx, column=2)))
            else:
                self.assertFalse(self._cell_has_fill(sheet.cell(row=row_idx, column=2)))
            for col_idx in range(1, 13):
                if col_idx == 2:
                    continue
                self.assertFalse(self._cell_has_fill(sheet.cell(row=row_idx, column=col_idx)))

    def test_unknown_significance_has_no_color_and_blank_cell(self):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        header_row_index = self._find_export_table_header_row(sheet)
        self.assertIsNotNone(header_row_index)

        row_count = int(self._get_input_summary_value(loaded['Input Summary'], 'row_count'))
        last_data_row = 8 + row_count
        target_row = None
        for row_idx in range(9, last_data_row + 1):
            if str(sheet.cell(row=row_idx, column=5).value or '').strip() == '17':
                target_row = row_idx
                break

        self.assertIsNotNone(target_row)
        significance = sheet.cell(row=target_row, column=2).value

        self.assertIn(significance, (None, ''))
        self.assertFalse(self._cell_has_fill(sheet.cell(row=target_row, column=2)))

    def test_preview_export_rowcount_parity_same_payload(self):
        payload = {
            'scope_mode': 'VESSEL',
            'vessel_id': str(self.vessel_id),
        }
        preview_response = self._post_preview(payload)
        self.assertEqual(preview_response.status_code, status.HTTP_200_OK)
        preview_row_count = preview_response.data['data']['summary']['row_count']

        export_response = self._post_export(payload)
        self.assertEqual(export_response.status_code, status.HTTP_200_OK)
        loaded = load_workbook(filename=BytesIO(export_response.content))
        summary_sheet = loaded['Input Summary']
        export_row_count = self._get_input_summary_value(summary_sheet, 'row_count')

        self.assertEqual(str(preview_row_count), str(export_row_count))

    def test_preview_export_metadata_parity_same_payload(self):
        payload = {
            'scope_mode': 'VESSEL',
            'vessel_id': str(self.vessel_id),
            'date_from': '2025-01-01',
            'date_to': '2025-01-31',
        }
        preview_response = self._post_preview(payload)
        self.assertEqual(preview_response.status_code, status.HTTP_200_OK)

        export_response = self._post_export(payload)
        self.assertEqual(export_response.status_code, status.HTTP_200_OK)
        loaded = load_workbook(filename=BytesIO(export_response.content))
        sheet = loaded['Preparation Checklist']
        date_range_line = str(sheet.cell(row=4, column=1).value or '')
        self.assertIn('2025-01-01', date_range_line)
        self.assertIn('2025-01-31', date_range_line)

    def test_export_vessel_header_resolves_name_when_vessel_id_present(self):
        self._ensure_vessel_data_table()
        VesselData.objects.filter(id=self.vessel_id).delete()
        VesselData.objects.create(
            id=self.vessel_id,
            vesselName='MV PARITY',
            vesselCode='MVP',
            imoNumber='1234567',
            is_active=True,
            is_deleted=False,
        )

        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        vessel_line = str(sheet.cell(row=2, column=1).value or '')
        self.assertIn('MV PARITY', vessel_line)

    @patch('apps.inspection.defintel_checklist._lookup_vessel_name', return_value='')
    def test_export_vessel_header_uses_request_vessel_name_when_lookup_unavailable(self, _mock_lookup):
        response = self._post_export(
            {
                'scope_mode': 'VESSEL',
                'vessel_id': str(self.vessel_id),
                'vessel_name': 'MV SELECTED',
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        loaded = load_workbook(filename=BytesIO(response.content))
        sheet = loaded['Preparation Checklist']
        vessel_line = str(sheet.cell(row=2, column=1).value or '')
        self.assertIn('MV SELECTED', vessel_line)

    def test_office_user_can_preview_and_export(self):
        self.user = make_office_user()

        preview_response = self._post_preview(
            {
                'scope_mode': 'FLEET',
            }
        )
        self.assertEqual(preview_response.status_code, status.HTTP_200_OK)

        export_response = self._post_export(
            {
                'scope_mode': 'FLEET',
            }
        )
        self.assertEqual(export_response.status_code, status.HTTP_200_OK)
