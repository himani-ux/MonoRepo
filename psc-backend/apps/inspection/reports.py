"""
Deficiency Excel export using openpyxl.

Source: DESIGN_SYSTEM.md Section 12 (PDF/Report Styling)
Implements: PRD.md FEAT-RPT-002

Sheets:
  1. Deficiency Summary — all deficiencies with inspection context
  2. CAR Status — CAR status overview per deficiency
  3. Applied Filters — filter criteria used for this export
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Design tokens from DESIGN_SYSTEM.md Section 12
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
DETENTION_FILL = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')

HEADER_FONT = Font(name='Arial', size=10, bold=True, color='1F2937')
BODY_FONT = Font(name='Arial', size=10, color='1F2937')
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='1F2937')

THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='top', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)


# ---------------------------------------------------------------------------
# Excel Builder
# ---------------------------------------------------------------------------
def generate_deficiency_excel(deficiencies: list, filters: dict) -> bytes:
    """
    Generate a multi-sheet Excel workbook for deficiency export.

    Args:
        deficiencies: List of deficiency dicts with nested inspection/CAR data.
        filters: Dict of filter criteria applied to this export.

    Returns:
        Excel file content as bytes.

    Source: DESIGN_SYSTEM.md Section 12
    Implements: PRD.md FEAT-RPT-002
    """
    wb = Workbook()

    # Sheet 1: Deficiency Summary
    _build_deficiency_summary_sheet(wb, deficiencies)

    # Sheet 2: CAR Status
    _build_car_status_sheet(wb, deficiencies)

    # Sheet 3: Applied Filters
    _build_filters_sheet(wb, filters)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ---------------------------------------------------------------------------
# Sheet 1: Deficiency Summary
# ---------------------------------------------------------------------------
def _build_deficiency_summary_sheet(wb: Workbook, deficiencies: list):
    """Build the main deficiency summary sheet."""
    ws = wb.active
    assert ws is not None
    ws.title = 'Deficiency Summary'

    headers = [
        'No.', 'Vessel', 'Inspection Type', 'Inspection Date',
        'Port', 'Def Code', 'Description', 'Action Code',
        'Target Date', 'Cleared', 'Cleared Date', 'Detention',
    ]

    col_widths = [6, 20, 15, 14, 20, 12, 40, 12, 14, 10, 14, 10]

    # Write headers
    _write_header_row(ws, headers, col_widths)

    # Write data rows
    for idx, deficiency in enumerate(deficiencies, 1):
        inspection = deficiency.get('inspection') or {}
        is_detention = inspection.get('is_detention', False)

        row = idx + 1
        values = [
            idx,
            inspection.get('vessel_name', ''),
            inspection.get('inspection_type', ''),
            _fmt_date(inspection.get('inspection_date')),
            inspection.get('port_place', ''),
            deficiency.get('def_code', ''),
            deficiency.get('description', ''),
            _safe(deficiency.get('action_code')),
            _fmt_date(deficiency.get('target_date')),
            'Yes' if deficiency.get('is_cleared') else 'No',
            _fmt_date(deficiency.get('cleared_date')),
            'Yes' if is_detention else 'No',
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN

        # Detention row highlighting per DESIGN_SYSTEM.md
        if is_detention:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = DETENTION_FILL
        elif idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = ALT_ROW_FILL

    # Auto-filter per FEAT-RPT-002
    if deficiencies:
        last_row = len(deficiencies) + 1
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A1:{last_col}{last_row}'


# ---------------------------------------------------------------------------
# Sheet 2: CAR Status
# ---------------------------------------------------------------------------
def _build_car_status_sheet(wb: Workbook, deficiencies: list):
    """Build the CAR status overview sheet."""
    ws = wb.create_sheet('CAR Status')

    headers = [
        'No.', 'CAR Number', 'Def Code', 'Description',
        'CAR Status', 'Root Cause Summary', 'Target Date',
        'Evidence Count', 'Actions Count',
    ]

    col_widths = [6, 18, 12, 35, 16, 40, 14, 14, 14]

    # Write headers
    _write_header_row(ws, headers, col_widths)

    # Write data rows
    for idx, deficiency in enumerate(deficiencies, 1):
        car = deficiency.get('car') or {}

        row = idx + 1
        values = [
            idx,
            car.get('car_number', '—'),
            deficiency.get('def_code', ''),
            deficiency.get('description', ''),
            car.get('status_display', car.get('status', '—')),
            _truncate(car.get('root_cause_summary', ''), 100),
            _fmt_date(car.get('target_date')),
            car.get('evidence_count', 0),
            car.get('actions_count', 0),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN

        # Alternating rows
        if idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = ALT_ROW_FILL

    # Auto-filter
    if deficiencies:
        last_row = len(deficiencies) + 1
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A1:{last_col}{last_row}'


# ---------------------------------------------------------------------------
# Sheet 3: Applied Filters
# ---------------------------------------------------------------------------
def _build_filters_sheet(wb: Workbook, filters: dict):
    """Build the filter criteria sheet."""
    ws = wb.create_sheet('Applied Filters')

    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws.cell(row=1, column=1, value='Export Filter Criteria')
    title_cell.font = TITLE_FONT

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40

    # Headers
    ws.cell(row=3, column=1, value='Filter').font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=1).border = THIN_BORDER
    ws.cell(row=3, column=2, value='Value').font = HEADER_FONT
    ws.cell(row=3, column=2).fill = HEADER_FILL
    ws.cell(row=3, column=2).border = THIN_BORDER

    filter_labels = {
        'vessel_id': 'Vessel ID',
        'vessel_name': 'Vessel Name',
        'inspection_type': 'Inspection Type',
        'status': 'Inspection Status',
        'date_from': 'Date From',
        'date_to': 'Date To',
    }

    row = 4
    for key, label in filter_labels.items():
        value = filters.get(key)
        if value:
            ws.cell(row=row, column=1, value=label).font = BODY_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=str(value)).font = BODY_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1

    # Generated timestamp
    row += 1
    ws.cell(row=row, column=1, value='Generated At').font = HEADER_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=datetime.now().strftime('%d %b %Y %H:%M')).font = BODY_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER

    ws.cell(row=row + 1, column=1, value='Total Records').font = HEADER_FONT
    ws.cell(row=row + 1, column=1).border = THIN_BORDER
    ws.cell(row=row + 1, column=2, value=filters.get('total_count', 0)).font = BODY_FONT
    ws.cell(row=row + 1, column=2).border = THIN_BORDER


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _write_header_row(ws, headers: list, col_widths: list):
    """Write styled header row with column widths."""
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _fmt_date(value) -> str:
    """Format a date string for display."""
    if not value:
        return ''
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d', '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%dT%H:%M:%SZ'):
            try:
                return datetime.strptime(value, fmt).strftime('%d %b %Y')
            except ValueError:
                continue
    return str(value) if value else ''


def _safe(value, default: str = '') -> str:
    """Return value or default if None/empty."""
    if value is None or value == '':
        return default
    return str(value)


def _truncate(text: str, max_len: int) -> str:
    """Truncate text with ellipsis if too long."""
    if not text:
        return ''
    if len(text) <= max_len:
        return text
    return text[:max_len - 3] + '...'
