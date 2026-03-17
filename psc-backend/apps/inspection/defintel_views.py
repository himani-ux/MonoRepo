"""
DefIntel/OpenSource import views (Phase 2 only).
"""

import hashlib
import io
import logging
from datetime import date, datetime
from zipfile import BadZipFile

from openpyxl import load_workbook
from django.db import DataError, DatabaseError, IntegrityError, ProgrammingError, transaction
from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .defintel_access import can_import_opensource
from .defintel_models import OpenSourceDeficiencyRecord, OpenSourceImportRun


MAX_TROUBLESHOOT_ROWS = 20
logger = logging.getLogger(__name__)

HEADER_ALIASES = {
    'no': 'no',
    'no.': 'no',
    'inspection port': 'inspection_port',
    'inspection country': 'inspection_country',
    'mou': 'mou',
    'year': 'year',
    'inspectiondate': 'inspection_date',
    'inspection date': 'inspection_date',
    'ship_type': 'ship_type',
    'ship type': 'ship_type',
    'deficiency code': 'deficiency_code',
    'actioncode': 'action_code',
    'action code': 'action_code',
    'description': 'description',
}

REQUIRED_HEADERS = {
    'no': 'No.',
    'inspection_port': 'Inspection Port',
    'inspection_country': 'Inspection Country',
    'mou': 'MOU',
    'inspection_date': 'InspectionDate',
    'ship_type': 'ship_type',
    'deficiency_code': 'Deficiency Code',
    'action_code': 'ActionCode',
    'description': 'Description',
}


def _normalize_header(value):
    if value is None:
        return ''
    return ' '.join(str(value).strip().lower().split())


def _sha256_hex(content):
    return hashlib.sha256(content).hexdigest()


def _normalize_def_code(value):
    if value is None:
        raise ValueError('def_code is required')

    text = str(value).strip()
    if not text:
        raise ValueError('def_code is required')

    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]

    if not text.isdigit():
        raise ValueError('def_code must be numeric')

    if len(text) > 5:
        raise ValueError('def_code must be at most 5 digits')

    return text.zfill(5)


def _normalize_action_code(value):
    if value is None:
        raise ValueError('action_code is required')

    text = str(value).strip()
    if not text:
        raise ValueError('action_code is required')

    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]

    if not text.isdigit():
        raise ValueError('action_code must be numeric')

    return int(text)


def _normalize_year(value):
    if value is None:
        raise ValueError('year is required')

    text = str(value).strip()
    if not text:
        raise ValueError('year is required')

    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]

    if not text.isdigit():
        raise ValueError('year must be numeric')

    year = int(text)
    if year <= 0:
        raise ValueError('year must be positive')

    return year


def _parse_inspection_date(value):
    if value is None:
        raise ValueError('inspection_date is required')

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        raise ValueError('inspection_date is required')

    for fmt in ('%Y.%m.%d', '%Y-%m-%d', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ValueError('inspection_date must be in YYYY.MM.DD format')


def _normalize_text(value, *, field_name, optional=False):
    if value is None:
        if optional:
            return None
        raise ValueError(f'{field_name} is required')

    text = ' '.join(str(value).strip().split())
    if not text:
        if optional:
            return None
        raise ValueError(f'{field_name} is required')

    return text.upper()


def _dedup_source_string(year, def_code_norm, action_code_norm, port_norm, mou_norm):
    return f'{year}|{def_code_norm}|{action_code_norm}|{port_norm}|{mou_norm}'


def _row_is_blank(row):
    return all(value is None or str(value).strip() == '' for value in row)


class OpenSourceImportView(APIView):
    """
    POST /api/psc/reports/opensource/import/

    Upload a monthly OpenSource xlsx file, normalize rows, and persist
    deduplicated records in DefIntel-only tables.
    """

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def _missing_storage_response(self):
        return Response(
            {
                'error': 'SERVER_ERROR',
                'message': 'OpenSource import storage is not initialized. Run backend migrations and retry.',
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    def _server_error_response(self):
        return Response(
            {
                'error': 'SERVER_ERROR',
                'message': 'OpenSource import failed due to an internal server error.',
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    def post(self, request, *args, **kwargs):
        if not can_import_opensource(request.user):
            return Response(
                {
                    'error': 'FORBIDDEN',
                    'message': 'Office only: OpenSource import is restricted to office users.',
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Missing required multipart file field: file',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        filename = uploaded_file.name or ''
        if not filename.lower().endswith('.xlsx'):
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Only .xlsx files are supported.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        file_bytes = uploaded_file.read()
        if not file_bytes:
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Uploaded file is empty.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        file_hash = _sha256_hex(file_bytes)
        logger.info(
            'OpenSource import started filename=%s bytes=%s user_id=%s',
            filename,
            len(file_bytes),
            getattr(request.user, 'id', None),
        )

        try:
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            worksheet = workbook.active
            logger.info('OpenSource import workbook loaded sheet=%s', getattr(worksheet, 'title', ''))
        except BadZipFile:
            logger.warning('OpenSource import rejected invalid zip/xlsx filename=%s', filename)
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Invalid Excel file. Could not read workbook.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception:
            logger.exception('OpenSource import workbook parse failed filename=%s', filename)
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Invalid Excel file. Could not read workbook.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Worksheet is empty.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        header_index = {}
        for idx, header_value in enumerate(header_row):
            canonical = HEADER_ALIASES.get(_normalize_header(header_value))
            if canonical and canonical not in header_index:
                header_index[canonical] = idx

        missing_headers = [
            display_name
            for key, display_name in REQUIRED_HEADERS.items()
            if key not in header_index
        ]
        if missing_headers:
            logger.warning(
                'OpenSource import missing headers filename=%s missing=%s',
                filename,
                missing_headers,
            )
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Missing required headers.',
                    'missing_headers': missing_headers,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        parsed_valid_rows = []
        invalid_rows_sample = []
        total_rows = 0
        invalid_rows = 0

        def cell(row_values, key):
            idx = header_index[key]
            return row_values[idx] if idx < len(row_values) else None

        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if _row_is_blank(row_values):
                continue

            total_rows += 1

            raw_snapshot = {
                'year': cell(row_values, 'year') if 'year' in header_index else None,
                'inspection_date': cell(row_values, 'inspection_date'),
                'deficiency_code': cell(row_values, 'deficiency_code'),
                'action_code': cell(row_values, 'action_code'),
                'inspection_port': cell(row_values, 'inspection_port'),
                'mou': cell(row_values, 'mou'),
                'inspection_country': cell(row_values, 'inspection_country')
                if 'inspection_country' in header_index
                else None,
                'description': cell(row_values, 'description')
                if 'description' in header_index
                else None,
            }

            try:
                inspection_date = _parse_inspection_date(raw_snapshot['inspection_date'])

                year = None
                raw_year = raw_snapshot.get('year')
                if raw_year is not None and str(raw_year).strip() != '':
                    try:
                        year = _normalize_year(raw_year)
                    except ValueError:
                        year = inspection_date.year
                if year is None:
                    year = inspection_date.year

                def_code_norm = _normalize_def_code(raw_snapshot['deficiency_code'])
                action_code_norm = _normalize_action_code(raw_snapshot['action_code'])
                port_norm = _normalize_text(
                    raw_snapshot['inspection_port'],
                    field_name='inspection_port',
                )
                mou_norm = _normalize_text(
                    raw_snapshot['mou'],
                    field_name='mou',
                )
                country_norm = _normalize_text(
                    raw_snapshot.get('inspection_country'),
                    field_name='inspection_country',
                    optional=True,
                )
                description_raw = raw_snapshot.get('description')
                description_raw = (
                    ' '.join(str(description_raw).strip().split())
                    if description_raw is not None and str(description_raw).strip() != ''
                    else None
                )

                dedup_source = _dedup_source_string(
                    year,
                    def_code_norm,
                    action_code_norm,
                    port_norm,
                    mou_norm,
                )
                dedup_key_hash = _sha256_hex(dedup_source.encode('utf-8'))

                parsed_valid_rows.append(
                    {
                        'row_number': row_number,
                        'year': year,
                        'def_code_norm': def_code_norm,
                        'action_code_norm': action_code_norm,
                        'port_norm': port_norm,
                        'mou_norm': mou_norm,
                        'country_norm': country_norm,
                        'description_raw': description_raw,
                        'dedup_key_hash': dedup_key_hash,
                    }
                )
            except ValueError as exc:
                invalid_rows += 1
                if len(invalid_rows_sample) < MAX_TROUBLESHOOT_ROWS:
                    invalid_rows_sample.append(
                        {
                            'row_number': row_number,
                            'reason': str(exc),
                            'raw': raw_snapshot,
                        }
                    )

        valid_rows = len(parsed_valid_rows)

        if invalid_rows:
            logger.warning(
                'OpenSource import invalid rows filename=%s total_rows=%s invalid_rows=%s',
                filename,
                total_rows,
                invalid_rows,
            )
        if valid_rows == 0:
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'No valid rows found in uploaded file.',
                    'invalid_rows': invalid_rows,
                    'invalid_rows_sample': invalid_rows_sample,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        candidate_hashes = [row['dedup_key_hash'] for row in parsed_valid_rows]
        try:
            existing_hashes = set(
                OpenSourceDeficiencyRecord.objects.filter(
                    dedup_key_hash__in=candidate_hashes
                ).values_list('dedup_key_hash', flat=True)
            )
        except ProgrammingError as exc:
            logger.exception('OpenSource import dedup lookup failed filename=%s', filename)
            if OpenSourceDeficiencyRecord._meta.db_table in str(exc):
                return self._missing_storage_response()
            return self._server_error_response()
        except DatabaseError:
            logger.exception('OpenSource import dedup lookup database error filename=%s', filename)
            return self._server_error_response()

        duplicate_rows_sample = []
        duplicate_rows = 0
        pending_hashes = set()
        records_to_insert = []

        try:
            with transaction.atomic():
                import_run = OpenSourceImportRun.objects.create(
                    uploaded_by=str(getattr(request.user, 'id', '')),
                    filename=filename,
                    file_hash=file_hash,
                    total_rows=total_rows,
                    valid_rows=valid_rows,
                    inserted_rows=0,
                    duplicate_rows=0,
                    invalid_rows=invalid_rows,
                )

                for parsed in parsed_valid_rows:
                    dedup_key_hash = parsed['dedup_key_hash']

                    duplicate_reason = None
                    if dedup_key_hash in existing_hashes:
                        duplicate_reason = 'duplicate_existing_record'
                    elif dedup_key_hash in pending_hashes:
                        duplicate_reason = 'duplicate_within_uploaded_file'

                    if duplicate_reason:
                        duplicate_rows += 1
                        if len(duplicate_rows_sample) < MAX_TROUBLESHOOT_ROWS:
                            duplicate_rows_sample.append(
                                {
                                    'row_number': parsed['row_number'],
                                    'reason': duplicate_reason,
                                    'year': parsed['year'],
                                    'def_code_norm': parsed['def_code_norm'],
                                    'action_code_norm': parsed['action_code_norm'],
                                    'port_norm': parsed['port_norm'],
                                    'mou_norm': parsed['mou_norm'],
                                }
                            )
                        continue

                    pending_hashes.add(dedup_key_hash)
                    records_to_insert.append(
                        OpenSourceDeficiencyRecord(
                            import_run=import_run,
                            year=parsed['year'],
                            def_code_norm=parsed['def_code_norm'],
                            action_code_norm=parsed['action_code_norm'],
                            port_norm=parsed['port_norm'],
                            mou_norm=parsed['mou_norm'],
                            country_norm=parsed['country_norm'],
                            description_raw=parsed['description_raw'],
                            dedup_key_hash=dedup_key_hash,
                        )
                    )

                if records_to_insert:
                    logger.info(
                        'OpenSource import bulk insert starting filename=%s records=%s',
                        filename,
                        len(records_to_insert),
                    )
                    OpenSourceDeficiencyRecord.objects.bulk_create(records_to_insert, batch_size=500)

                import_run.inserted_rows = len(records_to_insert)
                import_run.duplicate_rows = duplicate_rows
                import_run.save(update_fields=['inserted_rows', 'duplicate_rows'])
        except (DataError, IntegrityError) as exc:
            logger.exception('OpenSource import rejected by DB constraints filename=%s', filename)
            return Response(
                {
                    'error': 'VALIDATION_ERROR',
                    'message': 'Uploaded file has values that violate import constraints.',
                    'details': str(exc),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except ProgrammingError as exc:
            logger.exception('OpenSource import database programming error filename=%s', filename)
            if (
                OpenSourceDeficiencyRecord._meta.db_table in str(exc)
                or OpenSourceImportRun._meta.db_table in str(exc)
            ):
                return self._missing_storage_response()
            return self._server_error_response()
        except DatabaseError:
            logger.exception('OpenSource import database error filename=%s', filename)
            return self._server_error_response()

        logger.info(
            'OpenSource import completed filename=%s total=%s valid=%s inserted=%s duplicate=%s invalid=%s',
            filename,
            import_run.total_rows,
            import_run.valid_rows,
            import_run.inserted_rows,
            import_run.duplicate_rows,
            import_run.invalid_rows,
        )
        return Response(
            {
                'data': {
                    'import_run_id': str(import_run.id),
                    'total_rows': import_run.total_rows,
                    'valid_rows': import_run.valid_rows,
                    'inserted_rows': import_run.inserted_rows,
                    'duplicate_rows': import_run.duplicate_rows,
                    'invalid_rows': import_run.invalid_rows,
                    'invalid_rows_sample': invalid_rows_sample,
                    'duplicate_rows_sample': duplicate_rows_sample,
                },
                'message': 'OpenSource import completed successfully.',
            },
            status=status.HTTP_201_CREATED,
        )
