from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


PRINT_COLUMNS = [
    "Section",
    "Sub No.",
    "Certificate / Survey",
    "Certificate No.",
    "Issued By",
    "Issue Date",
    "Expiry Date",
    "Last Done",
    "Next Due",
    "Validity",
    "Status",
]


def render_print_excel(
    *,
    print_id: str,
    rows: list[dict[str, Any]],
    payload: dict[str, Any],
    system_state_hash: str,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SQE S 633"
    worksheet.append(["SQE S 633", "Certificates and Surveys"])
    worksheet.append(["Print ID", print_id])
    worksheet.append(["Scope", payload.get("scope")])
    worksheet.append(["System state hash", system_state_hash])
    worksheet.append([])
    worksheet.append(PRINT_COLUMNS)
    header_row = 6
    for cell in worksheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")

    for index, row in enumerate(_ordered_rows(rows), start=1):
        worksheet.append(
            [
                row.get("catalog_section_name") or row.get("catalog_section_code") or "",
                index,
                row.get("catalog_display_name") or row.get("catalog_code") or "",
                row.get("certificate_number") or "",
                row.get("issuing_authority") or "",
                _date_value(row.get("issue_date")),
                _date_value(row.get("expiry_date")) or "PERM",
                _date_value(row.get("last_done_date")),
                _date_value(row.get("next_due_date")),
                row.get("validity_type") or "",
                row.get("status") or "",
            ]
        )

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 42)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _ordered_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            str(row.get("vessel_name") or ""),
            int(row.get("catalog_print_order") or 0),
            str(row.get("catalog_display_name") or ""),
        ),
    )


def _date_value(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value)[:10]
