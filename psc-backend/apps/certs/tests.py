from django.conf import settings
from django.core.exceptions import SuspiciousFileOperation
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase, override_settings
from django.urls import resolve, reverse
from contextlib import nullcontext
from datetime import datetime, timezone
from decimal import Decimal
from django.db import DatabaseError
from io import BytesIO
import importlib
import inspect
import json
from pathlib import Path
from rest_framework.test import APIRequestFactory, force_authenticate
import tempfile
from types import SimpleNamespace
from unittest import TestCase
from unittest.mock import patch

from apps.certs.class_code_mapping_seed import KR_CLASS_CODE_MAPPING_ROWS, validate_class_code_mapping_seed_rows
from apps.certs.permissions import can_approve_tracked_item
from apps.certs.serializers.catalog import CatalogRowWriteSerializer
from apps.certs.services.catalog_repository import CatalogRepository
from apps.certs.services.ocr_pipeline import (
    DEFAULT_OCR_ENGINE_NAME,
    OcrEngineOutput,
    OcrFieldCandidate,
    OcrThresholds,
    _paddle_prediction_to_text,
    process_cert_pdf,
)
from apps.certs.services.parsers.base import (
    CLASS_SNAPSHOT_OCR_DET_LIMIT_SIDE_LEN,
    CLASS_SNAPSHOT_OCR_LANGUAGE,
    CLASS_SNAPSHOT_OCR_MAX_PAGES,
    CLASS_SNAPSHOT_OCR_REC_BATCH_SIZE,
    CLASS_SNAPSHOT_OCR_VERSION,
    BaseClassParser,
    ClassSnapshotParseError,
    ExtractedClassSnapshotText,
    extract_pdf_embedded_image_ocr_text,
    extract_pdf_image_ocr_text,
)
from apps.certs.services.parsers.bv import BVClassParser
from apps.certs.services.parsers.kr import KRClassParser
from apps.certs.services.parsers.nk import NKClassParser
from apps.certs.services.pdf_blob_storage import resolve_pdf_blob_path
from apps.certs.services.pdf_blob_repository import PdfBlobRepository
from apps.certs.services.print_delivery import PrintArtifactDeliveryService
from apps.certs.services.excel_renderer import render_print_excel
from apps.certs.services.pdf_renderer import ReportLabPdfRenderer
from apps.certs.services.reconciliation import (
    ReconciliationRepository,
    build_reconciliation_flags,
    dispatch_parser_anomaly_notifications,
)
from apps.certs.services.notification_dispatcher import CertNotificationRecipient
from apps.certs.services.audit_log_repository import AuditLogRepository
from apps.certs.services.snapshot_repository import ClassSnapshotRepository
from apps.certs.services.settings_repository import SettingsRepository
from apps.certs.services.slack_relay import (
    CertSlackRelay,
    DEFAULT_CERTS_SLACK_CHANNEL,
    DEFAULT_DPA_SLACK_CHANNEL,
    DEFAULT_MARINE_SLACK_CHANNEL,
    DEFAULT_OFFICE_SLACK_CHANNEL,
    DEFAULT_TECHNICAL_SLACK_CHANNEL,
)
from apps.certs.services.vessel_dashboard import _serialize_snapshot
from apps.certs.services import ocr_pipeline
from apps.certs.services.tracked_item_repository import TrackedItemRepository
from apps.certs.services.audit_log import record_audit_event
from apps.certs.serializers.print import ShareBundleRequestSerializer, serialize_print_artifact
from apps.certs.serializers.tracked_item import TrackedItemWriteSerializer, serialize_tracked_item
from apps.certs.views import notification_views, print_views, reconciliation_views, snapshot_views, tracked_item_views

notification_compat_migration = importlib.import_module(
    "apps.certs.migrations.0005_master_notification_certs_columns"
)


class CertsAppRegistrationTests(SimpleTestCase):
    def test_certs_app_is_registered(self):
        self.assertIn("apps.certs", settings.INSTALLED_APPS)

    def test_certs_routes_are_mounted(self):
        self.assertEqual(reverse("certs:health"), "/api/certs/health/")
        self.assertEqual(resolve("/api/certs/health/").url_name, "health")
        self.assertEqual(
            resolve("/api/certs/print/artifacts/SQE-S633-TEST/download/zip/").url_name,
            "print-artifact-download",
        )
        self.assertEqual(
            reverse("certs:tracked-item-reparse-pdf", kwargs={"tracked_item_id": "tracked-1"}),
            "/api/certs/tracked-items/tracked-1/reparse-pdf/",
        )
        self.assertEqual(
            reverse("certs:class-snapshot-pdf-view", kwargs={"snapshot_id": "11111111-1111-1111-1111-111111111111"}),
            "/api/certs/class-snapshots/11111111-1111-1111-1111-111111111111/pdf/view/",
        )
        self.assertEqual(
            resolve("/api/certs/class-snapshots/BE1F386E-A689-F111-ADEC-FDB8CC7078D1/pdf/view/").url_name,
            "class-snapshot-pdf-view",
        )
        self.assertEqual(
            resolve("/api/certs/audit-log/1080624C-3C8B-F111-ADEC-FDB8CC7078D1/").url_name,
            "audit-log-detail",
        )
        self.assertEqual(
            reverse("certs-auditor:signup", kwargs={"token": "sample"}),
            "/api/auditor/signup/sample/",
        )


class CertSlackRelayTests(SimpleTestCase):
    def test_builtin_certs_slack_routes_use_production_channel_id(self):
        self.assertEqual(DEFAULT_CERTS_SLACK_CHANNEL, "C0BMCASMNKS")
        self.assertEqual(DEFAULT_OFFICE_SLACK_CHANNEL, "C0BMCASMNKS")
        self.assertEqual(DEFAULT_DPA_SLACK_CHANNEL, "C0BMCASMNKS")
        self.assertEqual(DEFAULT_TECHNICAL_SLACK_CHANNEL, "C0BMCASMNKS")
        self.assertEqual(DEFAULT_MARINE_SLACK_CHANNEL, "C0BMCASMNKS")

    def test_missing_slack_token_reports_failed_delivery_to_vims_certs_channel(self):
        with patch.dict("os.environ", {}, clear=True):
            relay = CertSlackRelay()

        status = relay.send_office_notification(
            channel="",
            title="Cert alert",
            message="Certificate requires review.",
            payload={"certRowId": "cert-1"},
        )

        self.assertEqual(status["channel"], "slack")
        self.assertEqual(status["status"], "failed")
        self.assertEqual(status["slackChannel"], "C0BMCASMNKS")
        self.assertEqual(status["error"], "SLACK_BOT_TOKEN not configured")


class CatalogRepositoryPaginationTests(TestCase):
    def test_list_rows_uses_offset_fetch_when_page_requested(self):
        cursor = _FakeCatalogCursor()

        with patch("apps.certs.services.catalog_repository.connection.cursor", return_value=cursor):
            page = CatalogRepository().list_rows(is_active=True, page=2, page_size=25)

        self.assertEqual(page.count, 123)
        self.assertEqual(page.page, 2)
        self.assertEqual(page.page_size, 25)
        self.assertEqual(page.results, [{"catalog_id": "catalog-1"}])
        self.assertIn("OFFSET %s ROWS FETCH NEXT %s ROWS ONLY", cursor.executed[1][0])
        self.assertEqual(cursor.executed[1][1][-2:], [25, 25])


class TrackedItemRepositoryFilterTests(TestCase):
    def test_list_items_can_filter_by_approval_state(self):
        cursor = _FakeTrackedItemCursor()

        with patch("apps.certs.services.tracked_item_repository.connection.cursor", return_value=cursor):
            page = TrackedItemRepository().list_items(approval_state="pending_master_approval")

        self.assertEqual(page.count, 0)
        self.assertIn("t.approval_state = %s", cursor.executed[0][0])
        self.assertIn("t.approval_state = %s", cursor.executed[1][0])
        self.assertEqual(cursor.executed[0][1], ["pending_master_approval"])

    def test_list_items_uses_bounded_sql_pagination_when_requested(self):
        cursor = _FakeTrackedItemCursor(count=2754)

        with patch("apps.certs.services.tracked_item_repository.connection.cursor", return_value=cursor):
            page = TrackedItemRepository().list_items(page=3, page_size=250)

        self.assertEqual(page.count, 2754)
        self.assertEqual(page.page, 3)
        self.assertEqual(page.page_size, 100)
        self.assertIn("OFFSET %s ROWS FETCH NEXT %s ROWS ONLY", cursor.executed[1][0])
        self.assertEqual(cursor.executed[1][1][-2:], [200, 100])


class TrackedItemSerializerPerformanceTests(SimpleTestCase):
    def test_lightweight_list_serialization_skips_principal_display_lookup(self):
        row = _tracked_item_serializer_row()

        with patch("apps.certs.serializers.tracked_item.resolve_principal_display_name") as resolver:
            serialized = serialize_tracked_item(row, include_display_names=False)

        resolver.assert_not_called()
        self.assertEqual(serialized["createdBy"], "seed-user")
        self.assertIsNone(serialized["createdByDisplay"])


class CertsListPayloadRepositoryTests(TestCase):
    def test_class_snapshot_list_omits_parsed_payload_json(self):
        cursor = _FakeRepositoryListCursor(count=7)

        with patch("apps.certs.services.snapshot_repository.connection.cursor", return_value=cursor):
            ClassSnapshotRepository().list_snapshots(page=1, page_size=25)

        list_sql = cursor.executed[1][0]
        self.assertIn("CAST(NULL AS NVARCHAR(MAX)) AS parsed_payload_json", list_sql)
        self.assertNotIn("s.parsed_payload_json, s.parsed_payload_schema_version", list_sql)

    def test_reconciliation_run_list_omits_flags_json(self):
        cursor = _FakeRepositoryListCursor(count=7)

        with patch("apps.certs.services.reconciliation.connection.cursor", return_value=cursor):
            ReconciliationRepository().list_runs(page=1, page_size=25)

        list_sql = cursor.executed[1][0]
        self.assertNotIn("r.flags_json", list_sql)
        self.assertIn("r.notifications_sent_json", list_sql)

    def test_audit_log_list_omits_large_json_columns(self):
        cursor = _FakeRepositoryListCursor(count=626)

        with patch("apps.certs.services.audit_log_repository.connection.cursor", return_value=cursor):
            AuditLogRepository().list_events(filters={"page": 1, "pageSize": 25}, vessel_scope=None)

        list_sql = cursor.executed[1][0]
        self.assertIn("CAST(NULL AS NVARCHAR(MAX)) AS before_json", list_sql)
        self.assertIn("CAST(NULL AS NVARCHAR(MAX)) AS after_json", list_sql)
        self.assertIn("CAST(NULL AS NVARCHAR(MAX)) AS event_metadata", list_sql)


class CertsSettingsRepositoryTests(TestCase):
    def test_slack_routes_use_vesseldata_brownfield_column_names(self):
        cursor = _FakeRepositoryListCursor()

        with patch("apps.certs.services.settings_repository.connection.cursor", return_value=cursor):
            SettingsRepository().list_slack_routes()

        sql = cursor.executed[0][0]
        self.assertIn("v.vesselName AS vessel_name", sql)
        self.assertIn("v.imoNumber AS imo_number", sql)
        self.assertNotIn("v.vessel_name", sql)
        self.assertNotIn("v.imo_number", sql)


class CertNotificationListPaginationTests(SimpleTestCase):
    def test_notification_list_uses_sql_pagination(self):
        cursor = _FakeRepositoryListCursor(count=626)
        user = SimpleNamespace(is_authenticated=True, user_id="office-user")
        request = APIRequestFactory().get("/api/certs/notifications/?page=2&page_size=25")
        force_authenticate(request, user=user)

        with patch("apps.certs.views.notification_views.connection", SimpleNamespace(vendor="microsoft", cursor=lambda: cursor)):
            response = notification_views.CertNotificationListView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["pagination"]["total_count"], 626)
        self.assertEqual(response.data["pagination"]["total_pages"], 26)
        self.assertIn("OFFSET %s ROWS FETCH NEXT %s ROWS ONLY", cursor.executed[1][0])
        self.assertEqual(cursor.executed[1][1][-2:], [25, 25])


class TrackedItemApprovalAuthorityTests(SimpleTestCase):
    def test_dpa_can_approve_when_vessel_access_is_global(self):
        user = SimpleNamespace(user_type="OFFICE", role="DPA", has_global_vessel_access=True)

        self.assertTrue(can_approve_tracked_item(user, {"vessel_id": "VESSEL-1"}))

    def test_pic_can_approve_when_vessel_is_assigned(self):
        user = SimpleNamespace(user_type="OFFICE", role="OFFICE_PIC", vessel_ids=["VESSEL-1"])

        self.assertTrue(can_approve_tracked_item(user, {"vessel_id": "VESSEL-1"}))

    def test_non_approval_office_role_cannot_approve(self):
        user = SimpleNamespace(user_type="OFFICE", role="CHIEF ACCOUNTING OFFICER", has_global_vessel_access=True)

        self.assertFalse(can_approve_tracked_item(user, {"vessel_id": "VESSEL-1"}))


class AuditLogWriteTests(SimpleTestCase):
    def test_non_uuid_entity_id_is_stored_as_metadata_entity_ref(self):
        cursor = _FakeAuditLogCursor()
        actor = SimpleNamespace(user_id="Harman.S", role="DPA")

        with patch("apps.certs.services.audit_log.connection", SimpleNamespace(cursor=lambda: cursor)):
            record_audit_event(
                actor=actor,
                action="print_artifact_created",
                entity_type="print_artifact",
                entity_id="SQE-S633-9584293-20260723-009",
                vessel_id="11111111-1111-1111-1111-111111111111",
                before=None,
                after={"printId": "SQE-S633-9584293-20260723-009"},
                reason="Generated SQE S 633 print artifact.",
                metadata={"source": "api.certs.print"},
            )

        params = cursor.executed[0][1]
        self.assertIsNone(params[5])
        metadata = json.loads(params[9])
        self.assertEqual(metadata["source"], "api.certs.print")
        self.assertEqual(metadata["entityRef"], "SQE-S633-9584293-20260723-009")

    def test_uuid_entity_id_is_stored_in_entity_id_column(self):
        cursor = _FakeAuditLogCursor()
        actor = SimpleNamespace(user_id="Harman.S", role="DPA")
        entity_id = "22222222-2222-2222-2222-222222222222"

        with patch("apps.certs.services.audit_log.connection", SimpleNamespace(cursor=lambda: cursor)):
            record_audit_event(
                actor=actor,
                action="upload_class_snapshot",
                entity_type="class_status_snapshot",
                entity_id=entity_id,
                vessel_id="11111111-1111-1111-1111-111111111111",
                before=None,
                after={"id": entity_id},
                reason="Uploaded class snapshot.",
                metadata={"source": "api.certs.class_snapshots"},
            )

        params = cursor.executed[0][1]
        self.assertEqual(params[5], entity_id)
        metadata = json.loads(params[9])
        self.assertNotIn("entityRef", metadata)


class MasterNotificationCompatibilityMigrationTests(SimpleTestCase):
    def test_migration_guards_every_certs_master_notification_column(self):
        required_columns = {
            "module_code",
            "record_id",
            "recipient_ref",
            "notification_kind",
            "title",
            "message",
            "delivery_channel",
            "payload_json",
            "created_at",
        }

        migration_columns = {
            column_name
            for column_name, _column_definition in notification_compat_migration.MASTER_NOTIFICATION_CERTS_COLUMNS
        }
        migration_source = inspect.getsource(notification_compat_migration.add_master_notification_certs_columns)

        self.assertEqual(migration_columns, required_columns)
        self.assertIn("OBJECT_ID", migration_source)
        self.assertIn("COL_LENGTH", migration_source)


class PdfBlobStoragePathTests(SimpleTestCase):
    def test_resolves_pdf_blob_path_inside_upload_root(self):
        with tempfile.TemporaryDirectory() as upload_root, override_settings(UPLOAD_BASE_PATH=upload_root):
            resolved = resolve_pdf_blob_path({"blob_storage_path": "certs/vessel-1/class.pdf"})

        self.assertEqual(resolved, Path(upload_root).resolve() / "certs" / "vessel-1" / "class.pdf")

    def test_rejects_pdf_blob_path_outside_upload_root(self):
        with tempfile.TemporaryDirectory() as upload_root, override_settings(UPLOAD_BASE_PATH=upload_root):
            with self.assertRaises(SuspiciousFileOperation):
                resolve_pdf_blob_path({"blob_storage_path": "../outside.pdf"})


class PrintArtifactSerializationTests(SimpleTestCase):
    def test_print_artifact_serializer_exposes_download_urls(self):
        serialized = serialize_print_artifact(
            _print_artifact_row(
                print_id="SQE-S633-TEST-001",
                pdf_blob_id="pdf-1",
                excel_blob_id="excel-1",
                bundle_zip_blob_id="zip-1",
            )
        )

        self.assertEqual(
            serialized["downloadUrls"]["pdf"],
            "/api/certs/print/artifacts/SQE-S633-TEST-001/download/pdf/",
        )
        self.assertEqual(
            serialized["downloadUrls"]["excel"],
            "/api/certs/print/artifacts/SQE-S633-TEST-001/download/excel/",
        )
        self.assertEqual(
            serialized["downloadUrls"]["zip"],
            "/api/certs/print/artifacts/SQE-S633-TEST-001/download/zip/",
        )


class ShareBundleSerializerTests(SimpleTestCase):
    def test_accepts_sections_without_certificate_ids(self):
        serializer = ShareBundleRequestSerializer(
            data={
                "vesselIds": ["11111111-1111-1111-1111-111111111111"],
                "sections": ["STATUTORY"],
                "customCertIds": [],
                "recipientEmail": "agent@example.com",
            }
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
        self.assertEqual(serializer.validated_data["scope"], "share_bundle")
        self.assertEqual(serializer.validated_data["sections"], ["STATUTORY"])
        self.assertEqual(serializer.validated_data["customCertIds"], [])

    def test_requires_sections_or_certificate_ids(self):
        serializer = ShareBundleRequestSerializer(
            data={
                "vesselIds": ["11111111-1111-1111-1111-111111111111"],
                "sections": [],
                "customCertIds": [],
            }
        )

        self.assertFalse(serializer.is_valid())
        self.assertIn("sections", serializer.errors)


class PrintArtifactDeliveryTests(SimpleTestCase):
    def test_print_generation_view_sends_email_when_recipient_email_is_present(self):
        vessel_id = "11111111-1111-1111-1111-111111111111"
        artifact = _print_artifact_row(vessels=[vessel_id], recipient_email="agent@example.com")
        user = SimpleNamespace(
            is_authenticated=True,
            user_type="OFFICE",
            role="DPA",
            form_ids=["CERT_F_004"],
            process_ids=["CERT_P_005"],
            has_global_vessel_access=True,
        )
        request = APIRequestFactory().post(
            "/api/certs/print/",
            data={
                "scope": "per_vessel_full",
                "vesselIds": [vessel_id],
                "recipientEmail": "agent@example.com",
            },
            format="json",
        )
        force_authenticate(request, user=user)
        email_result = {
            "status": "sent",
            "message": "Email sent to agent@example.com.",
            "recipient": "agent@example.com",
            "attachmentKinds": ["pdf", "excel"],
        }

        with (
            patch.object(print_views.service, "generate_print", return_value=artifact),
            patch.object(print_views.delivery_service, "send_artifact_email", return_value=email_result) as send_email,
            patch("apps.certs.views.print_views.record_audit_event") as audit,
        ):
            response = print_views.PrintArtifactCreateView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["emailDeliveryStatus"], "sent")
        self.assertEqual(response.data["emailDeliveryMessage"], "Email sent to agent@example.com.")
        send_email.assert_called_once_with(artifact)
        self.assertEqual(audit.call_args.kwargs["metadata"]["emailDeliveryStatus"], "sent")

    def test_delivery_service_attaches_pdf_and_excel_to_print_email(self):
        with tempfile.TemporaryDirectory() as upload_root, override_settings(UPLOAD_BASE_PATH=upload_root, DEFAULT_FROM_EMAIL="certs@example.com"):
            _write_upload_blob(upload_root, "certs/prints/report.pdf", b"pdf-bytes")
            _write_upload_blob(upload_root, "certs/prints/report.xlsx", b"excel-bytes")
            connection_factory = _FakeEmailConnectionFactory()
            service = PrintArtifactDeliveryService(
                blob_repository=_FakePdfBlobRepository(
                    {
                        "pdf-1": {"blob_storage_path": "certs/prints/report.pdf", "filename": "report.pdf"},
                        "excel-1": {"blob_storage_path": "certs/prints/report.xlsx", "filename": "report.xlsx"},
                    }
                ),
                email_connection_factory=connection_factory,
            )

            result = service.send_artifact_email(
                _print_artifact_row(pdf_blob_id="pdf-1", excel_blob_id="excel-1", recipient_email="agent@example.com")
            )

        message = connection_factory.connections[0].sent_messages[0]
        self.assertEqual(result["status"], "sent")
        self.assertEqual(message.to, ["agent@example.com"])
        self.assertEqual(message.from_email, "certs@example.com")
        self.assertEqual([attachment[0] for attachment in message.attachments], ["report.pdf", "report.xlsx"])

    def test_delivery_service_attaches_zip_only_to_share_bundle_email(self):
        with tempfile.TemporaryDirectory() as upload_root, override_settings(UPLOAD_BASE_PATH=upload_root, DEFAULT_FROM_EMAIL="certs@example.com"):
            _write_upload_blob(upload_root, "certs/prints/bundle.zip", b"zip-bytes")
            connection_factory = _FakeEmailConnectionFactory()
            service = PrintArtifactDeliveryService(
                blob_repository=_FakePdfBlobRepository(
                    {
                        "zip-1": {"blob_storage_path": "certs/prints/bundle.zip", "filename": "bundle.zip"},
                    }
                ),
                email_connection_factory=connection_factory,
            )

            result = service.send_artifact_email(
                _print_artifact_row(scope="share_bundle", bundle_zip_blob_id="zip-1", recipient_email="agent@example.com")
            )

        message = connection_factory.connections[0].sent_messages[0]
        self.assertEqual(result["status"], "sent")
        self.assertEqual(result["attachmentKinds"], ["zip"])
        self.assertEqual([attachment[0] for attachment in message.attachments], ["bundle.zip"])

    def test_download_view_streams_requested_artifact_file(self):
        vessel_id = "11111111-1111-1111-1111-111111111111"
        user = SimpleNamespace(
            is_authenticated=True,
            user_type="OFFICE",
            role="DPA",
            form_ids=["CERT_F_004"],
            process_ids=[],
            has_global_vessel_access=True,
        )
        request = APIRequestFactory().get("/api/certs/print/artifacts/SQE-S633-TEST/download/zip/")
        force_authenticate(request, user=user)

        with tempfile.TemporaryDirectory() as upload_root:
            path = Path(upload_root) / "bundle.zip"
            path.write_bytes(b"zip-bytes")
            artifact_file = SimpleNamespace(
                absolute_path=path,
                filename="bundle.zip",
                content_type="application/zip",
            )
            with (
                patch.object(print_views.repository, "get_artifact", return_value=_print_artifact_row(vessels=[vessel_id])),
                patch.object(print_views.delivery_service, "get_download_file", return_value=artifact_file) as get_download_file,
            ):
                response = print_views.PrintArtifactDownloadView.as_view()(
                    request,
                    print_id="SQE-S633-TEST",
                    artifact_kind="zip",
                )

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response["Content-Type"], "application/zip")
            self.assertIn("attachment", response["Content-Disposition"])
            self.assertIn("bundle.zip", response["Content-Disposition"])
            get_download_file.assert_called_once()
            response.close()


class TrackedItemPdfReparseViewTests(SimpleTestCase):
    def test_reparse_active_pdf_updates_ocr_payload_and_auto_fields(self):
        user = SimpleNamespace(
            is_authenticated=True,
            user_type="OFFICE",
            role="DPA",
            form_ids=["CERT_F_002"],
            process_ids=["CERT_P_001"],
            has_global_vessel_access=True,
        )
        request = APIRequestFactory().post(
            "/api/certs/tracked-items/tracked-1/reparse-pdf/",
            data={"context": "office", "reason": "Read the uploaded PDF again."},
            format="json",
        )
        force_authenticate(request, user=user)

        ocr_payload = {
            "engine": DEFAULT_OCR_ENGINE_NAME,
            "status": "processed",
            "fields": {
                "certificate_number": {"value": "CERT-NEW", "confidence": 0.99, "mode": "auto_accept"},
                "expiry_date": {"value": "2030-07-24", "confidence": 0.98, "mode": "auto_accept"},
            },
        }
        item_before = _tracked_item_test_row(status="pending_first_upload", certificate_number=None, expiry_date=None)
        item_after = _tracked_item_test_row(status="ok", certificate_number="CERT-NEW", expiry_date="2030-07-24", version=2)

        class FakeTrackedItemRepository:
            def __init__(self):
                self.update_values = None

            def get_item(self, tracked_item_id):
                return item_before if tracked_item_id == "tracked-1" else None

            def update_item(self, tracked_item_id, values, *, actor_id):
                self.update_values = values
                self.actor_id = actor_id
                return item_before, item_after

        class FakePdfRepository:
            def __init__(self, before, after):
                self.before = before
                self.after = after
                self.updated_payload = None

            def get_blob(self, blob_id):
                return self.before if blob_id == "blob-1" else None

            def update_ocr_result(self, blob_id, payload):
                self.updated_payload = payload
                return self.after

        with tempfile.TemporaryDirectory() as upload_root, override_settings(UPLOAD_BASE_PATH=upload_root):
            relative_path = "certs/vessels/vessel-1/tracked-items/tracked-1/certificate.pdf"
            _write_upload_blob(upload_root, relative_path, b"%PDF-1.4")
            blob_before = _pdf_blob_test_row(blob_storage_path=relative_path, ocr_payload_json={})
            blob_after = _pdf_blob_test_row(blob_storage_path=relative_path, ocr_payload_json=ocr_payload)
            fake_repository = FakeTrackedItemRepository()
            fake_pdf_repository = FakePdfRepository(blob_before, blob_after)

            with (
                patch.object(tracked_item_views, "repository", fake_repository),
                patch.object(tracked_item_views, "pdf_repository", fake_pdf_repository),
                patch.object(tracked_item_views, "process_cert_pdf", return_value=ocr_payload) as process_pdf,
                patch.object(tracked_item_views, "record_audit_event") as audit,
                patch.object(tracked_item_views, "record_cert_change_log") as change_log,
                patch.object(tracked_item_views, "record_approval_event") as approval_event,
                patch.object(tracked_item_views.transaction, "atomic", return_value=nullcontext()),
            ):
                response = tracked_item_views.TrackedItemReparsePdfView.as_view()(request, tracked_item_id="tracked-1")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["trackedItem"]["certificateNumber"], "CERT-NEW")
        self.assertEqual(response.data["pdfBlob"]["id"], "blob-1")
        self.assertEqual(response.data["ocrPayload"], ocr_payload)
        process_pdf.assert_called_once()
        self.assertEqual(process_pdf.call_args.kwargs["context"], "office")
        self.assertEqual(fake_pdf_repository.updated_payload, ocr_payload)
        self.assertEqual(fake_repository.update_values["certificateNumber"], "CERT-NEW")
        self.assertEqual(fake_repository.update_values["expiryDate"], "2030-07-24")
        self.assertEqual(fake_repository.update_values["status"], "ok")
        self.assertFalse(fake_repository.update_values["pdfMissing"])
        self.assertEqual(audit.call_count, 2)
        self.assertEqual(change_log.call_count, 1)
        approval_event.assert_not_called()


class PdfBlobRepositoryTests(SimpleTestCase):
    def test_duplicate_sha_lookup_only_blocks_active_pdf_versions(self):
        class FakeCursor:
            description = [("blob_id",)]

            def __init__(self):
                self.sql = ""
                self.params = None

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, traceback):
                return False

            def execute(self, sql, params):
                self.sql = sql
                self.params = params

            def fetchone(self):
                return None

        class FakeConnection:
            vendor = "microsoft"

            def __init__(self):
                self.cursor_instance = FakeCursor()

            def cursor(self):
                return self.cursor_instance

        fake_connection = FakeConnection()
        with patch("apps.certs.services.pdf_blob_repository.connection", fake_connection):
            PdfBlobRepository().get_blob_for_tracked_item_sha(
                tracked_item_id="tracked-1",
                content_sha256="same-sha",
            )

        self.assertIn("AND is_active = 1", fake_connection.cursor_instance.sql)
        self.assertEqual(fake_connection.cursor_instance.params, ["tracked-1", "same-sha"])


class ShareBundleManifestRendererTests(SimpleTestCase):
    def test_manifest_removes_internal_header_footer_and_uses_printed_by_label(self):
        result = ReportLabPdfRenderer().render_share_bundle_manifest(
            print_id="SQE-S633-TEST-001",
            rows=[
                {
                    "vessel_name": "MV Test",
                    "vessel_imo": "1234567",
                    "vessel_flag": "PANAMA",
                    "class_society": "KR",
                    "catalog_display_name": "Safety Management Certificate",
                    "catalog_print_order": 1,
                    "tracked_item_id": "tracked-1",
                    "issue_date": "2026-01-01",
                    "expiry_date": "2031-01-01",
                    "blob_filename": "smc.pdf",
                }
            ],
            payload={"scope": "share_bundle", "watermarkApplied": "MASTER_COPY"},
            actor_id="Harman.S",
            actor_role="DPA",
            system_state_hash="ABC12345",
        )

        from PyPDF2 import PdfReader

        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(result.content)).pages)
        self.assertIn("Printed by: Harman.S (DPA)", text)
        self.assertIn("Safety Management Certificate", text)
        self.assertNotIn("SQE S 633 - Master Share Bundle Manifest", text)
        self.assertNotIn("Print ID:", text)
        self.assertNotIn("Hash:", text)
        self.assertNotIn("Validity: F=Full", text)
        self.assertNotIn("MASTER COPY", text)
        self.assertNotIn("Generation Footer", text)
        self.assertNotIn("User:", text)


class PrintArtifactRendererTests(SimpleTestCase):
    def test_normal_print_pdf_removes_internal_metadata_and_prints_bottom_right_watermark(self):
        result = ReportLabPdfRenderer().render_print_artifact(
            print_id="SQE-S633-TEST-001",
            rows=[
                {
                    "vessel_name": "MV Test",
                    "vessel_imo": "1234567",
                    "vessel_flag": "PANAMA",
                    "class_society": "KR",
                    "catalog_section_name": "Statutory",
                    "catalog_display_name": "Safety Management Certificate",
                    "catalog_print_order": 1,
                    "tracked_item_id": "tracked-1",
                    "certificate_number": "SMC-001",
                    "issuing_authority": "Class Society",
                    "issue_date": "2026-01-01",
                    "expiry_date": "2031-01-01",
                    "last_done_date": "2026-01-01",
                    "next_due_date": "2031-01-01",
                    "validity_type": "full",
                    "status": "valid",
                }
            ],
            payload={
                "scope": "per_vessel_full",
                "watermarkApplied": "MASTER_COPY",
                "watermarkRecipient": "Port Agent",
            },
            actor_id="Harman.S",
            actor_role="DPA",
            system_state_hash="ABC12345",
        )

        from PyPDF2 import PdfReader

        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(result.content)).pages)
        self.assertIn("Printed by: Harman.S (DPA)", text)
        self.assertIn("Safety Management Certificate", text)
        self.assertIn("MASTER COPY", text)
        self.assertNotIn("SQE S 633 - Certificates and Surveys", text)
        self.assertNotIn("Scope:", text)
        self.assertNotIn("Print ID:", text)
        self.assertNotIn("SQE-S633-TEST-001", text)
        self.assertNotIn("Hash:", text)
        self.assertNotIn("ABC12345", text)
        self.assertNotIn("Validity", text)
        self.assertNotIn("Port Agent", text)
        self.assertNotIn("Generation Footer", text)
        self.assertNotIn("User:", text)


class PrintArtifactExcelRendererTests(SimpleTestCase):
    def test_print_excel_omits_internal_metadata_rows(self):
        content = render_print_excel(
            print_id="SQE-S633-TEST-001",
            rows=[
                {
                    "vessel_name": "MV Test",
                    "catalog_section_name": "Statutory",
                    "catalog_display_name": "Safety Management Certificate",
                    "catalog_print_order": 1,
                    "certificate_number": "SMC-001",
                    "issuing_authority": "Class Society",
                    "issue_date": "2026-01-01",
                    "expiry_date": "2031-01-01",
                    "last_done_date": "2026-01-01",
                    "next_due_date": "2031-01-01",
                    "validity_type": "full",
                    "status": "valid",
                }
            ],
            payload={"scope": "per_vessel_full"},
            system_state_hash="ABC12345",
        )

        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), data_only=True)
        worksheet = workbook.active
        visible_values = [
            str(cell.value)
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value not in (None, "")
        ]

        self.assertIn("SQE S 633", visible_values)
        self.assertIn("Certificate / Survey", visible_values)
        self.assertIn("Safety Management Certificate", visible_values)
        self.assertNotIn("Print ID", visible_values)
        self.assertNotIn("SQE-S633-TEST-001", visible_values)
        self.assertNotIn("Scope", visible_values)
        self.assertNotIn("per_vessel_full", visible_values)
        self.assertNotIn("System state hash", visible_values)
        self.assertNotIn("ABC12345", visible_values)


class ClassSnapshotUploadParsingTests(SimpleTestCase):
    def test_upload_immediately_reparses_and_returns_run_ready_snapshot(self):
        vessel_id = "11111111-1111-1111-1111-111111111111"
        snapshot_id = "22222222-2222-2222-2222-222222222222"
        run_id = "33333333-3333-3333-3333-333333333333"
        blob_id = "44444444-4444-4444-4444-444444444444"
        user = SimpleNamespace(
            is_authenticated=True,
            user_type="OFFICE",
            role="DPA",
            form_ids=["CERT_F_003"],
            process_ids=["CERT_P_001"],
            has_global_vessel_access=True,
        )
        request = APIRequestFactory().post(
            "/api/certs/class-snapshots/",
            data={
                "vesselId": vessel_id,
                "classSociety": "NK",
                "printedOnDate": "2026-07-20",
                "file": SimpleUploadedFile("class-status.pdf", b"%PDF-1.4\n%", content_type="application/pdf"),
            },
            format="multipart",
        )
        force_authenticate(request, user=user)
        pending_snapshot = _class_snapshot_row(
            snapshot_id=snapshot_id,
            vessel_id=vessel_id,
            blob_id=blob_id,
            parse_status="pending",
            reconciliation_run_id=None,
        )
        parsed_snapshot = _class_snapshot_row(
            snapshot_id=snapshot_id,
            vessel_id=vessel_id,
            blob_id=blob_id,
            parse_status="success",
            reconciliation_run_id=run_id,
            parser_version="nk-pdfplumber-v1",
        )

        with (
            patch(
                "apps.certs.views.snapshot_views.save_uploaded_class_snapshot_pdf",
                return_value={"relative_path": "certs/vessels/test/class-status.pdf", "filename": "class-status.pdf", "sha256": "sha", "size": 10},
            ),
            patch("apps.certs.views.snapshot_views.resolve_actor_id", return_value="actor-1"),
            patch.object(snapshot_views.pdf_repository, "create_snapshot_blob", return_value={"blob_id": blob_id}),
            patch.object(snapshot_views.repository, "create_snapshot", return_value=pending_snapshot) as create_snapshot,
            patch("apps.certs.views.snapshot_views.run_class_snapshot_parser", return_value=(parsed_snapshot, {"run_id": run_id})) as parser_worker,
            patch("apps.certs.views.snapshot_views.record_audit_event"),
            patch("apps.certs.views.snapshot_views.transaction.atomic", return_value=nullcontext()),
        ):
            response = snapshot_views.ClassSnapshotListCreateView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["parseStatus"], "success")
        self.assertEqual(response.data["reconciliationRunId"], run_id)
        create_snapshot.assert_called_once()
        self.assertEqual(str(create_snapshot.call_args.kwargs["printed_on_date"]), "2026-07-20")
        parser_worker.assert_called_once_with(snapshot_id, repository=snapshot_views.repository)

    def test_class_snapshot_pdf_view_streams_uploaded_snapshot_for_accessible_vessel(self):
        vessel_id = "11111111-1111-1111-1111-111111111111"
        snapshot_id = "22222222-2222-2222-2222-222222222222"
        blob_id = "44444444-4444-4444-4444-444444444444"
        user = SimpleNamespace(
            is_authenticated=True,
            user_type="VESSEL",
            role_name="MASTER",
            rank="MASTER",
            form_ids=["CERT_F_002"],
            process_ids=[],
            vessel_id=vessel_id,
        )
        request = APIRequestFactory().get(f"/api/certs/class-snapshots/{snapshot_id}/pdf/view/")
        force_authenticate(request, user=user)

        with tempfile.TemporaryDirectory() as upload_dir:
            relative_path = Path("certs") / "vessels" / vessel_id / "class-snapshots" / "class-status.pdf"
            absolute_path = Path(upload_dir) / relative_path
            absolute_path.parent.mkdir(parents=True, exist_ok=True)
            absolute_path.write_bytes(b"%PDF-1.4\n% class status")
            snapshot = _class_snapshot_row(
                snapshot_id=snapshot_id,
                vessel_id=vessel_id,
                blob_id=blob_id,
                parse_status="success",
                reconciliation_run_id="33333333-3333-3333-3333-333333333333",
            )
            blob = {
                "blob_id": blob_id,
                "blob_storage_path": relative_path.as_posix(),
                "filename": "class-status.pdf",
            }

            with (
                override_settings(UPLOAD_BASE_PATH=Path(upload_dir)),
                patch.object(snapshot_views.repository, "get_snapshot", return_value=snapshot),
                patch.object(snapshot_views.pdf_repository, "get_blob", return_value=blob),
            ):
                response = snapshot_views.ClassSnapshotPdfInlineView.as_view()(request, snapshot_id=snapshot_id)

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response["Content-Type"], "application/pdf")
            self.assertIn("inline", response["Content-Disposition"])
            self.assertIn("class-status.pdf", response["Content-Disposition"])
            response.close()


class ReconciliationMasterMessageApiTests(SimpleTestCase):
    def test_vessel_user_can_list_office_messages_for_own_vessel(self):
        vessel_id = "11111111-1111-1111-1111-111111111111"
        user = _vessel_master_user(vessel_id=vessel_id)
        request = APIRequestFactory().get("/api/certs/reconciliation/master-messages/")
        force_authenticate(request, user=user)
        row = _master_message_row(vessel_id=vessel_id, master_reviewed_at=None)

        with patch.object(
            reconciliation_views.repository,
            "list_master_messages",
            return_value={"count": 1, "results": [row]},
        ) as list_messages:
            response = reconciliation_views.ReconciliationMasterMessageListView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["id"], row["flag_id"])
        self.assertEqual(response.data["results"][0]["officeNote"], "Please check expiry date with the vessel file.")
        list_messages.assert_called_once_with(
            vessel_id=vessel_id,
            include_reviewed=False,
            page=1,
            page_size=50,
        )

    def test_master_can_mark_office_message_reviewed(self):
        vessel_id = "11111111-1111-1111-1111-111111111111"
        flag_id = "22222222-2222-2222-2222-222222222222"
        user = _vessel_master_user(vessel_id=vessel_id)
        request = APIRequestFactory().post(
            f"/api/certs/reconciliation/master-messages/{flag_id}/ack/",
            data={"note": "Checked with the onboard class file."},
            format="json",
        )
        force_authenticate(request, user=user)
        before = _master_message_row(flag_id=flag_id, vessel_id=vessel_id, master_reviewed_at=None)
        after = _master_message_row(
            flag_id=flag_id,
            vessel_id=vessel_id,
            master_reviewed_at="2026-07-22T10:15:00Z",
            master_reviewed_by="KSM0229",
            master_review_note="Checked with the onboard class file.",
        )

        with (
            patch.object(reconciliation_views.repository, "get_master_message", side_effect=[before, after]),
            patch("apps.certs.views.reconciliation_views.record_audit_event") as audit_event,
        ):
            response = reconciliation_views.ReconciliationMasterMessageAcknowledgeView.as_view()(request, flag_id=flag_id)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["masterReviewedAt"], "2026-07-22T10:15:00Z")
        audit_event.assert_called_once()
        self.assertEqual(audit_event.call_args.kwargs["action"], "reconciliation_review")
        self.assertEqual(audit_event.call_args.kwargs["entity_id"], flag_id)
        self.assertEqual(audit_event.call_args.kwargs["metadata"]["resolution_action"], "master_reviewed")

    def test_non_master_vessel_user_cannot_mark_office_message_reviewed(self):
        vessel_id = "11111111-1111-1111-1111-111111111111"
        flag_id = "22222222-2222-2222-2222-222222222222"
        user = _vessel_master_user(vessel_id=vessel_id, role="CHIEF OFFICER", rank="CHIEF OFFICER")
        request = APIRequestFactory().post(
            f"/api/certs/reconciliation/master-messages/{flag_id}/ack/",
            data={"note": "Checked."},
            format="json",
        )
        force_authenticate(request, user=user)

        with patch.object(reconciliation_views.repository, "get_master_message") as get_message:
            response = reconciliation_views.ReconciliationMasterMessageAcknowledgeView.as_view()(request, flag_id=flag_id)

        self.assertEqual(response.status_code, 403)
        get_message.assert_not_called()


class ClassSnapshotParserAndReconciliationSmokeTests(SimpleTestCase):
    def test_class_snapshot_parser_reports_ocr_fallback_without_readable_text(self):
        parser = _NoopClassParser()

        with (
            patch("apps.certs.services.parsers.base.extract_pdf_text", return_value=ExtractedClassSnapshotText("", 19, "paddleocr_fallback")),
            self.assertRaisesRegex(ClassSnapshotParseError, "OCR fallback read no text"),
        ):
            parser.parse("image-only-class-status.pdf")

    def test_class_snapshot_parser_uses_ocr_fallback_when_pdf_has_no_text_layer(self):
        ocr_text = """KOREAN REGISTER
Ship Name EAST AYUTTHAYA Work ID VANS004726
Class No. 1000010 IMO No. 9584293
Printed on 07-Jul-2026
Certificates
Class Certificates
Cargo Gear(CG2) Certificate CG2 Full 2026-02-27 2031-02-26 CL26001506350
Statutory Certificates or Documents of Compliance issued by KR
Load Line Certificate ILL Full 2025-08-11 2030-07-11 0N25015926559
"""

        with tempfile.NamedTemporaryFile(suffix=".pdf") as pdf_file:
            with (
                patch("pdfplumber.open", return_value=_EmptyPdfPlumberDocument(page_count=3)),
                patch("apps.certs.services.parsers.base.extract_pdf_image_ocr_text", return_value=ocr_text),
            ):
                parsed = KRClassParser().parse(pdf_file.name)

        self.assertEqual(parsed.parse_status, "success")
        self.assertEqual(parsed.payload["source"], "ocr_text")
        self.assertEqual(parsed.payload["text_extraction"]["engine"], "paddleocr_fallback")
        self.assertEqual(parsed.payload["vessel"]["name"], "EAST AYUTTHAYA")
        self.assertEqual(parsed.payload["vessel"]["imo"], "9584293")
        self.assertEqual(parsed.payload["printed_on_date"], "2026-07-07")
        self.assertGreaterEqual(len(parsed.payload["rows"]), 2)

    def test_class_snapshot_parser_requires_pdf_printed_date(self):
        ocr_text = """KOREAN REGISTER
Ship Name EAST AYUTTHAYA Work ID VANS004726
Class No. 1000010 IMO No. 9584293
Certificates
Class Certificates
Cargo Gear(CG2) Certificate CG2 Full 2026-02-27 2031-02-26 CL26001506350
"""

        with tempfile.NamedTemporaryFile(suffix=".pdf") as pdf_file:
            with (
                patch("pdfplumber.open", return_value=_EmptyPdfPlumberDocument(page_count=3)),
                patch("apps.certs.services.parsers.base.extract_pdf_image_ocr_text", return_value=ocr_text),
                self.assertRaisesRegex(ClassSnapshotParseError, "printed/generated date"),
            ):
                KRClassParser().parse(pdf_file.name)

    def test_class_snapshot_parser_uses_manual_report_date_only_when_pdf_date_is_unreadable(self):
        text = """KOREAN REGISTER
VESSEL STATUS FOR SHIP'S OWNER EAST AYUTTHAYA Class No : 1000010 IMO No : 9584293
Certificates
Class Certificates
Classification Certificate(Full) CC Full 2025-08-11 2030-07-11 CL25004603552
"""

        with patch(
            "apps.certs.services.parsers.base.extract_pdf_text",
            return_value=ExtractedClassSnapshotText(text, 19, "pdfplumber"),
        ):
            parsed = KRClassParser().parse("class-status.pdf", printed_on_date="2026-07-20")

        self.assertEqual(parsed.payload["printed_on_date"], "2026-07-20")
        self.assertEqual(parsed.payload["printed_on_date_source"], "manual")

    def test_kr_parser_reads_printed_date_from_first_ocr_page(self):
        self.assertEqual(KRClassParser.ocr_page_numbers, (1, 5, 6, 7, 8, 9))

        payload = KRClassParser().parse_text(
            """KOREAN REGISTER
VESSEL STATUS FOR SHIP'S OWNER EAST AYUTTHAYA Class No : 1000010 IMO No : 9584293
Printed on 07-Jul-2026
Certificates
Class Certificates
Classification Certificate(Full) CC Full 2025-08-11 2030-07-11 CL25004603552
""",
            page_count=19,
        )

        self.assertEqual(payload["printed_on_date"], "2026-07-07")

    def test_kr_parser_reads_only_class_conditions(self):
        payload = KRClassParser().parse_text(
            """KOREAN REGISTER
VESSEL STATUS FOR SHIP'S OWNER EAST AYUTTHAYA Class No : 1000010 IMO No : 9584293
Printed on 07-Jul-2026
C.1 2026-07-01 2026-08-13 KRREPORT001 Class
Auxiliary Boiler automation system repair/replacement condition
C.2 2026-07-01 2026-09-01 KRREPORT002 Statutory
Statutory note that should not appear as Condition of Class
Actionable Note
Actionable note that should not appear as Condition of Class
Informative Notes
None
""",
            page_count=19,
        )

        self.assertEqual(payload["printed_on_date"], "2026-07-07")
        self.assertEqual(len(payload["conditions_of_class"]), 1)
        condition = payload["conditions_of_class"][0]
        self.assertEqual(condition["id"], "C.1")
        self.assertEqual(condition["section"], "Condition of Class")
        self.assertEqual(condition["due_date"], "2026-08-13")
        self.assertIn("Auxiliary Boiler", condition["text"])
        self.assertFalse(any("Statutory note" in condition["text"] for condition in payload["conditions_of_class"]))
        self.assertFalse(any("Actionable note" in condition["text"] for condition in payload["conditions_of_class"]))

    def test_nk_parser_reads_condition_of_class_with_due_date(self):
        payload = NKClassParser().parse_text(
            """NK-SHIPS Information Service
SFYC ARAYA Class No. NK 107011 IMO No. 9487043
Printed on 27.Jul.2026
Survey Status:: Class
Hull Annual Survey 29 Jul 2026
Condition & Note
Condition of Class
The malfunctioned Auxiliary Boiler automation system is to be repaired/replaced at the owner earliest convenience
but no later than the below mentioned due date.
(DueDate: 13 Aug 2026 )
Note
Nil.
""",
            page_count=22,
        )

        self.assertEqual(payload["printed_on_date"], "2026-07-27")
        self.assertEqual(len(payload["conditions_of_class"]), 1)
        condition = payload["conditions_of_class"][0]
        self.assertEqual(condition["section"], "Condition of Class")
        self.assertEqual(condition["due_date"], "2026-08-13")
        condition_rows = [row for row in payload["rows"] if row["row_type"] == "condition"]
        self.assertEqual(condition_rows[0]["source_section"], "Condition of Class")

    def test_nk_parser_ignores_non_class_condition_sections(self):
        payload = NKClassParser().parse_text(
            """NK-SHIPS Information Service
SFYC ARAYA Class No. NK 107011 IMO No. 9487043
Printed on 27.Jul.2026
Condition & Note
Condition of Class
Nil.
Condition of Installation
Installation condition that should not appear as Condition of Class
(DueDate: 13 Aug 2026 )
Condition of Statutory Survey
Statutory survey condition that should not appear as Condition of Class
(DueDate: 14 Aug 2026 )
Note
Nil.
""",
            page_count=22,
        )

        self.assertEqual(payload["printed_on_date"], "2026-07-27")
        self.assertEqual(payload["conditions_of_class"], [])
        self.assertFalse(any(row["row_type"] == "condition" for row in payload["rows"]))

    def test_bv_parser_ignores_memoranda_when_conditions_of_class_is_empty(self):
        payload = BVClassParser().parse_text(
            """MOVE Fleet in Service Survey Status Report
Ship name: YC FORTITUDE BV Nr: 25272W IMO Number: 9587178
Generated on 27 Jul 2026 Page 3 / 4
Conditions of Class / Statutory Recommendations
None
ISM Code Non-Conformities
None
Class Memoranda
Issued Description of Memoranda
07 Sep 2024 The following damages found on the propeller blade are to be repaired by owner not later than next dry docking survey.
07 Sep 2024 Permanent engrave for BV letter at plimsol marking to be carried out at next dry-docking.
Statutory Memoranda
Issued Description of Statutory Memoranda
07 Sep 2024 Sampling point shall be fitted not later than the first renewal survey of IAPP certificate.
""",
            page_count=20,
        )

        self.assertEqual(payload["printed_on_date"], "2026-07-27")
        self.assertEqual(payload["conditions_of_class"], [])
        self.assertFalse(any(row["row_type"] == "condition" for row in payload["rows"]))

    def test_reconciliation_surfaces_conditions_in_dedicated_review_bucket(self):
        result = build_reconciliation_flags(
            parsed_payload={
                "rows": [
                    {
                        "class_society": "NK",
                        "class_code_or_name": "NK-CONDITION-OF-CLASS-1",
                        "source_section": "Condition of Class",
                        "row_type": "condition",
                        "raw_text": "Repair boiler automation system.",
                        "due_date": "2026-08-13",
                        "confidence": 1.0,
                    }
                ]
            },
            tracked_items=[],
            mappings=[],
        )

        self.assertEqual(result.flags[0]["bucket"], "conditions_of_class")
        self.assertEqual(result.counts["missing_in_catalog_count"], 0)

    def test_reconciliation_does_not_surface_memoranda_as_conditions(self):
        result = build_reconciliation_flags(
            parsed_payload={
                "rows": [
                    {
                        "class_society": "BV",
                        "class_code_or_name": "BV-CLASS-MEMORANDA-1",
                        "source_section": "Class Memoranda",
                        "row_type": "condition",
                        "kind": "memorandum",
                        "raw_text": "Memorandum text should not be treated as Condition of Class.",
                        "confidence": 1.0,
                    }
                ]
            },
            tracked_items=[],
            mappings=[],
        )

        self.assertEqual(result.flags, [])
        self.assertEqual(result.counts["missing_in_catalog_count"], 0)

    def test_reconciliation_does_not_surface_non_class_condition_sections(self):
        result = build_reconciliation_flags(
            parsed_payload={
                "rows": [
                    {
                        "class_society": "KR",
                        "class_code_or_name": "KR-ACTIONABLE-NOTE-1",
                        "source_section": "Actionable Note",
                        "row_type": "condition",
                        "kind": "actionable_note",
                        "raw_text": "Actionable note should not be treated as Condition of Class.",
                        "confidence": 1.0,
                    },
                    {
                        "class_society": "NK",
                        "class_code_or_name": "NK-CONDITION-OF-INSTALLATION-1",
                        "source_section": "Condition of Installation",
                        "row_type": "condition",
                        "raw_text": "Installation condition should not be treated as Condition of Class.",
                        "confidence": 1.0,
                    },
                    {
                        "class_society": "KR",
                        "class_code_or_name": "C.2",
                        "source_section": "Statutory Condition",
                        "row_type": "condition",
                        "raw_text": "Statutory condition should not be treated as Condition of Class.",
                        "confidence": 1.0,
                    },
                ]
            },
            tracked_items=[],
            mappings=[],
        )

        self.assertEqual(result.flags, [])
        self.assertEqual(result.counts["missing_in_catalog_count"], 0)

    def test_vessel_dashboard_snapshot_age_uses_printed_report_date(self):
        printed_on_date = datetime.now(timezone.utc).date().isoformat()

        serialized = _serialize_snapshot(
            {
                "snapshot_id": "90975FDE-C384-F111-ADEC-FDB8CC7078D1",
                "class_society": "NK",
                "uploaded_at": "2000-01-01T00:00:00+00:00",
                "printed_on_date": printed_on_date,
                "parse_status": "success",
                "reconciliation_run_id": None,
            }
        )

        self.assertEqual(serialized["printedOnDate"], printed_on_date)
        self.assertEqual(serialized["daysAgo"], 0)

    def test_vessel_dashboard_snapshot_age_never_uses_upload_date_when_report_date_missing(self):
        serialized = _serialize_snapshot(
            {
                "snapshot_id": "90975FDE-C384-F111-ADEC-FDB8CC7078D1",
                "class_society": "NK",
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                "printed_on_date": None,
                "parse_status": "failed",
                "reconciliation_run_id": None,
            }
        )

        self.assertIsNone(serialized["printedOnDate"])
        self.assertIsNone(serialized["daysAgo"])

    def test_class_snapshot_ocr_fallback_uses_paddleocr_page_cap(self):
        fake_engine = _FakeOcrEngine(
            OcrEngineOutput(
                raw_text="KOREAN REGISTER\nIMO No. 9584293",
                mean_confidence=0.93,
                fields={},
            )
        )

        with patch("apps.certs.services.parsers.base.PaddleOcrEngine", return_value=fake_engine) as engine_class:
            text = extract_pdf_image_ocr_text("image-only-class-status.pdf")

        engine_class.assert_called_once_with(**_class_snapshot_ocr_kwargs())
        self.assertIn("KOREAN REGISTER", text)

    def test_class_snapshot_ocr_fallback_reads_embedded_page_images(self):
        page_image = _png_bytes(width=1273, height=1800)
        logo_image = _png_bytes(width=583, height=73)
        fake_reader = _FakePdfReader([_FakePdfPage([_FakePdfImage(logo_image), _FakePdfImage(page_image)])])
        fake_engine = _FakeClassSnapshotOcrEngine("KOREAN REGISTER\nIMO No. 9584293")

        with (
            patch("PyPDF2.PdfReader", return_value=fake_reader),
            patch("apps.certs.services.parsers.base.PaddleOcrEngine", return_value=fake_engine) as engine_class,
        ):
            text = extract_pdf_embedded_image_ocr_text("image-only-class-status.pdf")

        engine_class.assert_called_once_with(**_class_snapshot_ocr_kwargs())
        self.assertEqual(fake_engine.calls, 1)
        self.assertIn("--- PAGE 1 OCR ---", text)
        self.assertIn("IMO No. 9584293", text)

    def test_class_snapshot_embedded_ocr_can_limit_pages(self):
        first_page = _png_bytes(width=1273, height=1800)
        second_page = _png_bytes(width=1273, height=1800)
        fake_reader = _FakePdfReader(
            [
                _FakePdfPage([_FakePdfImage(first_page)]),
                _FakePdfPage([_FakePdfImage(second_page)]),
            ]
        )
        fake_engine = _FakeClassSnapshotOcrEngine("Certificate row")

        with (
            patch("PyPDF2.PdfReader", return_value=fake_reader),
            patch("apps.certs.services.parsers.base.PaddleOcrEngine", return_value=fake_engine),
        ):
            text = extract_pdf_embedded_image_ocr_text("image-only-class-status.pdf", ocr_page_numbers=(2,))

        self.assertEqual(fake_engine.calls, 1)
        self.assertNotIn("--- PAGE 1 OCR ---", text)
        self.assertIn("--- PAGE 2 OCR ---", text)

    def test_process_cert_pdf_defaults_to_paddleocr_payload_contract(self):
        output = OcrEngineOutput(
            raw_text=(
                "Certificate Type: Certificate of Classification\n"
                "Vessel Name: EAST AYUTTHAYA\n"
                "IMO No. 9584293\n"
                "Issue Date: 01/07/2026\n"
                "Expiry Date: 01/07/2031\n"
                "Certificate No.: KR-001\n"
                "Place of Issue: Busan\n"
                "Issuing Authority: Korean Register\n"
            ),
            mean_confidence=0.92,
            fields={"certificate_number": OcrFieldCandidate("KR-001", 0.94)},
        )

        ocr_pipeline._DEFAULT_CERTIFICATE_OCR_ENGINE = None
        try:
            with patch("apps.certs.services.ocr_pipeline.PaddleOcrEngine", return_value=_FakeOcrEngine(output)) as engine_class:
                payload = process_cert_pdf(
                    "certificate.png",
                    thresholds=OcrThresholds(office_auto_accept=0.80, vessel_auto_accept=0.85, manual_floor=0.60),
                )
        finally:
            ocr_pipeline._DEFAULT_CERTIFICATE_OCR_ENGINE = None

        engine_class.assert_called_once_with(
            max_pdf_pages=ocr_pipeline.CERTIFICATE_OCR_MAX_PDF_PAGES,
            pdf_render_scale=ocr_pipeline.CERTIFICATE_OCR_PDF_RENDER_SCALE,
            text_detection_model_name=ocr_pipeline.CERTIFICATE_OCR_TEXT_DETECTION_MODEL_NAME,
            text_recognition_model_name=ocr_pipeline.CERTIFICATE_OCR_TEXT_RECOGNITION_MODEL_NAME,
            text_det_limit_side_len=ocr_pipeline.CERTIFICATE_OCR_TEXT_DET_LIMIT_SIDE_LEN,
            text_recognition_batch_size=ocr_pipeline.CERTIFICATE_OCR_TEXT_RECOGNITION_BATCH_SIZE,
        )
        self.assertEqual(payload["engine"], DEFAULT_OCR_ENGINE_NAME)
        self.assertEqual(payload["status"], "processed")
        self.assertEqual(payload["fields"]["certificate_number"]["value"], "KR-001")
        self.assertEqual(payload["fields"]["imo_number"]["value"], "9584293")

    def test_process_cert_pdf_reuses_default_paddleocr_engine(self):
        output = OcrEngineOutput(
            raw_text="Certificate No.: KR-001\nIMO No. 9584293",
            mean_confidence=0.92,
            fields={},
        )
        ocr_pipeline._DEFAULT_CERTIFICATE_OCR_ENGINE = None
        try:
            with patch("apps.certs.services.ocr_pipeline.PaddleOcrEngine", return_value=_FakeOcrEngine(output)) as engine_class:
                process_cert_pdf(
                    "certificate-1.png",
                    thresholds=OcrThresholds(office_auto_accept=0.80, vessel_auto_accept=0.85, manual_floor=0.60),
                )
                process_cert_pdf(
                    "certificate-2.png",
                    thresholds=OcrThresholds(office_auto_accept=0.80, vessel_auto_accept=0.85, manual_floor=0.60),
                )
        finally:
            ocr_pipeline._DEFAULT_CERTIFICATE_OCR_ENGINE = None

        engine_class.assert_called_once()

    def test_paddleocr_v3_prediction_result_is_flattened_to_text_and_confidence(self):
        text, confidence = _paddle_prediction_to_text(
            [
                {
                    "res": {
                        "rec_texts": ["KOREAN REGISTER", "IMO No. 9584293"],
                        "rec_scores": [0.96, 0.92],
                    }
                }
            ]
        )

        self.assertEqual(text, "KOREAN REGISTER\nIMO No. 9584293")
        self.assertAlmostEqual(confidence, 0.94)

    def test_paddleocr_prediction_reconstructs_table_rows_from_boxes(self):
        text, confidence = _paddle_prediction_to_text(
            [
                {
                    "res": {
                        "rec_texts": [
                            "Classification Certificate(Full)",
                            "CC",
                            "Full",
                            "2025-08-11",
                            "2030-07-11",
                            "Annual Survey",
                            "2026-07-11",
                        ],
                        "rec_scores": [0.96, 0.95, 0.97, 0.98, 0.99, 0.94, 0.93],
                        "rec_boxes": _TruthlessBoxes(
                            [
                                [Decimal("10"), Decimal("10"), Decimal("260"), Decimal("30")],
                                [Decimal("300"), Decimal("11"), Decimal("330"), Decimal("30")],
                                [Decimal("360"), Decimal("10"), Decimal("410"), Decimal("30")],
                                [Decimal("450"), Decimal("10"), Decimal("540"), Decimal("30")],
                                [Decimal("570"), Decimal("10"), Decimal("660"), Decimal("30")],
                                [Decimal("10"), Decimal("55"), Decimal("120"), Decimal("75")],
                                [Decimal("450"), Decimal("55"), Decimal("540"), Decimal("75")],
                            ]
                        ),
                    }
                }
            ]
        )

        self.assertIn("Classification Certificate(Full) CC Full 2025-08-11 2030-07-11", text)
        self.assertIn("Annual Survey 2026-07-11", text)
        self.assertAlmostEqual(confidence, 0.96, places=2)

    def test_class_parser_text_smoke_covers_supported_societies(self):
        samples = {
            "NK": (
                NKClassParser(),
                """NK-SHIPS Information Service
Name of Ship: TEST VESSEL Class No. : NK 123456 IMO No. : 1234567
Printed on 01.Jan.2026
Survey Status:: Class
Hull Annual Survey 01 Jan 2027
Condition & Note
""",
            ),
            "KR": (
                KRClassParser(),
                """VESSEL STATUS FOR SHIP'S OWNER TEST VESSEL Class No : 98765 IMO No : 1234567 Printed on 01-Jan-2026
Certificates
International Oil Pollution Prevention Certificate IOPP Full 2024-01-01 2029-01-01
Class Surveys
Ship Name Work ID
Hull Annual Survey 2025-01-01 2026-01-01
""",
            ),
            "BV": (
                BVClassParser(),
                """Ship name: TEST VESSEL BV Nr BV123 IMO Number: 1234567 Generated on 1 Jan 2026
Classification Surveys
Hull Annual Survey 1 Jan 2025 1 Jan 2026
Conditions of Class / Statutory Recommendations
None
""",
            ),
        }

        for society, (parser, text) in samples.items():
            with self.subTest(society=society):
                payload = parser.parse_text(text, page_count=1)
                self.assertEqual(payload["class_society"], society)
                self.assertGreaterEqual(len(payload["rows"]), 1)

    def test_kr_parser_accepts_ocr_table_rows(self):
        payload = KRClassParser().parse_text(
            """Ship Name EAST AYUTTHAYA Work ID VANS004726
Class No. 1000010 IMO No. 9584293
Certificates
Class Certificates
Cargo Ship Safety Construction Certificate Sc Full 2025-08-11 2030-07-11 CN25015925557
Loading Instrument Certificate LI Permanence 2025-08-11 CL25004894584
Vessel General Permit VGP 2025-08-11 CN25015890568
Class Surveys
Annual Survey 2026-07-11 2026-04-11-2026-10-11
Hull Annual Survey 2025-01-01 2026-01-01
""",
            page_count=19,
        )

        rows_by_code = {row["class_code_or_name"]: row for row in payload["rows"]}
        self.assertEqual(rows_by_code["SC"]["expiry_date"], "2030-07-11")
        self.assertNotIn("expiry_date", rows_by_code["LI"])
        self.assertEqual(rows_by_code["VGP"]["issue_date"], "2025-08-11")

        annual = next(row for row in payload["rows"] if row["raw_text"].startswith("Annual Survey"))
        self.assertNotIn("last_done_date", annual)
        self.assertEqual(annual["next_due_date"], "2026-07-11")
        hull = next(row for row in payload["rows"] if row["raw_text"].startswith("Hull Annual Survey"))
        self.assertEqual(hull["last_done_date"], "2025-01-01")
        self.assertEqual(hull["next_due_date"], "2026-01-01")

    def test_reconciliation_buckets_snapshot_rows_against_tracked_items(self):
        result = build_reconciliation_flags(
            parsed_payload={
                "rows": [
                    {"class_code_or_name": "IOPP", "certificate_number": "C-001", "expiry_date": "2029-01-01", "confidence": 1.0},
                    {"class_code_or_name": "LOADLINE", "certificate_number": "C-002", "expiry_date": "2028-01-01", "confidence": 1.0},
                    {"class_code_or_name": "UNKNOWN", "confidence": 1.0},
                    {"class_code_or_name": "LOWCONF", "confidence": 0.5},
                    {"class_code_or_name": "STCROW", "type": "STC", "confidence": 1.0},
                    {"class_code_or_name": "POSTROW", "status": "POSTPONED", "postponed_until": "2027-01-01", "confidence": 1.0},
                ]
            },
            tracked_items=[
                {"tracked_item_id": "ti-1", "catalog_id": "cat-1", "catalog_is_class_tracked": True, "certificate_number": "C-001", "expiry_date": "2029-01-01"},
                {"tracked_item_id": "ti-2", "catalog_id": "cat-2", "catalog_is_class_tracked": True, "certificate_number": "C-002", "expiry_date": "2029-01-01"},
                {"tracked_item_id": "ti-3", "catalog_id": "cat-3", "catalog_is_class_tracked": True},
                {"tracked_item_id": "ti-4", "catalog_id": "cat-4", "catalog_is_class_tracked": True},
                {"tracked_item_id": "ti-5", "catalog_id": "cat-5", "catalog_is_class_tracked": True},
            ],
            mappings=[
                {"class_code_or_name": "IOPP", "catalog_id": "cat-1", "version": 1},
                {"class_code_or_name": "LOADLINE", "catalog_id": "cat-2", "version": 1},
                {"class_code_or_name": "STCROW", "catalog_id": "cat-3", "version": 1},
                {"class_code_or_name": "POSTROW", "catalog_id": "cat-4", "version": 1},
            ],
        )

        self.assertEqual(
            result.counts,
            {
                "matches_count": 1,
                "mismatches_count": 1,
                "missing_in_catalog_count": 1,
                "missing_in_class_count": 1,
                "conditional_stc_detected_count": 1,
                "extended_postponed_detected_count": 1,
                "unmapped_low_confidence_count": 1,
            },
        )

    def test_kr_class_code_mapping_seed_covers_ayuthya_parser_rows(self):
        expected_codes = {
            "Air Pollution Prevention Annual Survey",
            "Air Pollution Prevention Intermediate Survey",
            "Air Pollution Prevention Renewal Survey",
            "Annual Survey",
            "BWM",
            "Ballast Water Management Annual Survey",
            "Ballast Water Management Intermediate Survey",
            "Ballast Water Management Renewal Survey",
            "CC",
            "CDG",
            "CG2",
            "Cargo Gear Survey(Annual)",
            "Cargo Ship Safety Construction Annual Survey",
            "Cargo Ship Safety Construction Intermediate Survey",
            "Cargo Ship Safety Construction Renewal Survey",
            "Cargo Ship Safety Equipment Annual Survey",
            "Cargo Ship Safety Equipment Periodical Survey",
            "Cargo Ship Safety Equipment Renewal Survey",
            "Cargo Ship Safety Radio Periodical Survey",
            "Cargo Ship Safety Radio Renewal Survey",
            "Docking Survey",
            "Garbage Pollution Prevention Renewal Survey",
            "IAFS",
            "IAPP",
            "IEE",
            "IGPP",
            "IHM(EU)",
            "IIHM",
            "ILL",
            "IMSBC",
            "IOPP-A",
            "ISPP",
            "Intermediate Survey",
            "Inventory of Hazardous Materials Occasional Survey",
            "Inventory of Hazardous Materials Renewal Survey",
            "LI",
            "Maritime Solid Bulk Cargoes Code Renewal Survey",
            "No.1 Aux.Boiler Survey",
            "No.1 Propeller Shaft Survey",
            "Oil Pollution Prevention Annual Survey",
            "Oil Pollution Prevention Intermediate Survey",
            "Oil Pollution Prevention Renewal Survey",
            "Renewal Survey",
            "SC",
            "SE",
            "SR",
            "Sewage Pollution Prevention Renewal Survey",
            "Special Survey",
            "VGP",
        }

        validate_class_code_mapping_seed_rows(KR_CLASS_CODE_MAPPING_ROWS)
        rows_by_code = {row.class_code_or_name: row for row in KR_CLASS_CODE_MAPPING_ROWS}

        self.assertEqual(set(rows_by_code), expected_codes)
        self.assertEqual(rows_by_code["CC"].catalog_code, "CLASS-COC")
        self.assertEqual(rows_by_code["IOPP-A"].catalog_code, "STAT-INTERNATIONAL-OIL-POLLUTION-PREVENTION-IOPP-WITH")
        self.assertEqual(rows_by_code["Cargo Ship Safety Radio Periodical Survey"].cert_or_survey_kind, "periodic")

    def test_kr_seeded_mappings_turn_known_rows_into_matches(self):
        seed_by_code = {row.class_code_or_name: row for row in KR_CLASS_CODE_MAPPING_ROWS}
        parsed_rows = [
            {"class_code_or_name": "CC", "certificate_number": "KR-CC", "expiry_date": "2030-07-11", "confidence": 1.0},
            {"class_code_or_name": "IOPP-A", "certificate_number": "KR-IOPP", "expiry_date": "2030-07-11", "confidence": 1.0},
            {"class_code_or_name": "Air Pollution Prevention Annual Survey", "next_due_date": "2026-07-11", "confidence": 1.0},
            {"class_code_or_name": "No.1 Propeller Shaft Survey", "next_due_date": "2026-07-11", "confidence": 1.0},
        ]
        mappings = [
            {"class_code_or_name": row["class_code_or_name"], "catalog_id": seed_by_code[row["class_code_or_name"]].catalog_code, "version": 1}
            for row in parsed_rows
        ]
        tracked_items = [
            {
                "tracked_item_id": "ti-class",
                "catalog_id": "CLASS-COC",
                "catalog_is_class_tracked": True,
                "certificate_number": "KR-CC",
                "expiry_date": "2030-07-11",
            },
            {
                "tracked_item_id": "ti-iopp",
                "catalog_id": "STAT-INTERNATIONAL-OIL-POLLUTION-PREVENTION-IOPP-WITH",
                "catalog_is_class_tracked": True,
                "certificate_number": "KR-IOPP",
                "expiry_date": "2030-07-11",
            },
            {
                "tracked_item_id": "ti-iapp-survey",
                "catalog_id": "TRADE-IAPP-ANNUAL-PERIODICAL",
                "catalog_is_class_tracked": True,
                "next_due_date": "2026-07-11",
            },
            {
                "tracked_item_id": "ti-prop-shaft",
                "catalog_id": "CLASS-PROP-SHAFT-SURVEY",
                "catalog_is_class_tracked": True,
                "next_due_date": "2026-07-11",
            },
        ]

        result = build_reconciliation_flags(parsed_payload={"rows": parsed_rows}, tracked_items=tracked_items, mappings=mappings)

        self.assertEqual(result.counts["matches_count"], 4)
        self.assertEqual(result.counts["missing_in_catalog_count"], 0)
        self.assertEqual(result.counts["mismatches_count"], 0)

    def test_parser_anomaly_notification_schema_failure_does_not_abort_reconciliation(self):
        result = dispatch_parser_anomaly_notifications(
            run={"run_id": "run-1", "vessel_id": "vessel-1", "vessel_name": "MV Test", "class_society": "KR"},
            anomaly_breaches=[{"type": "parse_duration", "severity": "critical"}],
            flags=[],
            dispatcher=_FailingNotificationDispatcher(),
            candidate_recipients=[
                CertNotificationRecipient(user_id="dpa-1", role="DPA", side="office"),
                CertNotificationRecipient(user_id="tech-1", role="Technical Superintendent", side="office"),
            ],
        )

        self.assertFalse(result["dispatched"])
        self.assertEqual(result["reason"], "notification_dispatch_failed")
        self.assertEqual(result["notificationsSent"], [])


class CatalogRowSubmissionScopeTests(SimpleTestCase):
    def test_catalog_row_create_rejects_master_only_under_all_rank_policy(self):
        serializer = CatalogRowWriteSerializer(
            data=_catalog_row_payload("master_only"),
            context={"is_create": True},
        )

        self.assertFalse(serializer.is_valid())
        self.assertIn("submissionScope", serializer.errors)

    def test_catalog_row_create_accepts_all_rank_policy(self):
        serializer = CatalogRowWriteSerializer(
            data=_catalog_row_payload("all_ranks_with_approval"),
            context={"is_create": True},
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)


def _catalog_row_payload(submission_scope: str) -> dict[str, object]:
    return {
        "canonicalCode": "TEST-CATALOG-ROW",
        "sectionId": 2,
        "displayName": "Test Catalog Row",
        "printSectionLabel": "Statutory & Flag",
        "validityType": "full",
        "issuingAuthorityType": "flag",
        "submissionScope": submission_scope,
    }


def _class_snapshot_row(
    *,
    snapshot_id: str,
    vessel_id: str,
    blob_id: str,
    parse_status: str,
    reconciliation_run_id: str | None,
    parser_version: str = "pending-parser-v1",
) -> dict[str, object]:
    return {
        "snapshot_id": snapshot_id,
        "vessel_id": vessel_id,
        "vessel_name": "MV Test",
        "imo_number": "1234567",
        "class_society": "NK",
        "pdf_blob_id": blob_id,
        "filename": "class-status.pdf",
        "content_size_bytes": 10,
        "printed_on_date": "2026-07-20",
        "uploaded_by": "actor-1",
        "uploaded_at": "2026-07-20T10:00:00Z",
        "parser_version": parser_version,
        "parse_status": parse_status,
        "parse_started_at": None,
        "parse_completed_at": "2026-07-20T10:00:05Z" if parse_status != "pending" else None,
        "parser_timeout": False,
        "retry_count": 0,
        "parsed_payload_json": None,
        "parsed_payload_schema_version": 1,
        "reconciliation_run_id": reconciliation_run_id,
        "upload_sha256": "sha",
        "superseded_user_error": False,
    }


def _vessel_master_user(
    *,
    vessel_id: str,
    role: str = "VESSEL_MASTER",
    rank: str = "MASTER",
) -> SimpleNamespace:
    return SimpleNamespace(
        is_authenticated=True,
        user_type="VESSEL",
        role=role,
        role_name=role,
        rank=rank,
        crew_id="KSM0229",
        vessel_id=vessel_id,
        form_ids=["CERT_F_003"],
        process_ids=[],
    )


def _master_message_row(
    *,
    flag_id: str = "55555555-5555-5555-5555-555555555555",
    vessel_id: str,
    master_reviewed_at: str | None,
    master_reviewed_by: str | None = None,
    master_review_note: str | None = None,
) -> dict[str, object]:
    return {
        "flag_id": flag_id,
        "run_id": "33333333-3333-3333-3333-333333333333",
        "snapshot_id": "44444444-4444-4444-4444-444444444444",
        "vessel_id": vessel_id,
        "vessel_name": "MV Test",
        "imo_number": "1234567",
        "class_society": "KR",
        "printed_on_date": "2026-07-20",
        "ran_at": "2026-07-22T09:00:00Z",
        "bucket": "mismatch",
        "catalog_id": "cat-1",
        "catalog_display_name": "Load Line Certificate",
        "tracked_item_id": "tracked-1",
        "class_row_extract_json": {
            "display_name": "Load Line Certificate",
            "class_code_or_name": "ILL",
            "expiry_date": "2030-07-11",
        },
        "diff_json": {"expiry_date": {"tracked": "2029-07-11", "class": "2030-07-11"}},
        "reviewed_by": "office-1",
        "reviewed_at": "2026-07-22T09:05:00Z",
        "resolution_action": "notified_master",
        "resolved_at": "2026-07-22T09:05:00Z",
        "office_notified_at": "2026-07-22T09:05:00Z",
        "office_notified_by": "office-1",
        "office_notified_role": "DPA",
        "office_note": "Please check expiry date with the vessel file.",
        "master_reviewed_at": master_reviewed_at,
        "master_reviewed_by": master_reviewed_by,
        "master_reviewed_role": "VESSEL_MASTER" if master_reviewed_by else None,
        "master_review_note": master_review_note,
    }


def _print_artifact_row(
    *,
    print_id: str = "SQE-S633-TEST-001",
    scope: str = "per_vessel_full",
    vessels: list[str] | None = None,
    sections: list[str] | None = None,
    custom_cert_ids: list[str] | None = None,
    pdf_blob_id: str | None = "pdf-1",
    excel_blob_id: str | None = "excel-1",
    bundle_zip_blob_id: str | None = None,
    recipient_email: str = "",
) -> dict[str, object]:
    return {
        "print_id": print_id,
        "scope": scope,
        "vessels_json": json.dumps(vessels or ["11111111-1111-1111-1111-111111111111"]),
        "sections_json": json.dumps(sections or []),
        "filters_json": "{}",
        "custom_cert_ids_json": json.dumps(custom_cert_ids or []),
        "user_id": "Harman.S",
        "user_role": "DPA",
        "timestamp_utc": "2026-07-23T08:00:00Z",
        "system_state_hash": "ABC12345",
        "watermark_applied": "NONE",
        "watermark_recipient": "",
        "pdf_blob_id": pdf_blob_id,
        "excel_blob_id": excel_blob_id,
        "bundle_zip_blob_id": bundle_zip_blob_id,
        "recipient_email": recipient_email,
        "page_count": 2,
        "generation_status": "success",
        "failure_message": "",
    }


def _tracked_item_test_row(
    *,
    status: str = "ok",
    certificate_number: str | None = "CERT-OLD",
    expiry_date: str | None = "2029-07-24",
    version: int = 1,
) -> dict[str, object]:
    return {
        "tracked_item_id": "tracked-1",
        "vessel_id": "vessel-1",
        "vessel_name": "MV Test",
        "vessel_code": "TST",
        "vessel_imo_number": "1234567",
        "catalog_id": "catalog-1",
        "catalog_code": "SMC",
        "catalog_display_name": "Safety Management Certificate",
        "catalog_short_name": "SMC",
        "catalog_section_code": "STATUTORY",
        "catalog_section_name": "Statutory",
        "catalog_is_class_tracked": False,
        "catalog_retain_all_versions": False,
        "catalog_submission_scope": "all_ranks_with_approval",
        "type": "certificate",
        "validity_type": "full",
        "form_variant": None,
        "cadence_months": 60,
        "cadence_custom_days": None,
        "parent_id": None,
        "relationship_type": None,
        "supersedes_id": None,
        "issue_date": "2026-07-24",
        "expiry_date": expiry_date,
        "anniversary_date": None,
        "window_open": None,
        "window_close": None,
        "last_done_date": None,
        "next_due_date": None,
        "postponed_until": None,
        "status": status,
        "certificate_number": certificate_number,
        "issuing_authority": "Class Society",
        "place_of_issue": None,
        "extension_authority": None,
        "extension_letter_pdf_id": None,
        "extension_reason": None,
        "pdf_attachment_id": "blob-1",
        "pdf_missing": False,
        "source": "manual",
        "last_class_sync_id": None,
        "approval_state": "approved",
        "submitted_by": "Harman.S",
        "submitted_at": "2026-07-24T08:00:00Z",
        "approved_by": "Harman.S",
        "approved_at": "2026-07-24T08:00:00Z",
        "rejection_reason": None,
        "rejection_count": 0,
        "draft_expires_at": None,
        "lifecycle_status": "active",
        "row_version": b"\x00\x00\x00\x00\x00\x00\x00\x01",
        "version": version,
        "created_at": "2026-07-24T08:00:00Z",
        "created_by": "Harman.S",
        "updated_at": "2026-07-24T08:00:00Z",
        "updated_by": "Harman.S",
    }


def _pdf_blob_test_row(
    *,
    blob_storage_path: str,
    ocr_payload_json: dict[str, object],
) -> dict[str, object]:
    return {
        "blob_id": "blob-1",
        "tracked_item_id": "tracked-1",
        "snapshot_id": None,
        "blob_storage_path": blob_storage_path,
        "filename": "certificate.pdf",
        "content_sha256": "sha",
        "content_size_bytes": 8,
        "uploaded_by": "Harman.S",
        "uploaded_at": "2026-07-24T08:00:00Z",
        "is_active": True,
        "superseded_at": None,
        "retention_policy": "retain_18_months_then_purge",
        "scheduled_delete_at": None,
        "delete_pending_since": None,
        "dpa_retention_override_until": None,
        "ocr_payload_json": ocr_payload_json,
        "ocr_confidence_per_field": {},
        "ocr_processed_at": "2026-07-24T08:01:00Z" if ocr_payload_json else None,
        "ocr_engine_version": DEFAULT_OCR_ENGINE_NAME if ocr_payload_json else "",
    }


def _write_upload_blob(upload_root: str, relative_path: str, content: bytes) -> None:
    absolute_path = Path(upload_root) / relative_path
    absolute_path.parent.mkdir(parents=True, exist_ok=True)
    absolute_path.write_bytes(content)


class _FakePdfBlobRepository:
    def __init__(self, blobs: dict[str, dict[str, object]]):
        self.blobs = blobs

    def get_blob(self, blob_id: str):
        return self.blobs.get(blob_id)


class _FakeEmailConnectionFactory:
    def __init__(self):
        self.connections = []

    def __call__(self, **kwargs):
        connection = _FakeEmailConnection(kwargs)
        self.connections.append(connection)
        return connection


class _FakeEmailConnection:
    def __init__(self, kwargs):
        self.kwargs = kwargs
        self.sent_messages = []

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def send_messages(self, messages):
        self.sent_messages.extend(messages)
        return len(messages)


class _NoopClassParser(BaseClassParser):
    def parse_text(self, text: str, *, page_count: int) -> dict[str, object]:
        return {"rows": [], "schema_version": 1}


class _EmptyPdfPlumberDocument:
    def __init__(self, *, page_count: int):
        self.pages = [_EmptyPdfPlumberPage() for _ in range(page_count)]

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False


class _EmptyPdfPlumberPage:
    def extract_text(self, **_kwargs):
        return ""


class _FailingNotificationDispatcher:
    def dispatch(self, **_kwargs):
        raise DatabaseError("Invalid column name 'module_code'.")


class _FakeOcrEngine:
    engine_name = DEFAULT_OCR_ENGINE_NAME

    def __init__(self, output: OcrEngineOutput):
        self.output = output

    def extract(self, _source_path):
        return self.output


class _FakeClassSnapshotOcrEngine:
    def __init__(self, text: str):
        self.text = text
        self.calls = 0

    def extract_image_text(self, _source_path):
        self.calls += 1
        return self.text, 0.95


class _FakePdfReader:
    def __init__(self, pages):
        self.pages = pages


class _FakePdfPage:
    def __init__(self, images):
        self._images = images

    def get(self, key):
        if key == "/Resources":
            return {"/XObject": _FakePdfXObjects(self._images)}
        return None


class _FakePdfXObjects:
    def __init__(self, images):
        self._images = images

    def get_object(self):
        return {f"/X{index}": image for index, image in enumerate(self._images, start=1)}


class _FakePdfImage:
    def __init__(self, content: bytes):
        self._content = content

    def get_object(self):
        return self

    def get(self, key):
        if key == "/Subtype":
            return "/Image"
        return None

    def get_data(self):
        return self._content


class _TruthlessBoxes(list):
    def __bool__(self):
        raise ValueError("array truth value is ambiguous")


def _png_bytes(*, width: int, height: int) -> bytes:
    from PIL import Image

    buffer = BytesIO()
    Image.new("RGB", (width, height), "white").save(buffer, format="PNG")
    return buffer.getvalue()


def _class_snapshot_ocr_kwargs() -> dict[str, object]:
    return {
        "language": CLASS_SNAPSHOT_OCR_LANGUAGE,
        "max_pdf_pages": CLASS_SNAPSHOT_OCR_MAX_PAGES,
        "ocr_version": CLASS_SNAPSHOT_OCR_VERSION,
        "text_det_limit_side_len": CLASS_SNAPSHOT_OCR_DET_LIMIT_SIDE_LEN,
        "text_recognition_batch_size": CLASS_SNAPSHOT_OCR_REC_BATCH_SIZE,
    }


class _FakeCatalogCursor:
    def __init__(self):
        self.executed = []
        self.description = [("catalog_id",)]

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def execute(self, sql, params=None):
        self.executed.append((sql, list(params or [])))

    def fetchone(self):
        return (123,)

    def fetchall(self):
        return [("catalog-1",)]


class _FakeAuditLogCursor:
    def __init__(self):
        self.executed = []

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def execute(self, sql, params=None):
        self.executed.append((sql, list(params or [])))


class _FakeTrackedItemCursor:
    def __init__(self, *, count=0):
        self.executed = []
        self.description = [("tracked_item_id",)]
        self.count = count

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def execute(self, sql, params=None):
        self.executed.append((sql, list(params or [])))

    def fetchone(self):
        return (self.count,)

    def fetchall(self):
        return []


class _FakeRepositoryListCursor:
    def __init__(self, *, count=0):
        self.executed = []
        self.description = [("id",)]
        self.count = count

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def execute(self, sql, params=None):
        self.executed.append((sql, list(params or [])))

    def fetchone(self):
        return (self.count,)

    def fetchall(self):
        return []


def _tracked_item_serializer_row() -> dict[str, object]:
    return {
        "tracked_item_id": "11111111-1111-1111-1111-111111111111",
        "vessel_id": "22222222-2222-2222-2222-222222222222",
        "catalog_id": "33333333-3333-3333-3333-333333333333",
        "catalog_code": "CERT-001",
        "catalog_display_name": "Safety Management Certificate",
        "catalog_short_name": "SMC",
        "catalog_submission_scope": "all_ranks_with_approval",
        "type": "certificate",
        "validity_type": "full",
        "status": "pending_first_upload",
        "issuing_authority": "Class",
        "pdf_missing": True,
        "source": "manual",
        "approval_state": "approved",
        "lifecycle_status": "active",
        "created_by": "seed-user",
        "updated_by": "seed-user",
        "version": 1,
    }


class TrackedItemMetadataSerializerTests(SimpleTestCase):
    def test_partial_metadata_correction_payload_is_valid(self):
        serializer = TrackedItemWriteSerializer(
            data={
                "certificateNumber": "CERT-2026-001",
                "issuingAuthority": "ClassNK",
                "placeOfIssue": "Singapore",
                "issueDate": "2026-07-01",
                "expiryDate": "2027-07-01",
                "reason": "Metadata corrected after OCR review.",
            },
            partial=True,
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
