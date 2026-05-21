from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import os
from pathlib import Path

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from apps.safety.models.dashboard_rollup import SafetyDashboardRollup

from .composite_score import CompositeScoreService
from .dashboard_ca_aging import DashboardCorrectiveActionAgingService
from .dashboard_soi_compliance import DashboardSOIComplianceService
from .heinrich_ratio import HeinrichRatioService
from .pareto_screener import ParetoScreenerService
from .repeat_root_radar import RepeatRootRadarService


@dataclass(frozen=True)
class DashboardExportResult:
    content: bytes
    content_type: str
    export_path: str | None
    file_name: str
    format: str


class DashboardExportService:
    PDF = "pdf"
    EXCEL = "excel"
    XLSX = "xlsx"
    PDF_CONTENT_TYPE = "application/pdf"
    EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __init__(
        self,
        *,
        composite_service: CompositeScoreService | None = None,
        heinrich_service: HeinrichRatioService | None = None,
        repeat_root_service: RepeatRootRadarService | None = None,
        pareto_service: ParetoScreenerService | None = None,
        soi_compliance_service: DashboardSOIComplianceService | None = None,
        ca_aging_service: DashboardCorrectiveActionAgingService | None = None,
        export_root: str | os.PathLike[str] | None = None,
        now_func=timezone.now,
    ) -> None:
        self.composite_service = composite_service or CompositeScoreService(now_func=now_func)
        self.heinrich_service = heinrich_service or HeinrichRatioService(now_func=now_func)
        self.repeat_root_service = repeat_root_service or RepeatRootRadarService(now_func=now_func)
        self.pareto_service = pareto_service or ParetoScreenerService(now_func=now_func)
        self.soi_compliance_service = soi_compliance_service or DashboardSOIComplianceService()
        self.ca_aging_service = ca_aging_service or DashboardCorrectiveActionAgingService()
        self.now_func = now_func
        default_root = Path.cwd() / "var" / "www" / "ksm_uploads" / "safety"
        self.export_root = Path(export_root or os.getenv("SAFETY_EXPORT_ROOT") or default_root)

    def build_export(
        self,
        *,
        export_format: str,
        period_code: str | None,
        vessel_id: str | None,
        viewer_user,
        persist: bool = True,
    ) -> DashboardExportResult:
        normalized_format = self._normalize_format(export_format)
        scope = self.composite_service.resolve_scope(vessel_id=vessel_id, user=viewer_user)
        scoped_vessel_id = scope.scope_id if scope.scope_type == SafetyDashboardRollup.ScopeType.VESSEL else None
        exported_at = self.now_func()
        normalized_period = self.composite_service._normalize_period_code(  # noqa: SLF001
            period_code or SafetyDashboardRollup.PeriodCode.YEARS_3
        )

        payload = {
            "metadata": {
                "exported_at": exported_at.isoformat(),
                "exporter_name": self._exporter_name(viewer_user),
                "period_code": normalized_period,
                "scope_id": scope.scope_id,
                "scope_type": scope.scope_type,
            },
            "ca_aging": self.ca_aging_service.build_panel(vessel_id=scoped_vessel_id),
            "composite": self.composite_service.build_rollup(
                scope=scope,
                period_code=normalized_period,
                as_of=exported_at,
            ),
            "heinrich": self.heinrich_service.build_panel(vessel_id=scoped_vessel_id, as_of=exported_at),
            "pareto": self.pareto_service.build_panel(vessel_id=scoped_vessel_id, as_of=exported_at),
            "repeat_root": self.repeat_root_service.build_panel(vessel_id=scoped_vessel_id, as_of=exported_at),
            "soi_compliance": self.soi_compliance_service.build_panel(vessel_id=scoped_vessel_id),
        }

        if normalized_format == self.PDF:
            content = self._render_pdf(payload)
            content_type = self.PDF_CONTENT_TYPE
            suffix = "pdf"
        else:
            content = self._render_excel(payload)
            content_type = self.EXCEL_CONTENT_TYPE
            suffix = "xlsx"

        file_name = self._build_file_name(
            period_code=normalized_period,
            scope_type=scope.scope_type,
            scope_id=scope.scope_id,
            suffix=suffix,
        )
        export_path = self._persist_export(content=content, file_name=file_name, scope_type=scope.scope_type, scope_id=scope.scope_id) if persist else None
        return DashboardExportResult(
            content=content,
            content_type=content_type,
            export_path=export_path,
            file_name=file_name,
            format=normalized_format,
        )

    def _normalize_format(self, export_format: str) -> str:
        normalized = str(export_format or "").strip().lower()
        if normalized == self.XLSX:
            return self.EXCEL
        if normalized not in {self.PDF, self.EXCEL}:
            raise ValueError("export_format must be 'pdf', 'excel', or 'xlsx'.")
        return normalized

    def _render_pdf(self, payload: dict[str, object]) -> bytes:
        metadata = payload["metadata"]
        composite = payload["composite"]
        heinrich = payload["heinrich"]
        repeat_root = payload["repeat_root"]
        pareto = payload["pareto"]
        soi = payload["soi_compliance"]
        ca_aging = payload["ca_aging"]

        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4, invariant=1)
        width, height = A4
        y = height - 48

        def write_line(text: str, *, size: int = 10, bold: bool = False, gap: int = 14) -> None:
            nonlocal y
            if y < 60:
                pdf.showPage()
                y = height - 48
            pdf.setFont("Helvetica-Bold" if bold else "Helvetica", size)
            pdf.drawString(40, y, text[:120])
            y -= gap

        write_line("Safety Intelligence Dashboard Export", size=16, bold=True, gap=20)
        write_line(f"Generated At: {metadata['exported_at']}", bold=True)
        write_line(f"Exporter: {metadata['exporter_name']}")
        write_line(f"Period: {metadata['period_code']}")
        write_line(f"Scope: {metadata['scope_type']} {metadata['scope_id']}".strip())
        y -= 4

        write_line("Composite Score", size=13, bold=True, gap=18)
        write_line(f"Score: {composite['composite_score']} ({composite['score_status']})")
        metrics = composite["metrics"]
        write_line(f"Open incidents: {metrics['open_incidents']}")
        write_line(f"Open near misses: {metrics['open_near_misses']}")
        write_line(f"Open findings: {metrics['open_findings']}")
        write_line(f"Overdue corrective actions: {metrics['overdue_corrective_actions']}")
        write_line(f"SOI Compliance %: {metrics['soi_compliance_display']}")
        y -= 4

        write_line("CA Aging Pipeline", size=13, bold=True, gap=18)
        write_line(f"Open actions: {ca_aging['open_action_count']}")
        write_line(f"Oldest action age: {ca_aging['oldest_age_days']} days")
        for bucket in ca_aging["buckets"]:
            write_line(f"{bucket['label']}: {bucket['count']}")
        y -= 4

        write_line("Heinrich Ratio", size=13, bold=True, gap=18)
        write_line(f"Confidence: {heinrich['confidence']['status']} - {heinrich['confidence']['reason']}")
        for layer in heinrich["layers"][:5]:
            write_line(f"{layer['label']}: actual {layer['actual']} vs benchmark {layer['benchmark']}")
        y -= 4

        write_line("Repeat Root-Cause Radar", size=13, bold=True, gap=18)
        repeat_items = repeat_root["fleet"] if repeat_root["fleet"] else repeat_root["vessel"]
        if repeat_items:
            for item in repeat_items[:5]:
                write_line(f"{item['subcode_id']}: {item['description']} ({item['occurrences']})")
        else:
            write_line("No repeat root causes met the threshold.")
        y -= 4

        write_line("Pareto Screening", size=13, bold=True, gap=18)
        if pareto["entries"]:
            for entry in pareto["entries"][:5]:
                write_line(
                    f"#{entry['rank']} {entry['subcode_id']} {entry['description']} - {entry['share_percent']}%"
                )
        else:
            write_line("No Pareto rows are available for the current window.")
        y -= 4

        write_line("SOI Compliance", size=13, bold=True, gap=18)
        write_line(f"Current vessel: {soi['current_vessel']['display_value']}")
        write_line(f"Fleet average: {soi['fleet_average']['display_value']}")

        pdf.showPage()
        pdf.save()
        return buffer.getvalue()

    def _render_excel(self, payload: dict[str, object]) -> bytes:
        workbook = Workbook()
        metadata_sheet = workbook.active
        metadata_sheet.title = "Metadata"
        metadata_sheet["A1"] = "Safety Intelligence Dashboard Export"
        metadata_sheet["A1"].font = Font(bold=True, size=14)
        metadata_sheet["A2"] = "Generated At"
        metadata_sheet["B2"] = payload["metadata"]["exported_at"]
        metadata_sheet["A3"] = "Exporter"
        metadata_sheet["B3"] = payload["metadata"]["exporter_name"]
        metadata_sheet["A4"] = "Period"
        metadata_sheet["B4"] = payload["metadata"]["period_code"]
        metadata_sheet["A5"] = "Scope Type"
        metadata_sheet["B5"] = payload["metadata"]["scope_type"]
        metadata_sheet["A6"] = "Scope ID"
        metadata_sheet["B6"] = payload["metadata"]["scope_id"]

        summary_sheet = workbook.create_sheet("Summary")
        summary_sheet["A1"] = "Composite Score"
        summary_sheet["B1"] = payload["composite"]["composite_score"]
        summary_sheet["A2"] = "Score Status"
        summary_sheet["B2"] = payload["composite"]["score_status"]
        summary_sheet["A4"] = "Metric"
        summary_sheet["B4"] = "Value"
        summary_sheet["A4"].font = summary_sheet["B4"].font = Font(bold=True)
        row = 5
        for key, value in payload["composite"]["metrics"].items():
            summary_sheet[f"A{row}"] = key
            summary_sheet[f"B{row}"] = value
            row += 1

        ca_sheet = workbook.create_sheet("CA Aging")
        ca_sheet["A1"] = payload["ca_aging"]["label"]
        ca_sheet["A1"].font = Font(bold=True, size=14)
        ca_sheet["A2"] = "Open Actions"
        ca_sheet["B2"] = payload["ca_aging"]["open_action_count"]
        ca_sheet["A3"] = "Oldest Age (days)"
        ca_sheet["B3"] = payload["ca_aging"]["oldest_age_days"]
        ca_sheet["A5"] = "Bucket"
        ca_sheet["B5"] = "Count"
        ca_sheet["A5"].font = ca_sheet["B5"].font = Font(bold=True)
        for index, bucket in enumerate(payload["ca_aging"]["buckets"], start=6):
            ca_sheet[f"A{index}"] = bucket["label"]
            ca_sheet[f"B{index}"] = bucket["count"]

        heinrich_sheet = workbook.create_sheet("Heinrich")
        heinrich_sheet["A1"] = "Layer"
        heinrich_sheet["B1"] = "Actual"
        heinrich_sheet["C1"] = "Benchmark"
        heinrich_sheet["D1"] = "Variance"
        for cell in ("A1", "B1", "C1", "D1"):
            heinrich_sheet[cell].font = Font(bold=True)
        for index, layer in enumerate(payload["heinrich"]["layers"], start=2):
            heinrich_sheet[f"A{index}"] = layer["label"]
            heinrich_sheet[f"B{index}"] = layer["actual"]
            heinrich_sheet[f"C{index}"] = layer["benchmark"]
            heinrich_sheet[f"D{index}"] = layer["variance"]

        repeat_sheet = workbook.create_sheet("Repeat Root")
        repeat_sheet["A1"] = "Subcode"
        repeat_sheet["B1"] = "Description"
        repeat_sheet["C1"] = "Occurrences"
        for cell in ("A1", "B1", "C1"):
            repeat_sheet[cell].font = Font(bold=True)
        repeat_items = payload["repeat_root"]["fleet"] if payload["repeat_root"]["fleet"] else payload["repeat_root"]["vessel"]
        for index, item in enumerate(repeat_items, start=2):
            repeat_sheet[f"A{index}"] = item["subcode_id"]
            repeat_sheet[f"B{index}"] = item["description"]
            repeat_sheet[f"C{index}"] = item["occurrences"]

        pareto_sheet = workbook.create_sheet("Pareto")
        pareto_sheet["A1"] = "Rank"
        pareto_sheet["B1"] = "Subcode"
        pareto_sheet["C1"] = "Description"
        pareto_sheet["D1"] = "Occurrences"
        pareto_sheet["E1"] = "Share %"
        for cell in ("A1", "B1", "C1", "D1", "E1"):
            pareto_sheet[cell].font = Font(bold=True)
        for index, entry in enumerate(payload["pareto"]["entries"], start=2):
            pareto_sheet[f"A{index}"] = entry["rank"]
            pareto_sheet[f"B{index}"] = entry["subcode_id"]
            pareto_sheet[f"C{index}"] = entry["description"]
            pareto_sheet[f"D{index}"] = entry["occurrences"]
            pareto_sheet[f"E{index}"] = entry["share_percent"]

        soi_sheet = workbook.create_sheet("SOI")
        soi_sheet["A1"] = payload["soi_compliance"]["label"]
        soi_sheet["A1"].font = Font(bold=True, size=14)
        soi_sheet["A2"] = "Current Vessel"
        soi_sheet["B2"] = payload["soi_compliance"]["current_vessel"]["display_value"]
        soi_sheet["A3"] = "Fleet Average"
        soi_sheet["B3"] = payload["soi_compliance"]["fleet_average"]["display_value"]

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _exporter_name(user) -> str:
        return str(
            getattr(user, "username", None)
            or getattr(user, "id", None)
            or getattr(user, "pk", None)
            or "unknown"
        )

    def _persist_export(self, *, content: bytes, file_name: str, scope_type: str, scope_id: str) -> str:
        scope_folder = scope_id if scope_type == SafetyDashboardRollup.ScopeType.VESSEL and scope_id else "fleet"
        export_dir = self.export_root / scope_folder / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        output_path = export_dir / file_name
        output_path.write_bytes(content)
        return str(output_path.resolve())

    def _build_file_name(self, *, period_code: str, scope_type: str, scope_id: str, suffix: str) -> str:
        timestamp = self.now_func().strftime("%Y%m%d-%H%M%S")
        scope_folder = scope_id if scope_type == SafetyDashboardRollup.ScopeType.VESSEL and scope_id else "fleet"
        return f"safety-dashboard-{scope_folder}-{period_code.lower()}-{timestamp}.{suffix}"
