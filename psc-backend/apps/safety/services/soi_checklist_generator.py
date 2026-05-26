from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
import qrcode

from apps.safety.repositories import SOIRepository

from .checklist_version_resolver import ChecklistVersionResolutionError, ChecklistVersionResolver


@dataclass(frozen=True)
class SOIChecklistRenderResult:
    content: bytes
    content_type: str
    file_name: str
    output_format: str


class SOIChecklistGenerator:
    def __init__(
        self,
        *,
        soi_repository: SOIRepository | None = None,
        checklist_version_resolver: ChecklistVersionResolver | None = None,
    ) -> None:
        self.soi_repository = soi_repository or SOIRepository()
        self.checklist_version_resolver = checklist_version_resolver or ChecklistVersionResolver()

    def render_for_inspection(
        self,
        *,
        inspection_id,
        output_format: str,
    ) -> SOIChecklistRenderResult:
        inspection = self.soi_repository.read(inspection_id)
        if not inspection.checklist_unique_id:
            raise ValueError("SOI checklist generation requires a persisted checklist_unique_id.")

        selected_areas = self.soi_repository.list_selected_areas(inspection.id)
        checklist_items = self.soi_repository.list_checklist_items_for_areas(
            area_ids=[int(area["area_id"]) for area in selected_areas],
        )
        try:
            checklist_version = self.checklist_version_resolver.get_version_for_inspection(inspection)
        except ChecklistVersionResolutionError:
            checklist_version = None

        normalized_format = self._normalize_output_format(output_format)
        if normalized_format == "PDF":
            content = self._build_pdf(
                inspection=inspection,
                selected_areas=selected_areas,
                checklist_items=checklist_items,
                checklist_version=checklist_version,
            )
            content_type = "application/pdf"
            extension = "pdf"
        else:
            content = self._build_workbook(
                inspection=inspection,
                selected_areas=selected_areas,
                checklist_items=checklist_items,
                checklist_version=checklist_version,
            )
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            extension = "xlsx"

        return SOIChecklistRenderResult(
            content=content,
            content_type=content_type,
            file_name=self._build_file_name(
                inspection_reference=str(inspection.inspection_reference),
                checklist_unique_id=str(inspection.checklist_unique_id),
                extension=extension,
            ),
            output_format=normalized_format,
        )

    def _build_pdf(
        self,
        *,
        inspection,
        selected_areas: list[dict[str, object]],
        checklist_items: list[dict[str, object]],
        checklist_version,
    ) -> bytes:
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4, invariant=1)
        width, height = A4
        rows = self._build_pdf_rows(selected_areas=selected_areas, checklist_items=checklist_items)
        rows_per_page = 18
        chunks = [rows[index : index + rows_per_page] for index in range(0, len(rows), rows_per_page)] or [[]]
        total_pages = len(chunks)
        code_image = self._build_code_image(str(inspection.checklist_unique_id))

        for page_number, chunk in enumerate(chunks, start=1):
            self._draw_pdf_header(
                pdf,
                width=width,
                height=height,
                inspection=inspection,
                checklist_version=checklist_version,
                code_image=code_image,
            )
            self._draw_pdf_item_rows(pdf, rows=chunk, origin_y=height - 205)
            self._draw_pdf_signature_block(pdf, footer_y=110)
            self._draw_pdf_footer(
                pdf,
                width=width,
                checklist_unique_id=str(inspection.checklist_unique_id),
                page_number=page_number,
                total_pages=total_pages,
            )
            pdf.showPage()

        pdf.save()
        return buffer.getvalue()

    def _build_workbook(
        self,
        *,
        inspection,
        selected_areas: list[dict[str, object]],
        checklist_items: list[dict[str, object]],
        checklist_version,
    ) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "SOI Checklist"

        workbook.properties.creator = "Safety Module"
        workbook.properties.created = datetime(2026, 4, 17, 0, 0, 0)
        workbook.properties.modified = datetime(2026, 4, 17, 0, 0, 0)

        worksheet["A1"] = "Safety Officer Inspection Checklist"
        worksheet["A1"].font = Font(size=16, bold=True)
        worksheet["A2"] = "Inspection reference"
        worksheet["B2"] = inspection.inspection_reference
        worksheet["A3"] = "Checklist unique ID"
        worksheet["B3"] = inspection.checklist_unique_id
        worksheet["A4"] = "Cycle"
        worksheet["B4"] = inspection.cycle_label
        worksheet["A5"] = "Planned date"
        worksheet["B5"] = inspection.planned_date.isoformat()
        worksheet["A6"] = "Checklist version"
        worksheet["B6"] = checklist_version.version_label if checklist_version is not None else "Unresolved"
        worksheet["A7"] = "Paper-first note"
        worksheet["B7"] = "File paper in ship SMS filing system. No scan upload in VIMS."

        header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
        header_font = Font(bold=True)
        header_row = 9
        headers = ["Area", "Subsection", "Item", "Checklist requirement", "Yes/No/NA", "Notes"]
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        rows = self._build_workbook_rows(selected_areas=selected_areas, checklist_items=checklist_items)
        for row_index, row in enumerate(rows, start=header_row + 1):
            for column_index, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        signature_row = max(header_row + len(rows) + 3, 15)
        worksheet.cell(row=signature_row, column=1, value="Safety Officer signature:")
        worksheet.cell(row=signature_row + 1, column=1, value="Assistant signature:")
        worksheet.cell(row=signature_row + 2, column=1, value="Filed in ship SMS filing system:")

        for column_index, width in enumerate([32, 24, 16, 72, 16, 28, 18, 18], start=1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        code_image = self._build_code_image(str(inspection.checklist_unique_id))
        if code_image is not None:
            worksheet.add_image(ExcelImage(code_image), "G2")

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _draw_pdf_header(
        self,
        pdf,
        *,
        width: float,
        height: float,
        inspection,
        checklist_version,
        code_image: BytesIO | None,
    ) -> None:
        pdf.setFont("Helvetica-Bold", 17)
        pdf.drawString(40, height - 42, "Safety Officer Inspection Checklist")

        pdf.setFont("Helvetica", 10)
        pdf.drawString(40, height - 64, f"Inspection reference: {inspection.inspection_reference}")
        pdf.drawString(40, height - 80, f"Checklist unique ID: {inspection.checklist_unique_id}")
        pdf.drawString(40, height - 96, f"Cycle: {inspection.cycle_label}")
        pdf.drawString(40, height - 112, f"Planned date: {inspection.planned_date.isoformat()}")
        version_label = checklist_version.version_label if checklist_version is not None else "Unresolved"
        pdf.drawString(40, height - 128, f"Checklist version: {version_label}")
        pdf.drawString(40, height - 146, "Paper-first flow: download, inspect on paper, file in SMS, register findings digitally.")
        pdf.drawString(40, height - 162, "No scan upload exists in VIMS for the paper checklist.")

        if code_image is not None:
            pdf.drawImage(
                ImageReader(code_image),
                width - 150,
                height - 162,
                width=90,
                height=90,
                preserveAspectRatio=True,
                mask="auto",
            )

    def _draw_pdf_item_rows(
        self,
        pdf,
        *,
        rows: list[dict[str, object]],
        origin_y: float,
    ) -> None:
        if not rows:
            pdf.setFont("Helvetica", 11)
            pdf.drawString(40, origin_y, "No checklist items are currently selected for this checklist.")
            return

        current_y = origin_y
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(40, current_y, "Area")
        pdf.drawString(150, current_y, "Subsection")
        pdf.drawString(280, current_y, "Item")
        pdf.drawString(340, current_y, "Checklist requirement")
        pdf.drawString(515, current_y, "Y/N/NA")
        header_rule_y = current_y - 8
        pdf.line(40, header_rule_y, 555, header_rule_y)
        current_y = header_rule_y - 14
        for row in rows:
            pdf.setFont("Helvetica", 8)
            pdf.drawString(40, current_y, self._truncate(str(row["area"]), 28))
            pdf.drawString(150, current_y, self._truncate(str(row["subsection"]), 34))
            pdf.drawString(280, current_y, self._truncate(str(row["item_number"]), 12))
            description_lines = self._wrap_text(str(row["description"]), width=46, max_lines=2)
            for offset, line in enumerate(description_lines):
                pdf.drawString(340, current_y - (offset * 10), line)
            pdf.rect(520, current_y - 4, 28, 10, stroke=1, fill=0)
            current_y -= 26

    def _draw_pdf_signature_block(self, pdf, *, footer_y: float) -> None:
        pdf.setFont("Helvetica", 10)
        pdf.drawString(40, footer_y, "Safety Officer signature: ______________________________")
        pdf.drawString(40, footer_y - 18, "Assistant signature: __________________________________")
        pdf.drawString(40, footer_y - 36, "Filed in ship SMS filing system: ______________________")

    def _draw_pdf_footer(
        self,
        pdf,
        *,
        width: float,
        checklist_unique_id: str,
        page_number: int,
        total_pages: int,
    ) -> None:
        pdf.setFont("Helvetica", 9)
        pdf.drawString(40, 32, f"Unique checklist ID: {checklist_unique_id}")
        pdf.drawRightString(width - 40, 32, f"Page {page_number} of {total_pages}")

    def _build_code_image(self, checklist_unique_id: str) -> BytesIO | None:
        image = BytesIO()
        qr_image = qrcode.make(checklist_unique_id)
        qr_image.save(image, format="PNG")
        image.seek(0)
        return image

    def _build_pdf_rows(
        self,
        *,
        selected_areas: list[dict[str, object]],
        checklist_items: list[dict[str, object]],
    ) -> list[dict[str, object]]:
        area_names = {int(area["area_id"]): str(area["area_name"]) for area in selected_areas}
        return [
            {
                "area": f"{item['area_id']}: {area_names.get(int(item['area_id']), item['area_name'])}",
                "subsection": f"{item['subsection_id']}: {item['subsection_name']}",
                "item_number": item["item_number"],
                "description": item["description"],
            }
            for item in checklist_items
        ]

    def _build_workbook_rows(
        self,
        *,
        selected_areas: list[dict[str, object]],
        checklist_items: list[dict[str, object]],
    ) -> list[list[object]]:
        area_names = {int(area["area_id"]): str(area["area_name"]) for area in selected_areas}
        return [
            [
                f"Area {item['area_id']}: {area_names.get(int(item['area_id']), item['area_name'])}",
                f"{item['subsection_id']} - {item['subsection_name']}",
                item["item_number"],
                item["description"],
                "",
                "",
            ]
            for item in checklist_items
        ]

    @staticmethod
    def _wrap_text(value: str, *, width: int, max_lines: int) -> list[str]:
        words = value.split()
        lines: list[str] = []
        current = ""
        for word in words:
            candidate = word if not current else f"{current} {word}"
            if len(candidate) <= width:
                current = candidate
                continue
            if current:
                lines.append(current)
            current = word
            if len(lines) == max_lines:
                break
        if current and len(lines) < max_lines:
            lines.append(current)
        if not lines:
            return [""]
        if len(lines) == max_lines and len(" ".join(words)) > len(" ".join(lines)):
            lines[-1] = lines[-1][: max(width - 3, 0)].rstrip() + "..."
        return lines

    @staticmethod
    def _truncate(value: str, length: int) -> str:
        if len(value) <= length:
            return value
        return value[: max(length - 3, 0)].rstrip() + "..."

    def _build_file_name(
        self,
        *,
        inspection_reference: str,
        checklist_unique_id: str,
        extension: str,
    ) -> str:
        safe_reference = re.sub(r"[^A-Za-z0-9_-]+", "-", inspection_reference).strip("-")
        safe_unique_id = re.sub(r"[^A-Za-z0-9_-]+", "-", checklist_unique_id).strip("-")
        return f"{safe_reference}-{safe_unique_id}.{extension}"

    def _normalize_output_format(self, value: str) -> str:
        normalized = str(value).strip().upper()
        if normalized not in {"PDF", "XLSX"}:
            raise ValueError("SOI checklist output format must be PDF or XLSX.")
        return normalized
