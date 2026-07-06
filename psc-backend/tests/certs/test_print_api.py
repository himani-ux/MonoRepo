from __future__ import annotations

from io import BytesIO
import os
import unittest
from unittest.mock import MagicMock, patch
import uuid
from zipfile import ZipFile

import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings_test")
django.setup()

from openpyxl import load_workbook
from PyPDF2 import PdfReader
from rest_framework import status
from rest_framework.test import APIRequestFactory, force_authenticate

from apps.certs.services.excel_renderer import render_print_excel
from apps.certs.services.pdf_renderer import ReportLabPdfRenderer
from apps.certs.services.print_artifacts import (
    PrintGenerationFailed,
    PrintArtifactRepository,
    PrintArtifactService,
    derive_print_id_scope_token,
)
from apps.certs.services.system_state_hash import compute_system_state_hash
from apps.certs.services.zip_bundler import build_share_bundle_zip
from apps.certs.views.print_views import PrintArtifactCreateView, PrintShareBundleView
from tests.certs.test_tracked_item_api import make_user, tracked_item_row


def print_artifact_row(**overrides):
    row = {
        "print_id": "SQE-S633-9876543-20260629-001",
        "scope": "per_vessel_full",
        "vessels_json": '["vessel-1"]',
        "sections_json": "[]",
        "filters_json": "{}",
        "custom_cert_ids_json": "[]",
        "user_id": "dpa-1",
        "user_role": "DPA",
        "timestamp_utc": "2026-06-29T10:00:00Z",
        "system_state_hash": "abc12345",
        "watermark_applied": "INTERNAL",
        "watermark_recipient": "",
        "pdf_blob_id": uuid.uuid4(),
        "excel_blob_id": uuid.uuid4(),
        "bundle_zip_blob_id": None,
        "recipient_email": "",
        "page_count": 2,
        "generation_status": "success",
        "failure_message": "",
    }
    row.update(overrides)
    return row


def printable_row(**overrides):
    vessel_id = uuid.uuid4()
    row = tracked_item_row(vessel_id=vessel_id, pdf_attachment_id=uuid.uuid4())
    row.update(
        {
            "vessel_name": "KSM Fortitude",
            "vessel_code": "KSMF",
            "vessel_imo": "9876543",
            "vessel_flag": "Panama",
            "class_society": "NK",
            "catalog_section_code": "STAT",
            "catalog_section_name": "Statutory & Flag",
            "catalog_print_order": 10,
            "blob_storage_path": "certs/vessels/vessel-1/tracked/iopp.pdf",
            "blob_filename": "IOPP.pdf",
            "blob_content_sha256": "a" * 64,
        }
    )
    row.update(overrides)
    return row


class CertPrintApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.factory = APIRequestFactory()
        self.dpa = make_user(
            role="DPA",
            form_ids=["CERT_F_004"],
            process_ids=["CERT_P_005", "CERT_P_006"],
            has_global_vessel_access=True,
        )

    @patch("apps.certs.views.print_views.service")
    def test_print_requires_print_process_permission(self, service) -> None:
        request = self.factory.post(
            "/api/certs/print/",
            {"scope": "per_vessel_full", "vesselIds": [str(uuid.uuid4())]},
            format="json",
        )
        user = make_user(role="Fleet Manager", form_ids=["CERT_F_004"], process_ids=[])
        force_authenticate(request, user=user)

        response = PrintArtifactCreateView.as_view()(request)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        service.generate_print.assert_not_called()

    @patch("apps.certs.views.print_views.record_audit_event")
    @patch("apps.certs.views.print_views.service")
    def test_generate_per_vessel_print_returns_artifact_and_records_audit(self, service, record_audit_event) -> None:
        vessel_id = str(uuid.uuid4())
        service.generate_print.return_value = print_artifact_row(vessels_json=f'["{vessel_id}"]')
        request = self.factory.post(
            "/api/certs/print/",
            {
                "scope": "per_vessel_full",
                "vesselIds": [vessel_id],
                "watermarkApplied": "INTERNAL",
                "recipientEmail": "master@example.test",
            },
            format="json",
        )
        force_authenticate(request, user=self.dpa)

        response = PrintArtifactCreateView.as_view()(request)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["printId"], "SQE-S633-9876543-20260629-001")
        self.assertEqual(response.data["scope"], "per_vessel_full")
        self.assertEqual(response.data["vessels"], [vessel_id])
        self.assertEqual(response.data["watermarkApplied"], "INTERNAL")
        service.generate_print.assert_called_once()
        record_audit_event.assert_called_once()
        self.assertEqual(record_audit_event.call_args.kwargs["action"], "print_artifact_created")
        self.assertEqual(record_audit_event.call_args.kwargs["entity_type"], "print_artifact")

    @patch("apps.certs.views.print_views.record_audit_event")
    @patch("apps.certs.views.print_views.service")
    def test_share_bundle_requires_export_bundle_and_returns_zip_artifact(self, service, record_audit_event) -> None:
        vessel_id = str(uuid.uuid4())
        service.generate_share_bundle.return_value = print_artifact_row(
            scope="share_bundle",
            watermark_applied="MASTER_COPY",
            watermark_recipient="Port State Inspector",
            bundle_zip_blob_id=uuid.uuid4(),
            vessels_json=f'["{vessel_id}"]',
        )
        request = self.factory.post(
            "/api/certs/print/share-bundle/",
            {
                "vesselIds": [vessel_id],
                "customCertIds": [str(uuid.uuid4())],
                "watermarkRecipient": "Port State Inspector",
            },
            format="json",
        )
        force_authenticate(request, user=self.dpa)

        response = PrintShareBundleView.as_view()(request)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["scope"], "share_bundle")
        self.assertEqual(response.data["watermarkApplied"], "MASTER_COPY")
        self.assertIsNotNone(response.data["bundleZipBlobId"])
        service.generate_share_bundle.assert_called_once()
        self.assertEqual(record_audit_event.call_args.kwargs["action"], "share_bundle_created")

    def test_renderers_preserve_sqe_s633_pdf_excel_and_zip_manifest(self) -> None:
        row = printable_row()
        payload = {
            "scope": "per_vessel_full",
            "vesselIds": [str(row["vessel_id"])],
            "sections": [],
            "filters": {},
            "customCertIds": [],
            "watermarkApplied": "INTERNAL",
            "watermarkRecipient": "",
        }
        state_hash = compute_system_state_hash([row], payload)
        renderer = ReportLabPdfRenderer()

        pdf_result = renderer.render_print_artifact(
            print_id="SQE-S633-9876543-20260629-001",
            rows=[row],
            payload=payload,
            actor_id="dpa-1",
            actor_role="DPA",
            system_state_hash=state_hash,
        )
        excel_bytes = render_print_excel(
            print_id="SQE-S633-9876543-20260629-001",
            rows=[row],
            payload=payload,
            system_state_hash=state_hash,
        )
        manifest_pdf = renderer.render_share_bundle_manifest(
            print_id="SQE-S633-9876543-20260629-001",
            rows=[row],
            payload={**payload, "watermarkApplied": "MASTER_COPY", "watermarkRecipient": "Port State Inspector"},
            actor_id="master-1",
            actor_role="MASTER",
            system_state_hash=state_hash,
        ).content
        zip_bytes = build_share_bundle_zip(
            print_id="SQE-S633-9876543-20260629-001",
            rows=[row],
            manifest_pdf=manifest_pdf,
            read_blob=lambda blob_path: b"%PDF-1.4\ncertificate\n%%EOF",
        )

        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf_result.content)).pages)
        self.assertIn("SQE S 633", text)
        self.assertIn("INTERNAL", text)
        self.assertIn(state_hash, text)

        workbook = load_workbook(BytesIO(excel_bytes), read_only=True)
        values = [cell for worksheet in workbook.worksheets for row_values in worksheet.iter_rows(values_only=True) for cell in row_values]
        self.assertIn("SQE S 633", values)
        self.assertIn("SQE-S633-9876543-20260629-001", values)

        with ZipFile(BytesIO(zip_bytes)) as bundle:
            names = set(bundle.namelist())
            self.assertIn("manifest_SQE-S633-9876543-20260629-001.pdf", names)
            self.assertIn("certificates/01_IOPP.pdf", names)

    @patch("apps.certs.services.print_artifacts.connection")
    def test_repository_print_id_derivation_uses_s633_imo_date_sequence(self, connection) -> None:
        cursor = MagicMock()
        cursor.fetchone.return_value = (2,)
        connection.cursor.return_value.__enter__.return_value = cursor

        print_id = PrintArtifactRepository().next_print_id("9876543", now_yyyymmdd="20260629")

        self.assertEqual(print_id, "SQE-S633-9876543-20260629-003")
        self.assertIn("COUNT(*)", cursor.execute.call_args.args[0])
        self.assertIn("dbo.vims_certs_print_artifact", cursor.execute.call_args.args[0])

    def test_print_id_scope_token_uses_fleet_for_empty_or_multi_vessel_scopes(self) -> None:
        one_vessel = [printable_row(vessel_imo="9876543")]
        multi_vessel = [
            printable_row(vessel_imo="9876543"),
            printable_row(vessel_imo="7654321"),
        ]

        self.assertEqual(derive_print_id_scope_token(one_vessel), "9876543")
        self.assertEqual(derive_print_id_scope_token(multi_vessel), "FLEET")
        self.assertEqual(derive_print_id_scope_token([]), "FLEET")

    @patch("apps.certs.services.print_artifacts.record_audit_event")
    def test_service_records_high_volume_print_activity_without_blocking_print(self, record_audit_event) -> None:
        repository = MagicMock()
        blob_repository = MagicMock()
        repository.list_rows_for_scope.return_value = [printable_row()]
        repository.next_print_id.return_value = "SQE-S633-9876543-20260629-011"
        repository.insert_artifact.return_value = print_artifact_row(print_id="SQE-S633-9876543-20260629-011")
        repository.count_user_prints_since.return_value = 11
        blob_repository.create_artifact_blob.side_effect = [
            {"blob_id": uuid.uuid4()},
            {"blob_id": uuid.uuid4()},
        ]
        service = PrintArtifactService(
            repository=repository,
            blob_repository=blob_repository,
            save_artifact=lambda **kwargs: {"relative_path": kwargs["filename"], "filename": kwargs["filename"], "sha256": "a" * 64, "size": 123},
        )

        row = service.generate_print(
            payload={
                "scope": "per_vessel_full",
                "vesselIds": [str(uuid.uuid4())],
                "sections": [],
                "filters": {},
                "customCertIds": [],
                "watermarkApplied": "INTERNAL",
                "watermarkRecipient": "",
                "recipientEmail": "",
            },
            actor=self.dpa,
        )

        self.assertEqual(row["print_id"], "SQE-S633-9876543-20260629-011")
        repository.count_user_prints_since.assert_called_once()
        record_audit_event.assert_called_once()
        self.assertEqual(record_audit_event.call_args.kwargs["action"], "high_volume_print_activity")
        self.assertEqual(record_audit_event.call_args.kwargs["entity_type"], "print_artifact")
        self.assertEqual(record_audit_event.call_args.kwargs["entity_id"], "SQE-S633-9876543-20260629-011")
        self.assertEqual(record_audit_event.call_args.kwargs["metadata"]["printCountLastHour"], 11)
        self.assertEqual(record_audit_event.call_args.kwargs["metadata"]["thresholdPerHour"], 10)

    @patch("apps.certs.services.print_artifacts.record_audit_event")
    def test_service_hard_fails_print_with_support_ticket_and_no_auto_retry(self, record_audit_event) -> None:
        repository = MagicMock()
        blob_repository = MagicMock()
        renderer = MagicMock()
        repository.list_rows_for_scope.return_value = [printable_row()]
        repository.next_print_id.return_value = "SQE-S633-9876543-20260629-012"
        repository.insert_artifact.side_effect = lambda values: print_artifact_row(**values)
        renderer.render_print_artifact.side_effect = RuntimeError("ReportLab layout failed")
        service = PrintArtifactService(
            repository=repository,
            blob_repository=blob_repository,
            renderer=renderer,
            save_artifact=lambda **kwargs: {},
        )

        with self.assertRaises(PrintGenerationFailed) as exc:
            service.generate_print(
                payload={
                    "scope": "per_vessel_full",
                    "vesselIds": [str(uuid.uuid4())],
                    "sections": [],
                    "filters": {},
                    "customCertIds": [],
                    "watermarkApplied": "INTERNAL",
                    "watermarkRecipient": "",
                    "recipientEmail": "",
                },
                actor=self.dpa,
            )

        artifact = exc.exception.artifact
        self.assertEqual(artifact["print_id"], "SQE-S633-9876543-20260629-012")
        self.assertEqual(artifact["generation_status"], "failed")
        self.assertIn("SQE-S633-9876543-20260629-012-ERR", artifact["failure_message"])
        self.assertIn("Support ticket", artifact["failure_message"])
        renderer.render_print_artifact.assert_called_once()
        blob_repository.create_artifact_blob.assert_not_called()
        repository.count_user_prints_since.assert_not_called()
        record_audit_event.assert_called_once()
        self.assertEqual(record_audit_event.call_args.kwargs["action"], "print_generation_failed")
        self.assertEqual(record_audit_event.call_args.kwargs["entity_id"], "SQE-S633-9876543-20260629-012")
        self.assertEqual(record_audit_event.call_args.kwargs["metadata"]["supportTicketReference"], "SQE-S633-9876543-20260629-012-ERR")
        self.assertIn("RuntimeError", record_audit_event.call_args.kwargs["metadata"]["stackTrace"])

    @patch("apps.certs.views.print_views.record_audit_event")
    @patch("apps.certs.views.print_views.service")
    def test_print_view_returns_failed_artifact_ticket_without_success_audit(self, service, record_audit_event) -> None:
        failed_artifact = print_artifact_row(
            print_id="SQE-S633-9876543-20260629-013",
            generation_status="failed",
            failure_message="Generation failed. Support ticket SQE-S633-9876543-20260629-013-ERR was logged. Retry manually.",
            pdf_blob_id=None,
            excel_blob_id=None,
            page_count=0,
        )
        service.generate_print.side_effect = PrintGenerationFailed(failed_artifact)
        request = self.factory.post(
            "/api/certs/print/",
            {"scope": "per_vessel_full", "vesselIds": [str(uuid.uuid4())]},
            format="json",
        )
        force_authenticate(request, user=self.dpa)

        response = PrintArtifactCreateView.as_view()(request)

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertEqual(response.data["artifact"]["printId"], "SQE-S633-9876543-20260629-013")
        self.assertEqual(response.data["artifact"]["generationStatus"], "failed")
        self.assertIn("SQE-S633-9876543-20260629-013-ERR", response.data["detail"])
        record_audit_event.assert_not_called()

    @patch("apps.certs.services.print_artifacts.connection")
    def test_repository_custom_selection_keeps_vessel_scope_filter(self, connection) -> None:
        vessel_id = str(uuid.uuid4())
        tracked_id = str(uuid.uuid4())
        cursor = MagicMock()
        cursor.description = [("tracked_item_id",)]
        cursor.fetchall.return_value = []
        connection.cursor.return_value.__enter__.return_value = cursor

        PrintArtifactRepository().list_rows_for_scope(
            {
                "scope": "custom_selection",
                "vesselIds": [vessel_id],
                "sections": [],
                "filters": {},
                "customCertIds": [tracked_id],
            }
        )

        sql = cursor.execute.call_args.args[0]
        params = cursor.execute.call_args.args[1]
        self.assertIn("t.tracked_item_id IN (%s)", sql)
        self.assertIn("t.vessel_id IN (%s)", sql)
        self.assertEqual(params, [tracked_id, vessel_id])

    def test_service_rejects_share_bundle_rows_without_pdf_blobs(self) -> None:
        repository = MagicMock()
        blob_repository = MagicMock()
        repository.list_rows_for_scope.return_value = [printable_row(pdf_attachment_id=None, blob_storage_path=None)]
        service = PrintArtifactService(
            repository=repository,
            blob_repository=blob_repository,
            save_artifact=lambda **kwargs: {},
            read_blob=lambda path: b"",
        )

        with self.assertRaises(ValueError) as exc:
            service.generate_share_bundle(
                payload={
                    "scope": "share_bundle",
                    "vesselIds": [str(uuid.uuid4())],
                    "sections": [],
                    "filters": {},
                    "customCertIds": [str(uuid.uuid4())],
                    "watermarkApplied": "MASTER_COPY",
                    "watermarkRecipient": "Auditor",
                    "recipientEmail": "",
                },
                actor=self.dpa,
            )

        self.assertIn("attached certificate PDF", str(exc.exception))


if __name__ == "__main__":
    unittest.main()
