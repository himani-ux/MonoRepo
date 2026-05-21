from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from types import SimpleNamespace
import unittest

from openpyxl import load_workbook
from PyPDF2 import PdfReader
from django.utils import timezone

from tests.safety.support import (
    bootstrap_django,
    recreate_incident_table,
    recreate_phase5_reference_tables,
    recreate_scm_tables,
    recreate_soi_tables,
)


bootstrap_django(root_urlconf="config.urls")

from django.db import connection
from rest_framework.test import APIRequestFactory, force_authenticate

from apps.safety.models import CorrectiveAction, Incident, Recommendation, SOIInspection
from apps.safety.views.dashboard_export import DashboardExportView


def aware(year: int, month: int, day: int, hour: int = 0, minute: int = 0):
    return timezone.make_aware(datetime(year, month, day, hour, minute))


def build_user(*, role_name: str, process_ids: list[str]) -> SimpleNamespace:
    return SimpleNamespace(
        id=f"{role_name.lower()}-1",
        username=f"{role_name.lower()}-1",
        role_name=role_name,
        form_ids=["SAF_F_015"],
        process_ids=process_ids,
        vessel_ids=["7"],
        is_global=False,
    )


class DashboardExportViewTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        super().setUpClass()
        bootstrap_django(root_urlconf="config.urls")

    def setUp(self) -> None:
        recreate_incident_table()
        recreate_phase5_reference_tables()
        recreate_soi_tables()
        recreate_scm_tables()
        self.current_at = aware(2026, 4, 30, 12, 0)
        self.factory = APIRequestFactory()
        self.view = DashboardExportView.as_view()

        incident = Incident.objects.create(
            incident_number="INC/2026/EXP-1",
            vessel_id="7",
            record_type=Incident.RecordType.INCIDENT,
            state="UNDER_REVIEW",
            current_phase=6,
            risk_band=Incident.RiskBand.YELLOW,
            created_by="dpa-1",
            updated_by="dpa-1",
            schema_version=1,
        )
        recommendation = Recommendation.objects.create(
            incident=incident,
            tier=Recommendation.Tier.CORRECTIVE,
            title="Replace failed guard",
            description="Corrective action export fixture.",
            created_by="dpa-1",
            updated_by="dpa-1",
            schema_version=1,
        )
        action = CorrectiveAction.objects.create(
            source_table="vims_safety_incident",
            source_id=incident.pk,
            recommendation=recommendation,
            title="Overdue corrective action",
            description="Waiting on execution.",
            due_date=(self.current_at - timedelta(days=2)).date(),
            status=CorrectiveAction.Status.OPEN,
            created_by="dpa-1",
            updated_by="dpa-1",
            schema_version=1,
        )
        CorrectiveAction.objects.filter(pk=action.pk).update(created_date=self.current_at - timedelta(days=22))

        SOIInspection.objects.create(
            vessel_id="7",
            inspection_reference="SOI/EXP/7/001",
            cycle_label="Q2/2026",
            state=SOIInspection.State.REPORTED,
            planned_date=self.current_at.date(),
            safety_officer_crew_id="co-7",
            safety_officer_department="DECK",
            assistant_crew_id="2e-7",
            assistant_department="ENGINE",
            created_by="co-7",
            updated_by="co-7",
            schema_version=1,
        )

        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO master_soi_area (
                    area_id,
                    area_name,
                    section_12_flag,
                    display_order,
                    active,
                    seeded_version
                ) VALUES (%s, %s, %s, %s, %s, %s)
                """,
                [3, "Navigating Bridge & Monkey Island", False, 3, True, "v1.0"],
            )
            cursor.execute(
                """
                INSERT INTO vims_safety_soi_vessel_area_map (
                    vessel_id,
                    area_id,
                    applicable,
                    last_inspected_at,
                    due_at,
                    schema_version
                ) VALUES (%s, %s, %s, %s, %s, %s)
                """,
                [
                    "7",
                    3,
                    True,
                    self.current_at - timedelta(days=20),
                    self.current_at + timedelta(days=70),
                    1,
                ],
            )

    def test_excel_export_is_dpa_only_and_contains_metadata(self) -> None:
        request = self.factory.post(
            "/api/safety/dashboard/export/",
            {"format": "excel", "period": "12M", "vessel_id": "7"},
            format="json",
        )
        force_authenticate(request, user=build_user(role_name="DPA", process_ids=["SAF_P_023"]))

        response = self.view(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        metadata = workbook["Metadata"]
        self.assertEqual(metadata["A1"].value, "Safety Intelligence Dashboard Export")
        self.assertEqual(metadata["B3"].value, "dpa-1")
        self.assertEqual(metadata["B4"].value, "12M")
        self.assertEqual(metadata["B5"].value, "VESSEL")
        self.assertEqual(str(metadata["B6"].value), "7")

    def test_xlsx_alias_returns_excel_workbook(self) -> None:
        request = self.factory.post(
            "/api/safety/dashboard/export/",
            {"format": "xlsx", "period": "12M", "vessel_id": "7"},
            format="json",
        )
        force_authenticate(request, user=build_user(role_name="DPA", process_ids=["SAF_P_023"]))

        response = self.view(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertIn("Metadata", workbook.sheetnames)

    def test_pdf_export_rejects_non_dpa_even_with_export_process_permission(self) -> None:
        request = self.factory.post(
            "/api/safety/dashboard/export/",
            {"format": "pdf", "period": "3Y", "vessel_id": "7"},
            format="json",
        )
        force_authenticate(request, user=build_user(role_name="FM", process_ids=["SAF_P_023"]))

        response = self.view(request)

        self.assertEqual(response.status_code, 403)

    def test_pdf_export_returns_pdf_payload_for_dpa(self) -> None:
        request = self.factory.post(
            "/api/safety/dashboard/export/",
            {"format": "pdf", "period": "3Y", "vessel_id": "7"},
            format="json",
        )
        force_authenticate(request, user=build_user(role_name="DPA", process_ids=["SAF_P_023"]))

        response = self.view(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        reader = PdfReader(BytesIO(response.content))
        self.assertGreaterEqual(len(reader.pages), 1)
        self.assertIn(
            "Safety Intelligence Dashboard Export",
            reader.pages[0].extract_text(),
        )
